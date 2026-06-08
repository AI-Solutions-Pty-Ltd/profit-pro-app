import openpyxl

wb_val = openpyxl.load_workbook(
    "LEPHADIMISHA SS - OUTSTANDING WORKS BOQ PC2 - 15 May 2026.xlsx", data_only=True
)
wb_form = openpyxl.load_workbook(
    "LEPHADIMISHA SS - OUTSTANDING WORKS BOQ PC2 - 15 May 2026.xlsx", data_only=False
)

sh_val = wb_val["Summary"]
sh_form = wb_form["Summary"]

for r in range(1, sh_val.max_row + 1):
    row_val = [
        sh_val.cell(row=r, column=c).value for c in range(1, sh_val.max_column + 1)
    ]
    row_form = [
        sh_form.cell(row=r, column=c).value for c in range(1, sh_form.max_column + 1)
    ]
    if any(v is not None for v in row_val):
        print(f"Row {r:02d}: {row_val}  | Formulas: {row_form}")
