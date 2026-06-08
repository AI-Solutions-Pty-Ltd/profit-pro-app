import openpyxl

wb_val = openpyxl.load_workbook(
    "LEPHADIMISHA SS - OUTSTANDING WORKS BOQ PC2 - 15 May 2026.xlsx", data_only=True
)
sh_val_summary = wb_val["Summary"]
sh_val_pg = wb_val["Preliminary and General"]

print("Summary Sheet Row 8 values:")
for col in range(1, 8):
    print(
        f"  Col {col} ({openpyxl.utils.get_column_letter(col)}): {sh_val_summary.cell(row=8, column=col).value}"
    )

print("\nPreliminary and General Sheet Row 50 values:")
for col in range(1, 15):
    print(
        f"  Col {col} ({openpyxl.utils.get_column_letter(col)}): {sh_val_pg.cell(row=50, column=col).value}"
    )
