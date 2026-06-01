from pathlib import Path

import openpyxl

template_path = Path("app/BillOfQuantities/data/Project set-up Template.xlsx")
wb = openpyxl.load_workbook(template_path)
print("Sheet names:", wb.sheetnames)
ws = wb.active
print("Title of active sheet:", ws.title)
print("Dimensions:", ws.dimensions)

for row in range(1, 5):
    row_vals = [ws.cell(row=row, column=col).value for col in range(1, 15)]
    print(f"Row {row}:", row_vals)
