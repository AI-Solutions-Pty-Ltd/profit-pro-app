import openpyxl

# Load workbook with data_only=True to get evaluated values
wb_val = openpyxl.load_workbook(
    "LEPHADIMISHA SS - OUTSTANDING WORKS BOQ PC2 - 15 May 2026.xlsx", data_only=True
)
# Load workbook with data_only=False to get formulas
wb_form = openpyxl.load_workbook(
    "LEPHADIMISHA SS - OUTSTANDING WORKS BOQ PC2 - 15 May 2026.xlsx", data_only=False
)


def print_sheet_comparison(sheet_name, max_r=75):
    print(f"\n==================== SHEET: {sheet_name} ====================")
    sh_val = wb_val[sheet_name]
    sh_form = wb_form[sheet_name]

    for r in range(1, min(max_r + 1, sh_val.max_row + 1)):
        row_val = [
            sh_val.cell(row=r, column=c).value for c in range(1, sh_val.max_column + 1)
        ]
        row_form = [
            sh_form.cell(row=r, column=c).value
            for c in range(1, sh_form.max_column + 1)
        ]

        # Check if row is not empty
        if any(v is not None for v in row_val):
            # Print row nicely
            print(f"Row {r:02d}:")
            for c in range(len(row_val)):
                v = row_val[c]
                f = row_form[c]
                if v is not None or f is not None:
                    col_letter = openpyxl.utils.get_column_letter(c + 1)
                    if v == f or f is None:
                        print(f"  {col_letter}: {v}")
                    else:
                        print(f"  {col_letter}: {v}  [Formula: {f}]")


print_sheet_comparison("Front", 80)
print_sheet_comparison("Summary", 80)
