from pathlib import Path

import openpyxl
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation

template_path = Path("app/BillOfQuantities/data/Project set-up Template.xlsx")
wb = openpyxl.load_workbook(template_path)
ws = wb["Setup Template"]

# Enable sheet protection
ws.protection.sheet = True
ws.protection.password = "profitpro123"  # Standard default protection password

# Lock all cells by default, then unlock input areas
# Header row (row 1) and Contract Amount (column J) should remain locked
for row in range(2, 1001):
    # Unlock Columns A through I (columns 1 to 9)
    for col in range(1, 10):
        cell = ws.cell(row=row, column=col)
        cell.protection = Protection(locked=False)

    # Lock Column J (column 10) and set formula
    amount_cell = ws.cell(row=row, column=10)
    amount_cell.protection = Protection(locked=True)
    amount_cell.value = f"=H{row}*I{row}"

# Add Data Validation for Unit Column (Column G / 7)
unit_validation = DataValidation(
    type="list",
    formula1='"m,m²,m³,kg,t,no,hr,sum,item,l"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Unit",
    error="Please select a valid unit from the dropdown list.",
)
ws.add_data_validation(unit_validation)
unit_validation.add("G2:G1000")

# Add Decimal Validation for Contract Quantity (Column H / 8)
qty_validation = DataValidation(
    type="decimal",
    operator="greaterThanOrEqual",
    formula1="0",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Quantity",
    error="Please enter a valid positive number for Quantity.",
)
ws.add_data_validation(qty_validation)
qty_validation.add("H2:H1000")

# Add Decimal Validation for Contract Rate (Column I / 9)
rate_validation = DataValidation(
    type="decimal",
    operator="greaterThanOrEqual",
    formula1="0",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Rate",
    error="Please enter a valid positive number for Rate.",
)
ws.add_data_validation(rate_validation)
rate_validation.add("I2:I1000")

wb.save(template_path)
print("Spreadsheet guardrails applied successfully!")
