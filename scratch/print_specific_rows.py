import openpyxl

wb_val = openpyxl.load_workbook(
    "LEPHADIMISHA SS - OUTSTANDING WORKS BOQ PC2 - 15 May 2026.xlsx", data_only=True
)
wb_form = openpyxl.load_workbook(
    "LEPHADIMISHA SS - OUTSTANDING WORKS BOQ PC2 - 15 May 2026.xlsx", data_only=False
)


def print_row(sheet_name, row_num):
    sh_val = wb_val[sheet_name]
    sh_form = wb_form[sheet_name]
    row_val = [
        sh_val.cell(row=row_num, column=c).value
        for c in range(1, sh_val.max_column + 1)
    ]
    row_form = [
        sh_form.cell(row=row_num, column=c).value
        for c in range(1, sh_form.max_column + 1)
    ]
    print(f"{sheet_name} Row {row_num}: {row_val}")
    print(f"  Formulas: {row_form}")


print_row("Preliminary and General", 50)
print_row("Preliminary and General", 100)
print_row("Media Centre", 57)
print_row("Media Centre", 118)
