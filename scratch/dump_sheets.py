import openpyxl

wb = openpyxl.load_workbook(
    "LEPHADIMISHA SS - OUTSTANDING WORKS BOQ PC2 - 15 May 2026.xlsx", data_only=False
)

with open("scratch/excel_structure.txt", "w", encoding="utf-8") as f:
    f.write("SHEETS IN WORKBOOK:\n")
    f.write(", ".join(wb.sheetnames) + "\n\n")

    # Inspect Front sheet in full
    if "Front" in wb.sheetnames:
        f.write("=== FRONT SHEET ===\n")
        sheet = wb["Front"]
        for r in range(1, sheet.max_row + 1):
            row_vals = [
                sheet.cell(row=r, column=c).value
                for c in range(1, sheet.max_column + 1)
            ]
            # filter out entirely empty rows
            if any(val is not None for val in row_vals):
                f.write(f"Row {r:03d}: {row_vals}\n")
        f.write("\n")

    # Inspect Summary sheet
    if "Summary" in wb.sheetnames:
        f.write("=== SUMMARY SHEET ===\n")
        sheet = wb["Summary"]
        for r in range(1, sheet.max_row + 1):
            row_vals = [
                sheet.cell(row=r, column=c).value
                for c in range(1, sheet.max_column + 1)
            ]
            if any(val is not None for val in row_vals):
                f.write(f"Row {r:03d}: {row_vals}\n")
        f.write("\n")

    # Inspect Media Centre sheet (first 100 rows to understand detail layout)
    if "Media Centre" in wb.sheetnames:
        f.write("=== MEDIA CENTRE SHEET (FIRST 120 ROWS) ===\n")
        sheet = wb["Media Centre"]
        for r in range(1, min(120, sheet.max_row + 1)):
            row_vals = [
                sheet.cell(row=r, column=c).value
                for c in range(1, sheet.max_column + 1)
            ]
            if any(val is not None for val in row_vals):
                f.write(f"Row {r:03d}: {row_vals}\n")
        f.write("\n")

print("Dumped structure to scratch/excel_structure.txt")
