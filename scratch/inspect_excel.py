import os

import openpyxl

file_path = "LEPHADIMISHA SS - OUTSTANDING WORKS BOQ PC2 - 15 May 2026.xlsx"
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    # Let's list files starting with LEPHADIMISHA
    for f in os.listdir("."):
        if "LEPHADIMISHA" in f:
            print(f"Found match: {f}")
else:
    print(f"File found! Size: {os.path.getsize(file_path)} bytes")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    print("Sheets in workbook:", wb.sheetnames)

    # Inspect Front sheet
    if "Front" in wb.sheetnames:
        print("\n--- Front Sheet Details ---")
        sheet = wb["Front"]
        for r in range(1, 40):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
            if any(row_vals):
                print(f"Row {r:02d}: {row_vals}")

    # Inspect Summary sheet
    if "Summary" in wb.sheetnames:
        print("\n--- Summary Sheet Details ---")
        sheet = wb["Summary"]
        for r in range(1, 40):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
            if any(row_vals):
                print(f"Row {r:02d}: {row_vals}")

    # Inspect Media Centre sheet
    if "Media Centre" in wb.sheetnames:
        print("\n--- Media Centre Sheet Details ---")
        sheet = wb["Media Centre"]
        for r in range(1, 40):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
            if any(row_vals):
                print(f"Row {r:02d}: {row_vals}")
