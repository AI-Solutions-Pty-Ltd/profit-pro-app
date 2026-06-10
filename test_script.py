import openpyxl
wb = openpyxl.load_workbook('test_export.xlsx')
print('Sheetnames:', wb.sheetnames)
ws = wb.active
for row in ws.iter_rows(min_row=1, max_row=10):
    print([c.value for c in row[:2]])
