from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_group_summary_report_to_xlsx(group_name, province_reports):
    """
    Export the Valuation Summary for a project group to XLSX format,
    organizing projects by province with sub-totals and grand totals.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuation Summary"

    # Styles
    font_bold = Font(bold=True)
    font_bold_white = Font(bold=True, color="FFFFFFFF")
    font_subtitle = Font(italic=True, color="FF666666")
    font_title = Font(bold=True, size=14)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    border_bottom_thick = Border(bottom=Side(style="medium"))
    border_bottom_light = Border(bottom=Side(style="thin", color="FFE5E5E5"))

    fill_header = PatternFill(
        start_color="FF1E3A5F", end_color="FF1E3A5F", fill_type="solid"
    )
    fill_province_header = PatternFill(
        start_color="FFEAEAEA", end_color="FFEAEAEA", fill_type="solid"
    )
    fill_subtotal = PatternFill(
        start_color="FFF5F5F5", end_color="FFF5F5F5", fill_type="solid"
    )
    fill_footer = PatternFill(
        start_color="FFC5922E", end_color="FFC5922E", fill_type="solid"
    )

    # Header Row
    ws.cell(row=1, column=1, value="[ LOGO ]").font = font_bold
    title_cell = ws.cell(
        row=1, column=2, value=f"VALUATION SUMMARY — {group_name.upper()}"
    )
    title_cell.font = font_title
    title_cell.alignment = align_center
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)

    ws.row_dimensions[1].height = 40

    # Subtitle Row
    sub_cell = ws.cell(row=2, column=1, value="Summary Grouped by Province")
    sub_cell.font = font_subtitle
    sub_cell.alignment = align_center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    # Column Headers
    headers = [
        "PROJECT NAME",
        "STATUS",
        "TENDER AMOUNT (R)",
        "APPROVED VARIATIONS (R)",
        "REVISED CONTRACT VALUE (R)",
        "PREVIOUS CERTIFIED (R)",
        "PROGRESSIVE TO DATE (R)",
        "NET CLAIMED (R)",
        "FORECAST AT COMPLETION (R)",
        "VARIANCE (R)",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = font_bold_white
        cell.fill = fill_header
        cell.alignment = align_center

    ws.row_dimensions[4].height = 25

    current_row = 5

    # Grand Totals
    t_budget = Decimal("0.00")
    t_variations = Decimal("0.00")
    t_revised = Decimal("0.00")
    t_previous = Decimal("0.00")
    t_progressive = Decimal("0.00")
    t_net = Decimal("0.00")
    t_forecast = Decimal("0.00")
    t_variance = Decimal("0.00")

    for rep in province_reports:
        # Write Province Header Row
        ws.row_dimensions[current_row].height = 22
        prov_header_cell = ws.cell(
            row=current_row, column=1, value=f"Province: {rep['province_name']}"
        )
        prov_header_cell.font = font_bold
        prov_header_cell.fill = fill_province_header
        for col_idx in range(2, 11):
            ws.cell(row=current_row, column=col_idx).fill = fill_province_header
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=10
        )
        current_row += 1

        # Write Projects
        for p in rep["projects"]:
            project = p["project"]
            ws.cell(row=current_row, column=1, value=project.name)
            ws.cell(row=current_row, column=2, value=project.get_status_display())

            ws.cell(row=current_row, column=3, value=p["contract_value"])
            ws.cell(row=current_row, column=4, value=p["variations"])
            ws.cell(row=current_row, column=5, value=p["revised_contract_value"])
            ws.cell(row=current_row, column=6, value=p["certified_previous"])
            ws.cell(row=current_row, column=7, value=p["certified_amount"])
            ws.cell(row=current_row, column=8, value=p["net_claimed"])
            ws.cell(row=current_row, column=9, value=p["forecast_amount"])
            ws.cell(row=current_row, column=10, value=p["variance"])

            for col_idx in range(3, 11):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.number_format = "R #,##0.00;[Red]-R #,##0.00;R 0.00"
                cell.alignment = align_right

            for col_idx in range(1, 11):
                ws.cell(row=current_row, column=col_idx).border = border_bottom_light

            current_row += 1

        # Write Province Sub-total Row
        ws.row_dimensions[current_row].height = 22
        sub_label = ws.cell(
            row=current_row, column=1, value=f"{rep['province_name']} Sub-total"
        )
        sub_label.font = font_bold
        sub_label.fill = fill_subtotal

        ws.cell(row=current_row, column=2, value="").fill = fill_subtotal

        sub_cells = [
            (3, rep["budget"]),
            (4, rep["variations"]),
            (5, rep["revised"]),
            (6, rep["previous"]),
            (7, rep["cumulative"]),
            (8, rep["net"]),
            (9, rep["forecast"]),
            (10, rep["variance"]),
        ]

        for col_idx, val in sub_cells:
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_bold
            cell.fill = fill_subtotal
            cell.number_format = "R #,##0.00;[Red]-R #,##0.00;R 0.00"
            cell.alignment = align_right

        for col_idx in range(1, 11):
            ws.cell(row=current_row, column=col_idx).border = border_bottom_light

        t_budget += rep["budget"]
        t_variations += rep["variations"]
        t_revised += rep["revised"]
        t_previous += rep["previous"]
        t_progressive += rep["cumulative"]
        t_net += rep["net"]
        t_forecast += rep["forecast"] or Decimal("0.00")
        t_variance += rep["variance"] or Decimal("0.00")

        current_row += 1

    # Grand Totals Row
    ws.row_dimensions[current_row].height = 25
    tot_label = ws.cell(row=current_row, column=1, value="TOTALS")
    tot_label.font = font_bold_white
    tot_label.fill = fill_footer

    ws.cell(row=current_row, column=2, value="").fill = fill_footer

    tot_cells = [
        (3, t_budget),
        (4, t_variations),
        (5, t_revised),
        (6, t_previous),
        (7, t_progressive),
        (8, t_net),
        (9, t_forecast),
        (10, t_variance),
    ]

    for col_idx, val in tot_cells:
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = font_bold_white
        cell.fill = fill_footer
        cell.number_format = "R #,##0.00;[Red]-R #,##0.00;R 0.00"
        cell.alignment = align_right

    for col_idx in range(1, 11):
        ws.cell(row=current_row, column=col_idx).border = border_bottom_thick

    # Signatory spaces
    current_row += 3
    ws.row_dimensions[current_row].height = 20
    cell_p = ws.cell(
        row=current_row, column=1, value="Prepared By (Quantity Surveyor):"
    )
    cell_p.font = font_bold
    cell_r = ws.cell(row=current_row, column=4, value="Reviewed By (Contractor):")
    cell_r.font = font_bold
    cell_a = ws.cell(row=current_row, column=7, value="Approved By (Client):")
    cell_a.font = font_bold

    current_row += 2
    ws.row_dimensions[current_row].height = 18
    ws.cell(
        row=current_row, column=1, value="___________________________"
    ).font = font_bold
    ws.cell(
        row=current_row, column=4, value="___________________________"
    ).font = font_bold
    ws.cell(
        row=current_row, column=7, value="___________________________"
    ).font = font_bold

    current_row += 1
    ws.cell(row=current_row, column=1, value="Signature / Date").font = font_subtitle
    ws.cell(row=current_row, column=4, value="Signature / Date").font = font_subtitle
    ws.cell(row=current_row, column=7, value="Signature / Date").font = font_subtitle

    # Auto fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    return wb
