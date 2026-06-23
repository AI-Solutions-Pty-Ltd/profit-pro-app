import datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_cover_page_to_xlsx(payment_certificate, wb=None):
    """
    Export the cover page report for a payment certificate to XLSX format.
    Mirrors the precise styling, colors, and layout of '01_Front.xlsx'.
    """
    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cover Page"
    else:
        ws = wb.create_sheet(title="Cover Page")

    project = payment_certificate.project
    cover_page_config = project.get_cover_page_config()

    cert_num = str(payment_certificate.certificate_number).zfill(2)
    cert_date = (
        payment_certificate.approved_on.strftime("%d %b %Y")
        if payment_certificate.approved_on
        else datetime.datetime.now().strftime("%d %b %Y")
    )

    # Precise Colors and Styles
    c_black = "FF111111"
    c_grey_text = "FF717171"
    c_white = "FFFFFFFF"

    font_bold = Font(bold=True, color=c_black)
    font_normal = Font(bold=False, color=c_black)
    font_title = Font(bold=True, size=14, color=c_black)
    font_white_bold = Font(bold=True, color=c_white)
    font_less = Font(color=c_grey_text, italic=True)
    font_italic_small = Font(italic=True, size=9, color=c_grey_text)

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    fill_light_grey = PatternFill(
        start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid"
    )
    fill_white = PatternFill(
        start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid"
    )
    fill_dark_blue = PatternFill(
        start_color="FF1B3A6B", end_color="FF1B3A6B", fill_type="solid"
    )
    fill_gold = PatternFill(
        start_color="FFC8963E", end_color="FFC8963E", fill_type="solid"
    )
    fill_separator = PatternFill(
        start_color="FF003300", end_color="FF003300", fill_type="solid"
    )  # Dark green/black line

    border_thin_top_bottom = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    border_thin_bottom = Border(bottom=Side(style="thin"))

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 25

    # Row Heights
    ws.row_dimensions[1].height = 55.5
    ws.row_dimensions[2].height = 13.5
    ws.row_dimensions[3].height = 3.0
    ws.row_dimensions[4].height = 6.0

    # Initialize all cells to white fill to avoid default gridlines
    for row in range(1, 55):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill_white
            ws.cell(row=row, column=col).font = font_normal

    # Row 1: Logo & Title
    ws.cell(row=1, column=1, value="[ LOGO ]").font = font_bold
    ws.cell(row=1, column=1).alignment = align_center

    main_title = cover_page_config.get("title") or "PAYMENT CERTIFICATE"
    title_cell = ws.cell(row=1, column=3, value=main_title)
    title_cell.font = font_title
    title_cell.alignment = align_center
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)

    cert_cell = ws.cell(row=1, column=7, value=f"Cert No. {cert_num}\n{cert_date}")
    cert_cell.alignment = Alignment(
        wrap_text=True, horizontal="right", vertical="center"
    )
    cert_cell.font = Font(color=c_grey_text)

    # Row 1 Borders
    for col in range(1, 8):
        ws.cell(row=1, column=col).border = border_thin_top_bottom
        ws.cell(row=1, column=col).fill = fill_light_grey

    # Row 2: Project Info
    project_info = f"{project.name} - Contract No. {project.contract_number or ''}"
    ws.cell(row=2, column=1, value=project_info).font = Font(
        italic=True, color=c_grey_text
    )
    ws.cell(row=2, column=1).alignment = align_center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    # Row 2 Borders
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border_thin_bottom
        ws.cell(row=2, column=col).fill = fill_light_grey

    # Row 3: Thick separator line
    for col in range(1, 8):
        ws.cell(row=3, column=col).fill = fill_separator

    # Dynamic Info Rows starting at Row 5
    orig_val = project.original_contract_value or Decimal("0.00")
    amends_val = project.addendum_contract_value or Decimal("0.00")
    sub_total_contract = orig_val + amends_val
    vat_val_contract = (
        sub_total_contract * Decimal("0.15") if project.vat else Decimal("0.00")
    )
    total_contract = sub_total_contract + vat_val_contract

    work_progressive_previous = (
        payment_certificate.work_progressive_previous or Decimal("0.00")
    )
    contract_current_claim_total = (
        payment_certificate.contract_current_claim_total or Decimal("0.00")
    )
    addendum_current_claim_total = (
        payment_certificate.addendum_current_claim_total or Decimal("0.00")
    )
    work_progressive_to_date = payment_certificate.work_progressive_to_date or Decimal(
        "0.00"
    )
    ap_current = payment_certificate.get_advance_payment_total()
    ap_prev = payment_certificate.previous_advance_payment_total
    advance_payment = ap_current + ap_prev

    ret_current = payment_certificate.get_retention_total()
    ret_prev = payment_certificate.previous_retention_total
    retention = ret_current + ret_prev

    mat_current = payment_certificate.get_materials_on_site_total()
    mat_prev = payment_certificate.previous_materials_on_site_total
    material_on_site = mat_current + mat_prev

    # Sum only those SpecialItemTransaction entries specifically categorized as type OTHER
    totals_by_type = payment_certificate.get_special_item_totals_by_type()
    prev_totals_by_type = payment_certificate.previous_special_item_totals_by_type
    other_current = totals_by_type.get("OTHER", Decimal("0.00"))
    other_prev = prev_totals_by_type.get("OTHER", Decimal("0.00"))
    other_specify = other_current + other_prev

    progressive_to_date = (
        (payment_certificate.progressive_to_date or Decimal("0.00"))
        + advance_payment
        + retention
        + material_on_site
        + other_specify
    )
    progressive_previous = (
        (payment_certificate.progressive_previous or Decimal("0.00"))
        + ap_prev
        + ret_prev
        + mat_prev
        + other_prev
    )
    current_claim_total = progressive_to_date - progressive_previous

    vat_val_payment = (
        current_claim_total * Decimal("0.15") if project.vat else Decimal("0.00")
    )
    total_certified = current_claim_total + vat_val_payment


    section_order = cover_page_config.get("section_order") or [
        "section_a",
        "section_b",
        "section_c",
    ]
    current_row = 5

    for sec_id in section_order:
        if sec_id == "section_a":
            info_rows = []
            sec_a_fields = cover_page_config["sections"]["section_a"]["fields"]
            for field in sec_a_fields:
                fid = field["id"]
                label = field["label"]
                enabled = field["enabled"]
                if not enabled:
                    continue

                val = ""
                if fid == "contract_name":
                    val = project.name
                elif fid == "contract_number":
                    val = project.contract_number or ""
                elif fid == "contract_clause":
                    val = project.contract_clause or ""
                elif fid == "description":
                    val = project.description or ""
                elif fid == "client":
                    val = project.client.name if project.client else ""
                elif fid == "status":
                    val = payment_certificate.get_status_display()
                elif fid == "assessment_date":
                    val = cert_date
                elif fid == "certificate_date":
                    val = cert_date

                info_rows.append((label, val))

            for label, val in info_rows:
                ws.row_dimensions[current_row].height = 18.0
                ws.cell(row=current_row, column=1, value=label).font = Font(
                    color=c_grey_text
                )
                ws.cell(row=current_row, column=1).alignment = align_right
                ws.cell(row=current_row, column=2, value=val).font = font_bold
                current_row += 1

            current_row += 1  # Leave an empty row

        elif sec_id == "section_b":
            sec_b_title = cover_page_config["sections"]["section_b"]["title"]
            ws.row_dimensions[current_row].height = 18.0
            ws.cell(row=current_row, column=1, value=sec_b_title).font = font_bold
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).fill = fill_light_grey
                ws.cell(row=current_row, column=col).border = border_thin_top_bottom
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=7,
            )
            current_row += 1

            ws.row_dimensions[current_row].height = 15.75
            ws.cell(row=current_row, column=1, value="Description").font = Font(
                color=c_grey_text
            )
            ws.cell(row=current_row, column=7, value="Value (R)").font = Font(
                color=c_grey_text
            )
            ws.cell(row=current_row, column=7).alignment = align_right
            current_row += 1

            contract_rows = []
            sec_b_fields = cover_page_config["sections"]["section_b"]["fields"]
            for field in sec_b_fields:
                fid = field["id"]
                label = field["label"]
                enabled = field["enabled"]
                if not enabled:
                    continue

                val = Decimal("0.00")
                if fid == "original_value":
                    val = orig_val
                elif fid == "amendments_value":
                    val = amends_val
                elif fid == "sub_total":
                    val = sub_total_contract
                elif fid == "vat":
                    val = vat_val_contract
                elif fid == "total_value":
                    val = total_contract

                contract_rows.append((label, val, fid))

            for i, (desc, val, fid) in enumerate(contract_rows):
                ws.row_dimensions[current_row].height = 16.5
                ws.cell(row=current_row, column=1, value=desc)
                if (
                    desc
                    and "less" in desc.lower()
                    and isinstance(val, (Decimal, float, int))
                    and val > 0
                ):
                    val = -val
                cell_val = ws.cell(row=current_row, column=7, value=val)
                cell_val.number_format = "#,##0.00"

                if fid == "total_value":
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).fill = fill_dark_blue
                    ws.cell(row=current_row, column=1).font = font_white_bold
                    cell_val.font = font_white_bold
                else:
                    if fid in ["sub_total", "total_value"]:
                        ws.cell(row=current_row, column=1).font = font_bold
                        cell_val.font = font_bold
                    if i % 2 != 0:
                        for col in range(1, 8):
                            ws.cell(row=current_row, column=col).fill = fill_light_grey
                current_row += 1

            current_row += 1  # Leave an empty row

        elif sec_id == "section_c":
            sec_c_title = cover_page_config["sections"]["section_c"]["title"]
            ws.row_dimensions[current_row].height = 18.0
            ws.cell(
                row=current_row,
                column=1,
                value=f"{sec_c_title} — CERTIFICATE NO. {cert_num}",
            ).font = font_bold
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).fill = fill_light_grey
                ws.cell(row=current_row, column=col).border = border_thin_top_bottom
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=7,
            )
            current_row += 1

            ws.row_dimensions[current_row].height = 15.75
            ws.cell(row=current_row, column=1, value="Description").font = Font(
                color=c_grey_text
            )
            ws.cell(row=current_row, column=7, value="Amount (R)").font = Font(
                color=c_grey_text
            )
            ws.cell(row=current_row, column=7).alignment = align_right
            current_row += 1

            payment_rows = []
            sec_c_fields = cover_page_config["sections"]["section_c"]["fields"]
            for field in sec_c_fields:
                fid = field["id"]
                label = field["label"]
                enabled = field["enabled"]
                if not enabled:
                    continue

                if fid == "work_progressive_previous":
                    payment_rows.append((label, work_progressive_previous, fid))
                elif fid == "contract_current_claim_total":
                    payment_rows.append((label, contract_current_claim_total, fid))
                elif fid == "addendum_current_claim_total":
                    payment_rows.append((label, addendum_current_claim_total, fid))
                elif fid == "work_progressive_to_date":
                    payment_rows.append((label, work_progressive_to_date, fid))
                elif fid == "special_items":
                    for special_item in payment_certificate.special_items_annotated:
                        total = special_item.total or Decimal("0.00")
                        if total != 0:
                            prefix = "ADD" if total > 0 else "LESS"
                            payment_rows.append(
                                (
                                    f"{prefix}: {special_item.description}",
                                    total,
                                    "special_item",
                                )
                            )
                elif fid == "advance_payment":
                    payment_rows.append((label, advance_payment, fid))
                elif fid == "retention":
                    payment_rows.append((label, retention, fid))
                elif fid == "material_on_site":
                    payment_rows.append((label, material_on_site, fid))
                elif fid == "other_specify":
                    payment_rows.append((label, other_specify, fid))
                elif fid == "progressive_to_date":
                    payment_rows.append((label, progressive_to_date, fid))
                elif fid == "progressive_previous":
                    payment_rows.append((label, progressive_previous, fid))
                elif fid == "current_claim_total":
                    payment_rows.append((label, current_claim_total, fid))
                elif fid == "vat_now":
                    payment_rows.append((label, vat_val_payment, fid))
                elif fid == "total_certified":
                    payment_rows.append((label, total_certified, fid))

            for i, (desc, val, fid) in enumerate(payment_rows):
                ws.row_dimensions[current_row].height = 16.5
                ws.cell(row=current_row, column=1, value=desc)
                if (
                    desc
                    and "less" in desc.lower()
                    and isinstance(val, (Decimal, float, int))
                    and val > 0
                ):
                    val = -val
                cell_val = ws.cell(row=current_row, column=7, value=val)
                cell_val.number_format = "#,##0.00"

                if fid == "total_certified":
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).fill = fill_gold
                    ws.cell(row=current_row, column=1).font = font_white_bold
                    cell_val.font = font_white_bold
                else:
                    if fid in [
                        "progressive_to_date",
                        "current_claim_total",
                        "work_progressive_to_date",
                    ]:
                        ws.cell(row=current_row, column=1).font = font_bold
                        cell_val.font = font_bold
                    if fid == "progressive_previous":
                        ws.cell(row=current_row, column=1).font = font_less
                    if i % 2 != 0:
                        for col in range(1, 8):
                            ws.cell(row=current_row, column=col).fill = fill_light_grey
                current_row += 1

            current_row += 1  # Leave an empty row

    # Retention Note
    note = ""
    ws.cell(row=current_row, column=1, value=note).font = font_italic_small
    ws.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=7
    )
    current_row += 2

    # Authorisation Section
    ws.row_dimensions[current_row].height = 18.0
    ws.cell(row=current_row, column=1, value="AUTHORISATION").font = font_bold
    for col in range(1, 8):
        ws.cell(row=current_row, column=col).fill = fill_light_grey
        ws.cell(row=current_row, column=col).border = border_thin_top_bottom
    ws.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=7
    )
    current_row += 1

    auth_rows = [
        ("Project Manager", ""),
        ("Quantity Surveyor", ""),
        ("Section Manager – Project Governance", ""),
        ("Senior Manager PMU", ""),
    ]

    for title, name in auth_rows:
        ws.row_dimensions[current_row].height = 18.0
        ws.cell(row=current_row, column=1, value=title).font = font_bold

        ws.cell(row=current_row, column=2, value=name)
        ws.cell(row=current_row, column=2).fill = fill_light_grey
        ws.merge_cells(
            start_row=current_row, start_column=2, end_row=current_row, end_column=3
        )

        ws.cell(row=current_row, column=4, value="Signature: ______________________")
        ws.cell(row=current_row, column=4).fill = fill_light_grey
        ws.cell(row=current_row, column=4).font = Font(color=c_grey_text)
        ws.merge_cells(
            start_row=current_row, start_column=4, end_row=current_row, end_column=5
        )

        ws.cell(row=current_row, column=7, value="Date: _______________")
        ws.cell(row=current_row, column=7).fill = fill_light_grey
        ws.cell(row=current_row, column=7).font = Font(color=c_grey_text)
        current_row += 1

    current_row += 1

    # Footer
    footer_text = (
        f"Sedgepro | {project.name} | Payment Certificate No. {cert_num} | {cert_date}"
    )
    ws.cell(row=current_row, column=1, value=footer_text).font = font_italic_small
    ws.cell(row=current_row, column=1).alignment = align_center
    ws.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=7
    )

    return wb
