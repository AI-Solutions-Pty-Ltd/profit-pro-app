import openpyxl

from app.BillOfQuantities.exporters.cover_page_exporter import export_cover_page_to_xlsx
from app.BillOfQuantities.exporters.detailed_report_exporter import (
    export_detailed_report_to_xlsx,
)
from app.BillOfQuantities.exporters.summary_report_exporter import (
    export_summary_report_to_xlsx,
)


def export_unified_xlsx(payment_certificate, sections, is_abridged=False):
    """
    Generate a single XLSX file combining the requested sections.
    """
    wb = openpyxl.Workbook()

    front = sections.get("front", True)
    summary = sections.get("summary", True)
    detailed = sections.get("detailed", True)

    if front:
        export_cover_page_to_xlsx(payment_certificate, wb=wb)

    if summary:
        export_summary_report_to_xlsx(
            payment_certificate, is_abridged=is_abridged, wb=wb
        )

    if detailed:
        export_detailed_report_to_xlsx(
            payment_certificate, is_abridged=is_abridged, wb=wb
        )

    # Remove the default sheet if it is empty and other sheets were created
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # If no sections were selected, just return an empty sheet with a message
    if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
        ws = wb.active
        ws.append(["No sections selected for export."])

    return wb
