from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def generate_payment_certificate_excel(payment_certificate, is_abridged=False):
    """
    Generate a complete payment certificate Excel workbook using openpyxl.
    Matches the RPM/Valterra layout with Front, Summary, and detailed sheets.
    """
    wb = openpyxl.Workbook()

    # Re-fetch project
    project = payment_certificate.project

    # Styles
    font_family = "Arial"
    font_normal = Font(name=font_family, size=9)
    font_bold = Font(name=font_family, size=9, bold=True)
    font_title = Font(name=font_family, size=16, bold=True)

    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    fill_header = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    fill_section = PatternFill(
        start_color="EAEAEA", end_color="EAEAEA", fill_type="solid"
    )

    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )

    double_bottom_border = Border(
        bottom=Side(border_style="double", color="000000"),
        top=Side(border_style="thin", color="000000"),
    )

    # ----------------------------------------------------
    # Detailed Structure Sheets
    # ----------------------------------------------------
    structures = project.structures.all().order_by("name")

    # We will fetch line items once and group them
    from openpyxl.utils import get_column_letter

    from app.BillOfQuantities.models import LineItem

    # Resolve active columns based on project config
    all_columns = project.get_column_config()
    active_columns = [col for col in all_columns if col.get("enabled", True)]

    # Map column id to its 1-based index and letter coordinate
    col_index_map = {col["id"]: idx for idx, col in enumerate(active_columns, start=1)}
    col_letter_map = {col["id"]: get_column_letter(idx) for idx, col in enumerate(active_columns, start=1)}

    if is_abridged:
        annotated_line_items = LineItem.abridged_payment_certificate(
            payment_certificate
        ).order_by("row_index")
    else:
        annotated_line_items = LineItem.construct_payment_certificate(
            payment_certificate
        ).order_by("row_index")

    # Map structure and bills to their subtotal rows
    bill_subtotal_map = {}  # (structure_id, bill_id) -> (sheet_name, row_index)

    for structure in structures:
        # Sanitize sheet name (Excel doesn't allow \, /, ?, *, :, [, ])
        sheet_name = structure.name
        for char in ["\\", "/", "?", "*", ":", "[", "]"]:
            sheet_name = sheet_name.replace(char, " ")
        sheet_name = sheet_name.strip()[:30]

        # Ensure name is unique in workbook and doesn't conflict with reserved names
        base_name = sheet_name
        counter = 1
        existing_names = [s.lower() for s in wb.sheetnames] + ["front", "summary"]
        while sheet_name.lower() in existing_names:
            suffix = f"_{counter}"
            sheet_name = f"{base_name[: 30 - len(suffix)]}{suffix}"
            counter += 1

        # Create sheet
        ws = wb.create_sheet(title=sheet_name)

        # Setup dynamic column widths and headers
        header_vals = []
        for col in active_columns:
            header_vals.append(col["label"].upper())
            col_letter = col_letter_map[col["id"]]
            if col["id"] == "description":
                ws.column_dimensions[col_letter].width = 45
            elif col["id"] == "payment_reference":
                ws.column_dimensions[col_letter].width = 15
            elif col["id"] in ["item_number", "unit_measurement"]:
                ws.column_dimensions[col_letter].width = 10
            else:
                ws.column_dimensions[col_letter].width = 15

        # Row 1 Headers
        ws.append(header_vals)
        ws.append([])  # Row 2 empty

        # Apply header styling to Row 1
        for col in range(1, len(active_columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = font_bold
            cell.alignment = align_center
            cell.fill = fill_header
            cell.border = thin_border

        # Fill data
        struct_items = [
            item for item in annotated_line_items if item.structure_id == structure.id
        ]

        current_bill = None
        bill_start_row = 3
        row_idx = 3

        def write_bill_subtotal(ws, bill, start, end):
            subtotal_row = end + 1
            ws.cell(
                row=subtotal_row, column=1, value="Carried to Summary"
            ).font = font_bold

            # Write SUM formulas only for columns that are enabled
            for key in ["total_price", "total_claimed", "previous_claimed", "current_claim"]:
                if key in col_index_map:
                    col_idx = col_index_map[key]
                    col_let = col_letter_map[key]
                    ws.cell(
                        row=subtotal_row,
                        column=col_idx,
                        value=f"=SUM({col_let}{start}:{col_let}{end})"
                    ).font = font_bold
                    ws.cell(row=subtotal_row, column=col_idx).number_format = "#,##0.00"

            # Apply double underline border
            for col in range(1, len(active_columns) + 1):
                ws.cell(row=subtotal_row, column=col).border = double_bottom_border

            return subtotal_row

        for item in struct_items:
            # If bill changed, write previous subtotal
            if item.bill_id != (current_bill.id if current_bill else None):
                if current_bill:
                    sub_row = write_bill_subtotal(
                        ws, current_bill, bill_start_row, row_idx - 1
                    )
                    bill_subtotal_map[(structure.id, current_bill.id)] = (
                        sheet_name,
                        sub_row,
                    )
                    row_idx = sub_row + 1

                # Write Bill Header
                current_bill = item.bill
                ws.cell(
                    row=row_idx, column=1, value=f"Section No. {structure.name.upper()}"
                ).font = font_bold
                ws.cell(
                    row=row_idx + 1,
                    column=1,
                    value=f"BILL NO. {current_bill.name.upper() if current_bill else ''}",
                ).font = font_bold
                row_idx += 3
                bill_start_row = row_idx

            # Write active columns for item row
            for col in active_columns:
                cid = col["id"]
                col_idx = col_index_map[cid]
                cell = ws.cell(row=row_idx, column=col_idx)

                if cid == "item_number":
                    cell.value = item.item_number
                    cell.alignment = align_center
                elif cid == "payment_reference":
                    cell.value = item.payment_reference
                    cell.alignment = align_center
                elif cid == "description":
                    cell.value = item.description
                    cell.alignment = align_left
                elif cid == "unit_measurement":
                    cell.value = item.unit_measurement
                    cell.alignment = align_center
                elif cid == "budgeted_quantity":
                    if item.is_work:
                        cell.value = float(item.budgeted_quantity or 0)
                        cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif cid == "unit_price":
                    if item.is_work:
                        cell.value = float(item.unit_price or 0)
                        cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif cid == "total_price":
                    if item.is_work:
                        if "budgeted_quantity" in col_letter_map and "unit_price" in col_letter_map:
                            cell.value = f"={col_letter_map['budgeted_quantity']}{row_idx}*{col_letter_map['unit_price']}{row_idx}"
                        else:
                            cell.value = float(item.total_price or 0)
                        cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif cid == "total_qty":
                    if item.is_work:
                        cell.value = float(item.total_qty or 0)
                        cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif cid == "total_claimed":
                    if item.is_work:
                        if "total_qty" in col_letter_map and "unit_price" in col_letter_map:
                            cell.value = f"={col_letter_map['total_qty']}{row_idx}*{col_letter_map['unit_price']}{row_idx}"
                        else:
                            cell.value = float(item.total_claimed or 0)
                        cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif cid == "previous_qty":
                    if item.is_work:
                        cell.value = float(item.previous_qty or 0)
                        cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif cid == "previous_claimed":
                    if item.is_work:
                        if "previous_qty" in col_letter_map and "unit_price" in col_letter_map:
                            cell.value = f"={col_letter_map['previous_qty']}{row_idx}*{col_letter_map['unit_price']}{row_idx}"
                        else:
                            cell.value = float(item.previous_claimed or 0)
                        cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif cid == "current_qty":
                    if item.is_work:
                        if "total_qty" in col_letter_map and "previous_qty" in col_letter_map:
                            cell.value = f"={col_letter_map['total_qty']}{row_idx}-{col_letter_map['previous_qty']}{row_idx}"
                        else:
                            cell.value = float(item.current_qty or 0)
                        cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif cid == "current_claim":
                    if item.is_work:
                        if "current_qty" in col_letter_map and "unit_price" in col_letter_map:
                            cell.value = f"={col_letter_map['current_qty']}{row_idx}*{col_letter_map['unit_price']}{row_idx}"
                        else:
                            cell.value = float(item.current_claim or 0)
                        cell.number_format = "#,##0.00"
                    cell.alignment = align_right

            # Apply fonts and borders
            for col in range(1, len(active_columns) + 1):
                ws.cell(row=row_idx, column=col).font = font_normal
                ws.cell(row=row_idx, column=col).border = thin_border

            row_idx += 1

        # Write subtotal for the last bill
        if current_bill:
            sub_row = write_bill_subtotal(ws, current_bill, bill_start_row, row_idx - 1)
            bill_subtotal_map[(structure.id, current_bill.id)] = (sheet_name, sub_row)

    # ----------------------------------------------------
    # Summary (Valuation Summary) Sheet
    # ----------------------------------------------------
    ws_sum = wb.create_sheet(title="Summary", index=1)
    ws_sum.column_dimensions["A"].width = 12
    ws_sum.column_dimensions["B"].width = 40
    ws_sum.column_dimensions["C"].width = 18
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 18
    ws_sum.column_dimensions["F"].width = 18

    # Headers
    ws_sum.append(
        [
            "SCHED No.",
            "DESCRIPTION",
            "AMOUNT",
            "Cumulative Amount Certified",
            "Previous Amount Certified",
            "Amount Due",
        ]
    )
    for col in range(1, 7):
        cell = ws_sum.cell(row=1, column=col)
        cell.font = font_bold
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = thin_border

    row_idx = 2
    section_rows = []  # list of section subtotal row numbers

    struct_idx = 1
    for structure in structures:
        bills = structure.bills.all().order_by("name")
        if not bills.exists():
            continue

        start_bill_row = row_idx + 1
        ws_sum.cell(
            row=row_idx, column=1, value=f"Section {struct_idx}"
        ).font = font_bold
        ws_sum.cell(
            row=row_idx, column=2, value=structure.name.upper()
        ).font = font_bold
        ws_sum.cell(row=row_idx, column=1).fill = fill_section
        ws_sum.cell(row=row_idx, column=2).fill = fill_section

        row_idx += 1
        bill_idx = 1
        for bill in bills:
            # Get cell mapping
            map_key = (structure.id, bill.id)
            if map_key in bill_subtotal_map:
                sheet_name, sub_row = bill_subtotal_map[map_key]
                ws_sum.cell(
                    row=row_idx, column=1, value=f"Bill No. {bill_idx}"
                ).font = font_normal
                ws_sum.cell(row=row_idx, column=2, value=bill.name).font = font_normal

                # Fetch dynamically using col_letter_map
                total_price_let = col_letter_map.get("total_price", "G")
                total_claimed_let = col_letter_map.get("total_claimed", "I")
                previous_claimed_let = col_letter_map.get("previous_claimed", "K")
                current_claim_let = col_letter_map.get("current_claim", "M")

                ws_sum.cell(
                    row=row_idx,
                    column=3,
                    value=f"='{sheet_name}'!{total_price_let}{sub_row}" if "total_price" in col_letter_map else 0
                ).number_format = "#,##0.00"
                ws_sum.cell(
                    row=row_idx,
                    column=4,
                    value=f"='{sheet_name}'!{total_claimed_let}{sub_row}" if "total_claimed" in col_letter_map else 0
                ).number_format = "#,##0.00"
                ws_sum.cell(
                    row=row_idx,
                    column=5,
                    value=f"='{sheet_name}'!{previous_claimed_let}{sub_row}" if "previous_claimed" in col_letter_map else 0
                ).number_format = "#,##0.00"
                ws_sum.cell(
                    row=row_idx,
                    column=6,
                    value=f"='{sheet_name}'!{current_claim_let}{sub_row}" if "current_claim" in col_letter_map else 0
                ).number_format = "#,##0.00"

                for col in range(1, 7):
                    ws_sum.cell(row=row_idx, column=col).border = thin_border

                row_idx += 1
                bill_idx += 1

        end_bill_row = row_idx - 1

        # Apply section formulas on the Section Header Row
        ws_sum.cell(
            row=start_bill_row - 1,
            column=3,
            value=f"=SUM(C{start_bill_row}:C{end_bill_row})",
        ).number_format = "#,##0.00"
        ws_sum.cell(row=start_bill_row - 1, column=3).font = font_bold
        ws_sum.cell(
            row=start_bill_row - 1,
            column=4,
            value=f"=SUM(D{start_bill_row}:D{end_bill_row})",
        ).number_format = "#,##0.00"
        ws_sum.cell(row=start_bill_row - 1, column=4).font = font_bold
        ws_sum.cell(
            row=start_bill_row - 1,
            column=5,
            value=f"=SUM(E{start_bill_row}:E{end_bill_row})",
        ).number_format = "#,##0.00"
        ws_sum.cell(row=start_bill_row - 1, column=5).font = font_bold
        ws_sum.cell(
            row=start_bill_row - 1,
            column=6,

            value=f"=SUM(F{start_bill_row}:F{end_bill_row})",
        ).number_format = "#,##0.00"
        ws_sum.cell(row=start_bill_row - 1, column=6).font = font_bold

        section_rows.append(start_bill_row - 1)
        struct_idx += 1
        row_idx += 1  # Empty row separating sections

    # Totals at the bottom
    row_idx += 1
    total_row = row_idx
    ws_sum.cell(row=total_row, column=2, value="TOTAL EXCLUDING VAT").font = font_bold

    # Sum of all sections
    c_sum = "+".join([f"C{r}" for r in section_rows])
    d_sum = "+".join([f"D{r}" for r in section_rows])
    e_sum = "+".join([f"E{r}" for r in section_rows])
    f_sum = "+".join([f"F{r}" for r in section_rows])

    ws_sum.cell(
        row=total_row, column=3, value=f"={c_sum}" if c_sum else "=0"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=total_row, column=3).font = font_bold
    ws_sum.cell(
        row=total_row, column=4, value=f"={d_sum}" if d_sum else "=0"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=total_row, column=4).font = font_bold
    ws_sum.cell(
        row=total_row, column=5, value=f"={e_sum}" if e_sum else "=0"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=total_row, column=5).font = font_bold
    ws_sum.cell(
        row=total_row, column=6, value=f"={f_sum}" if f_sum else "=0"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=total_row, column=6).font = font_bold

    # Compensation Events row
    comp_row = total_row + 2
    ws_sum.cell(row=comp_row, column=2, value="COMPENSATION EVENTS").font = font_bold
    for col in range(3, 7):
        ws_sum.cell(row=comp_row, column=col, value=0).number_format = "#,##0.00"
        ws_sum.cell(row=comp_row, column=col).font = font_bold

    # Materials on Site row
    mos_row = total_row + 4
    ws_sum.cell(row=mos_row, column=2, value="MOS").font = font_bold
    for col in range(3, 7):
        ws_sum.cell(row=mos_row, column=col, value=0).number_format = "#,##0.00"
        ws_sum.cell(row=mos_row, column=col).font = font_bold

    # Grand Total row
    grand_row = total_row + 6
    ws_sum.cell(row=grand_row, column=2, value="GRAND TOTAL").font = font_bold
    ws_sum.cell(
        row=grand_row, column=3, value=f"=SUM(C{total_row},C{comp_row},C{mos_row})"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=grand_row, column=3).font = font_bold
    ws_sum.cell(
        row=grand_row, column=4, value=f"=SUM(D{total_row},D{comp_row},D{mos_row})"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=grand_row, column=4).font = font_bold
    ws_sum.cell(
        row=grand_row, column=5, value=f"=SUM(E{total_row},E{comp_row},E{mos_row})"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=grand_row, column=5).font = font_bold
    ws_sum.cell(
        row=grand_row, column=6, value=f"=SUM(F{total_row},F{comp_row},F{mos_row})"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=grand_row, column=6).font = font_bold

    # Retention row
    ret_row = total_row + 8
    ws_sum.cell(row=ret_row, column=2, value="RETENTION ACCOUNT").font = font_bold
    ws_sum.cell(
        row=ret_row, column=3, value=f"=C{grand_row}*0.1"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=ret_row, column=3).font = font_bold
    ws_sum.cell(
        row=ret_row, column=4, value=f"=D{grand_row}*0.1"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=ret_row, column=4).font = font_bold
    ws_sum.cell(
        row=ret_row, column=5, value=f"=E{grand_row}*0.1"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=ret_row, column=5).font = font_bold
    ws_sum.cell(
        row=ret_row, column=6, value=f"=F{grand_row}*0.05"
    ).number_format = "#,##0.00"
    ws_sum.cell(row=ret_row, column=6).font = font_bold

    # Apply double underline borders to totals
    for r in [total_row, grand_row]:
        for col in range(1, 7):
            ws_sum.cell(row=r, column=col).border = double_bottom_border

    # ----------------------------------------------------
    # Front Cover Page Sheet
    # ----------------------------------------------------
    ws_front = wb.active  # Use the first worksheet
    ws_front.title = "Front"
    ws_front.column_dimensions["A"].width = 25
    ws_front.column_dimensions["B"].width = 5
    ws_front.column_dimensions["C"].width = 35
    ws_front.column_dimensions["D"].width = 20
    ws_front.column_dimensions["E"].width = 5
    ws_front.column_dimensions["F"].width = 5
    ws_front.column_dimensions["G"].width = 20

    # Write Cover details
    ws_front["A1"] = "Certificate no"
    ws_front["C1"] = f"{payment_certificate.certificate_number:02d}"
    ws_front["A2"] = "Assessment Date"
    ws_front["C2"] = (
        payment_certificate.assessment_date.date()
        if payment_certificate.assessment_date
        else datetime.now().date()
    )
    ws_front["C2"].number_format = "YYYY-MM-DD"
    ws_front["A3"] = "Contract Amendment"
    ws_front["A4"] = "Retention Free Amount"
    ws_front["C4"] = "=G24*20%"
    ws_front["C4"].number_format = "#,##0.00"
    ws_front["A5"] = "Date of Cert"
    ws_front["C5"] = (
        payment_certificate.approved_on.date()
        if payment_certificate.approved_on
        else datetime.now().date()
    )
    ws_front["C5"].number_format = "YYYY-MM-DD"

    for r in range(1, 6):
        ws_front.cell(row=r, column=1).font = font_bold
        ws_front.cell(row=r, column=3).font = font_normal

    ws_front["A12"] = "Assessment Date"
    ws_front["C12"] = "=C2"
    ws_front["C12"].number_format = "YYYY-MM-DD"
    ws_front["A14"] = "Date of Certificate"
    ws_front["C14"] = "=C5"
    ws_front["C14"].number_format = "YYYY-MM-DD"

    ws_front["A16"] = "CONTRACT:"
    ws_front["A16"].font = font_bold
    ws_front["C16"] = project.name.upper()
    ws_front["C16"].font = font_title

    ws_front["A19"] = (
        '="Herewith certificate no. "&C1&"  for the above mentioned project duly signed and certified for payment in accordance with clause 50 and 51"'
    )
    ws_front["A19"].font = font_bold

    # Values block
    ws_front["A21"] = "Contract No."
    ws_front["C21"] = "Contract"
    ws_front["G21"] = "Value"
    for col in [1, 3, 7]:
        ws_front.cell(row=21, column=col).font = font_bold
        ws_front.cell(row=21, column=col).border = thin_border

    ws_front["A22"] = project.contract_number
    ws_front["C22"] = "Original Contract Value"
    ws_front["G22"] = f"=Summary!C{total_row}"
    ws_front["G22"].number_format = "#,##0.00"

    ws_front["C23"] = "Total Contract Amendments To Date"
    ws_front["G23"] = 0
    ws_front["G23"].number_format = "#,##0.00"

    ws_front["C24"] = "Sub Total"
    ws_front["G24"] = "=SUM(G22:G23)"
    ws_front["G24"].number_format = "#,##0.00"

    ws_front["C25"] = "VAT 15%"
    ws_front["G25"] = "=G24*0.15"
    ws_front["G25"].number_format = "#,##0.00"

    ws_front["C26"] = "Total Contract Value"
    ws_front["G26"] = "=G24+G25"
    ws_front["G26"].number_format = "#,##0.00"

    for r in range(22, 27):
        ws_front.cell(row=r, column=1).font = font_normal
        ws_front.cell(row=r, column=3).font = font_normal
        ws_front.cell(row=r, column=7).font = font_normal
        ws_front.cell(row=r, column=1).border = thin_border
        ws_front.cell(row=r, column=3).border = thin_border
        ws_front.cell(row=r, column=7).border = thin_border

    # Double border under total contract value
    for col in [1, 3, 7]:
        ws_front.cell(row=26, column=col).border = double_bottom_border

    # Claim calculations
    ws_front["A29"] = "Payment Due:"
    ws_front["A29"].font = font_bold
    ws_front["G30"] = "Value"
    ws_front["G30"].font = font_bold

    ws_front["A31"] = '="Value of work done up to the assesment interval no. "&C1'
    ws_front["G31"] = f"=Summary!F{total_row}"
    ws_front["G31"].number_format = "#,##0.00"

    ws_front["A32"] = "Plus : Compensation Events"
    ws_front["G32"] = f"=Summary!D{comp_row}"
    ws_front["G32"].number_format = "#,##0.00"

    ws_front["A33"] = "Plus : Material On Site"
    ws_front["G33"] = f"=Summary!D{mos_row}"
    ws_front["G33"].number_format = "#,##0.00"

    ws_front["A34"] = "Plus : Price adjustment for inflation"
    ws_front["G34"] = 0
    ws_front["G34"].number_format = "#,##0.00"

    ws_front["A35"] = "Total value of work done"
    ws_front["G35"] = "=SUM(G31:G34)"
    ws_front["G35"].number_format = "#,##0.00"

    ws_front["A36"] = "Less : Retention"
    ws_front["G36"] = (
        "=G35*0"  # Matching analyzed RPM Cover sheet where retention is not deducted
    )
    ws_front["G36"].number_format = "#,##0.00"

    ws_front["A37"] = "Sub Total"
    ws_front["G37"] = "=G35-G36"
    ws_front["G37"].number_format = "#,##0.00"

    ws_front["A38"] = "Less : Previous Amount Due"
    ws_front["G38"] = f"=Summary!E{total_row}"
    ws_front["G38"].number_format = "#,##0.00"

    ws_front["A39"] = "Sub Total"
    ws_front["G39"] = "=G37-G38"
    ws_front["G39"].number_format = "#,##0.00"

    ws_front["A40"] = "Plus : V.A.T. at 15%"
    ws_front["G40"] = "=G39*15%"
    ws_front["G40"].number_format = "#,##0.00"

    ws_front["A41"] = "TOTAL AMOUNT NOW CERTIFIED"
    ws_front["G41"] = "=G39+G40"
    ws_front["G41"].number_format = "#,##0.00"

    ws_front["A44"] = "TOTAL AMOUNT NOW DUE FOR PAYMENT TO CONTRACTOR"
    ws_front["G44"] = "=G41"
    ws_front["G44"].number_format = "#,##0.00"

    ws_front["A46"] = (
        "5% Retention (Not deducted on Certificate, Deduction done by Valterra GSS)"
    )
    ws_front["G46"] = "=G39*0.05"
    ws_front["G46"].number_format = "#,##0.00"

    # Fonts and borders for payment due cells
    for r in range(31, 47):
        if ws_front.cell(row=r, column=1).value or ws_front.cell(row=r, column=7).value:
            ws_front.cell(row=r, column=1).font = font_normal
            ws_front.cell(row=r, column=7).font = font_normal
            ws_front.cell(row=r, column=1).border = thin_border
            ws_front.cell(row=r, column=7).border = thin_border

    ws_front.cell(row=35, column=1).font = font_bold
    ws_front.cell(row=35, column=7).font = font_bold
    ws_front.cell(row=37, column=1).font = font_bold
    ws_front.cell(row=37, column=7).font = font_bold
    ws_front.cell(row=39, column=1).font = font_bold
    ws_front.cell(row=39, column=7).font = font_bold
    ws_front.cell(row=41, column=1).font = font_bold
    ws_front.cell(row=41, column=7).font = font_bold
    ws_front.cell(row=44, column=1).font = font_bold
    ws_front.cell(row=44, column=7).font = font_bold

    ws_front.cell(row=41, column=1).border = double_bottom_border
    ws_front.cell(row=41, column=7).border = double_bottom_border
    ws_front.cell(row=44, column=1).border = double_bottom_border
    ws_front.cell(row=44, column=7).border = double_bottom_border

    # Signatures
    ws_front["A57"] = "Signed :...............................……….."
    ws_front["D57"] = "Project Manager  (M  Susela)"
    ws_front["G57"] = "Date :"

    ws_front["A60"] = "Signed :...............................……….."
    ws_front["D60"] = "Quantity Surveyor (M. Hachinavamwe)"
    ws_front["G60"] = "Date :"

    ws_front["A63"] = "Signed :...............................……….."
    ws_front["D63"] = "Section Manager Project Governance  (S. Tsokalamtengo)"
    ws_front["G63"] = "Date :"

    ws_front["A66"] = "Signed :...............................……….."
    ws_front["D66"] = "Senior Manager PMU (H. Sifeni)"
    ws_front["G66"] = "Date :"

    for r in [57, 60, 63, 66]:
        ws_front.cell(row=r, column=1).font = font_bold
        ws_front.cell(row=r, column=4).font = font_bold
        ws_front.cell(row=r, column=7).font = font_bold

    ws_front["A69"] = "RUSTENBURG PLATINUM MINES (PTY) LTD"
    ws_front["A69"].font = font_bold
    ws_front["A70"] = "144 Oxford Street,Rosebank, Melrose, 2196"
    ws_front["A70"].font = font_normal
    ws_front["A71"] = "P O Box 62179, Marshalltown, 2107"
    ws_front["A71"].font = font_normal

    return wb
