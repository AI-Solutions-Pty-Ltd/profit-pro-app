from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.BillOfQuantities.models import LineItem


def export_detailed_report_to_xlsx(payment_certificate, is_abridged=False, wb=None):
    """
    Export the detailed report for a payment certificate to XLSX format.
    Mirrors the layout of '03_MediaCentre_Detailed (1).xlsx'.
    """
    from app.BillOfQuantities.tasks import group_line_items_by_hierarchy

    if wb is None:
        wb = openpyxl.Workbook()

    # We will create a sheet for each structure (section)
    if is_abridged:
        all_line_items = LineItem.abridged_payment_certificate(payment_certificate)
        line_items = all_line_items.filter(addendum=False, special_item=False)
    else:
        line_items = LineItem.construct_payment_certificate(payment_certificate)

    grouped_data = group_line_items_by_hierarchy(line_items)

    # active columns logic
    project = payment_certificate.project
    all_columns = project.get_column_config()
    active_columns = [col for col in all_columns if col.get("enabled", True)]
    num_cols = len(active_columns)
    col_idx_map = {col["id"]: idx for idx, col in enumerate(active_columns, 1)}

    # Styles
    font_bold = Font(bold=True)
    font_bold_white = Font(bold=True, color="FFFFFFFF")
    Font(italic=True, color="FF666666")
    font_subtitle = Font(italic=True, color="FF666666")
    Font(bold=True, size=11)
    font_title = Font(bold=True, size=14)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    Border(bottom=Side(style="thin"))
    border_bottom_thick = Border(bottom=Side(style="medium"))
    border_bottom_light = Border(bottom=Side(style="thin", color="FFE5E5E5"))

    fill_bill_header = PatternFill(
        start_color="FF333333", end_color="FF333333", fill_type="solid"
    )
    fill_package_header = PatternFill(
        start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid"
    )
    fill_section_footer = PatternFill(
        start_color="FFD19B3D", end_color="FFD19B3D", fill_type="solid"
    )
    fill_zebra_even = PatternFill(
        start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid"
    )
    fill_zebra_odd = PatternFill(
        start_color="FFF9F9F9", end_color="FFF9F9F9", fill_type="solid"
    )
    fill_column_headers = PatternFill(
        start_color="FF111111", end_color="FF111111", fill_type="solid"
    )

    if not grouped_data:
        ws = wb.active
        ws.title = "No Data"
        ws.append(["No data available for this report."])
        return wb

    # Remove the default sheet created by openpyxl

    import re

    for structure_idx, structure_data in enumerate(grouped_data, 1):
        structure = structure_data["structure"]
        # Excel sheet names max 31 chars and cannot contain \ * ? : / [ ]
        sheet_name = re.sub(r"[\\*?:/\[\]]", "_", structure.name)[:31]

        # If it's the first sheet being added, we can rename the default or just create new and delete default later
        ws = wb.create_sheet(title=sheet_name)

        cert_num = str(payment_certificate.certificate_number).zfill(2)
        cert_date = payment_certificate.created_at.strftime("%d %b %Y")
        project_name = payment_certificate.project.name

        # Row 1: Header
        ws.cell(row=1, column=1, value="[ LOGO ]").font = font_bold

        # Calculate merge ranges based on num_cols
        title_end_col = max(2, num_cols - 2)
        cert_start_col = title_end_col + 1

        title_cell = ws.cell(
            row=1, column=2, value=f"SECTION {structure_idx} — {structure.name.upper()}"
        )
        title_cell.font = font_title
        title_cell.alignment = align_center
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=title_end_col)

        cert_cell = ws.cell(
            row=1, column=cert_start_col, value=f"Cert No. {cert_num}\n{cert_date}"
        )
        cert_cell.alignment = Alignment(
            wrap_text=True, horizontal="right", vertical="center"
        )
        cert_cell.font = font_bold
        if cert_start_col < num_cols:
            ws.merge_cells(
                start_row=1, start_column=cert_start_col, end_row=1, end_column=num_cols
            )

        for col in range(cert_start_col, num_cols + 1):
            ws.cell(row=1, column=col).fill = fill_package_header

        ws.row_dimensions[1].height = 60

        # Row 2: Subtitle
        sub_cell = ws.cell(
            row=2,
            column=1,
            value=f"{project_name} - Payment Certificate No. {cert_num} - {cert_date}",
        )
        sub_cell.font = font_subtitle
        sub_cell.alignment = align_center
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        for col in range(1, num_cols + 1):
            ws.cell(row=2, column=col).fill = fill_package_header

        # Row 3: Empty

        # Row 4: Column Headers
        ws.row_dimensions[4].height = 30
        for col_idx, col_config in enumerate(active_columns, 1):
            cell = ws.cell(row=4, column=col_idx, value=col_config["label"])
            cell.font = font_bold_white
            cell.fill = fill_column_headers
            cell.alignment = align_center

        current_row = 5

        # Data Rows
        for bill_idx, bill_data in enumerate(structure_data["bills"], 1):
            bill = bill_data["bill"]

            # Bill Header
            ws.row_dimensions[current_row].height = 25
            ws.cell(
                row=current_row,
                column=1,
                value=f"BILL NO. {bill_idx} — {bill.name.upper()}",
            )

            if "total_price" in col_idx_map:
                ws.cell(
                    row=current_row,
                    column=col_idx_map["total_price"],
                    value=bill_data["budget"],
                ).alignment = align_right
            if "total_claimed" in col_idx_map:
                ws.cell(
                    row=current_row,
                    column=col_idx_map["total_claimed"],
                    value=bill_data["cumulative"],
                ).alignment = align_right
            if "previous_claimed" in col_idx_map:
                ws.cell(
                    row=current_row,
                    column=col_idx_map["previous_claimed"],
                    value=bill_data["previous"],
                ).alignment = align_right
            if "current_claim" in col_idx_map:
                ws.cell(
                    row=current_row,
                    column=col_idx_map["current_claim"],
                    value=bill_data["current"],
                ).alignment = align_right

            for col in range(1, num_cols + 1):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_bill_header
                cell.font = font_bold_white
            current_row += 1

            for package_data in bill_data["packages"]:
                package = package_data["package"]
                if package and package.name:
                    # Package Header
                    ws.row_dimensions[current_row].height = 20
                    ws.cell(row=current_row, column=1, value=f"{package.name}")
                    for col in range(1, num_cols + 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = font_subtitle
                    current_row += 1

                item_row_count = 0
                for item in package_data["line_items"]:
                    ws.row_dimensions[current_row].height = 20

                    val_map = {
                        "item_number": getattr(item, "item_number", ""),
                        "payment_reference": getattr(item, "payment_reference", ""),
                        "description": getattr(item, "description", ""),
                        "unit_measurement": getattr(item, "unit_measurement", ""),
                        "budgeted_quantity": getattr(item, "budgeted_quantity", None),
                        "unit_price": getattr(item, "unit_price", None),
                        "total_price": getattr(item, "total_price", None),
                        "total_qty": getattr(item, "total_qty", None),
                        "total_claimed": getattr(item, "total_claimed", None),
                        "previous_qty": getattr(item, "previous_qty", None),
                        "previous_claimed": getattr(item, "previous_claimed", None),
                        "current_qty": getattr(item, "current_qty", None),
                        "current_claim": getattr(item, "current_claim", None),
                    }

                    for col_idx, col_config in enumerate(active_columns, 1):
                        col_id = col_config["id"]
                        val = val_map.get(col_id)
                        if val is None:
                            val = ""

                        cell = ws.cell(row=current_row, column=col_idx, value=val)

                        if col_id in ("item_number", "unit_measurement"):
                            cell.alignment = align_center
                        elif col_id == "description":
                            cell.alignment = align_wrap
                        elif col_id in ("payment_reference",):
                            cell.alignment = align_left
                        else:
                            cell.alignment = align_right

                    fill_to_use = (
                        fill_zebra_even if item_row_count % 2 == 0 else fill_zebra_odd
                    )
                    for col in range(1, num_cols + 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.fill = fill_to_use
                        cell.border = border_bottom_light

                    current_row += 1
                    item_row_count += 1

            # Bill Footer
            ws.row_dimensions[current_row].height = 20
            ws.cell(
                row=current_row,
                column=1,
                value=f"Carried to Summary — Bill No. {bill_idx}",
            )

            # Merge footer text up to right before amounts
            footer_merge_end = (
                min(
                    col_idx_map.get("total_price", num_cols),
                    col_idx_map.get("total_claimed", num_cols),
                    col_idx_map.get("previous_claimed", num_cols),
                    col_idx_map.get("current_claim", num_cols),
                )
                - 1
            )
            if footer_merge_end < 1:
                footer_merge_end = 1

            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=footer_merge_end,
            )
            ws.cell(row=current_row, column=1).alignment = align_right

            if "total_price" in col_idx_map:
                ws.cell(
                    row=current_row,
                    column=col_idx_map["total_price"],
                    value=bill_data["budget"],
                ).alignment = align_right
            if "total_claimed" in col_idx_map:
                ws.cell(
                    row=current_row,
                    column=col_idx_map["total_claimed"],
                    value=bill_data["cumulative"],
                ).alignment = align_right
            if "previous_claimed" in col_idx_map:
                ws.cell(
                    row=current_row,
                    column=col_idx_map["previous_claimed"],
                    value=bill_data["previous"],
                ).alignment = align_right
            if "current_claim" in col_idx_map:
                ws.cell(
                    row=current_row,
                    column=col_idx_map["current_claim"],
                    value=bill_data["current"],
                ).alignment = align_right

            for col in range(1, num_cols + 1):
                cell = ws.cell(row=current_row, column=col)
                cell.font = font_bold
                cell.border = border_bottom_thick
            current_row += 1

        # Structure Footer
        ws.cell(
            row=current_row,
            column=1,
            value=f"SECTION {structure_idx} TOTAL — {structure.name.upper()}",
        )

        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=footer_merge_end,
        )
        ws.cell(row=current_row, column=1).alignment = align_right

        if "total_price" in col_idx_map:
            ws.cell(
                row=current_row,
                column=col_idx_map["total_price"],
                value=structure_data["budget"],
            ).alignment = align_right
        if "total_claimed" in col_idx_map:
            ws.cell(
                row=current_row,
                column=col_idx_map["total_claimed"],
                value=structure_data["cumulative"],
            ).alignment = align_right
        if "previous_claimed" in col_idx_map:
            ws.cell(
                row=current_row,
                column=col_idx_map["previous_claimed"],
                value=structure_data["previous"],
            ).alignment = align_right
        if "current_claim" in col_idx_map:
            ws.cell(
                row=current_row,
                column=col_idx_map["current_claim"],
                value=structure_data["current"],
            ).alignment = align_right

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = fill_section_footer
            cell.font = font_bold_white
        current_row += 2

        # Final Footer
        company = "Sedgepro"  # Or get from settings/project
        footer_text = (
            f"{company}  |  Payment Certificate No. {cert_num}  |  {cert_date}"
        )
        ws.cell(row=current_row, column=1, value=footer_text)
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=num_cols,
        )

        # Set Column Widths dynamically
        for col_idx, col_config in enumerate(active_columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            col_id = col_config["id"]
            if col_id in ("item_number",):
                width = 10
            elif col_id in ("payment_reference",):
                width = 12
            elif col_id == "description":
                width = 65
            else:
                # Give enough space for the custom label or default 15
                width = max(15, len(col_config.get("label", "")) + 2)
            ws.column_dimensions[col_letter].width = width

        # Format numbers
        for row in ws.iter_rows(min_row=5, max_row=current_row):
            for col_idx, col_config in enumerate(active_columns, 1):
                col_id = col_config["id"]
                if col_id not in (
                    "item_number",
                    "payment_reference",
                    "description",
                    "unit_measurement",
                ):
                    cell = row[col_idx - 1]
                    if isinstance(cell.value, (int, float, Decimal)):
                        cell.number_format = "#,##0.00"

    # Remove the default sheet if we added custom sheets
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    return wb
