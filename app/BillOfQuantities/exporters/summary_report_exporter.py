from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_summary_report_to_xlsx(payment_certificate, is_abridged=False, wb=None):
    """
    Export the summary report for a payment certificate to XLSX format.
    Mirrors the layout of '02_Summary.xlsx' and the UI valuation summary.
    """
    from app.BillOfQuantities.tasks import get_valuation_summary_data

    mode_text = "Abridged" if is_abridged else "Full"
    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Summary - {mode_text}"
    else:
        ws = wb.create_sheet(title=f"Summary - {mode_text}")

    cert_num = str(payment_certificate.certificate_number).zfill(2)
    cert_date = payment_certificate.created_at.strftime("%d %b %Y")
    project_name = payment_certificate.project.name

    # Styles
    font_bold = Font(bold=True)
    font_bold_white = Font(bold=True, color="FFFFFFFF")
    font_subtitle = Font(italic=True, color="FF666666")
    font_title = Font(bold=True, size=14)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    border_bottom_thick = Border(bottom=Side(style="medium"))
    border_bottom_light = Border(bottom=Side(style="thin", color="FFE5E5E5"))

    fill_header = PatternFill(
        start_color="FF111111", end_color="FF111111", fill_type="solid"
    )
    fill_section = PatternFill(
        start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid"
    )
    fill_footer = PatternFill(
        start_color="FFD19B3D", end_color="FFD19B3D", fill_type="solid"
    )

    # Header Row
    ws.cell(row=1, column=1, value="[ LOGO ]").font = font_bold
    title_cell = ws.cell(
        row=1, column=2, value=f"VALUATION SUMMARY — {mode_text.upper()}"
    )
    title_cell.font = font_title
    title_cell.alignment = align_center
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)

    cert_cell = ws.cell(row=1, column=5, value=f"Cert No. {cert_num}\n{cert_date}")
    cert_cell.alignment = Alignment(
        wrap_text=True, horizontal="right", vertical="center"
    )
    cert_cell.font = font_bold
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
    ws.row_dimensions[1].height = 40

    # Subtitle Row
    sub_cell = ws.cell(
        row=2, column=1, value=f"{project_name} - Payment Certificate No. {cert_num}"
    )
    sub_cell.font = font_subtitle
    sub_cell.alignment = align_center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # Empty Row 3

    # Column Headers
    headers = [
        # "SCHED",
        "DESCRIPTION",
        "TENDER AMOUNT (R)",
        "CUMULATIVE CERTIFIED (R)",
        "PREVIOUS CERTIFIED (R)",
        "AMOUNT DUE (R)",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = font_bold_white
        cell.fill = fill_header
        cell.alignment = align_center
    ws.row_dimensions[4].height = 25

    # Fetch Data
    summary_data = get_valuation_summary_data(payment_certificate, abridged=is_abridged)
    grouped_sections = summary_data["grouped_sections"]
    total_budget = summary_data["total_budget"]
    total_cumulative = summary_data["total_cumulative"]
    total_previous = summary_data["total_previous"]
    total_current = summary_data["total_current"]

    current_row = 5

    if not grouped_sections:
        ws.cell(row=current_row, column=1, value="No valuation data available.")
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=6
        )
        ws.cell(row=current_row, column=1).alignment = align_center
        current_row += 1
    else:
        for section in grouped_sections:
            # Section Header
            ws.cell(row=current_row, column=1, value="")
            ws.cell(
                row=current_row, column=2, value=str(section["name"]).upper()
            ).font = font_bold
            ws.cell(
                row=current_row, column=3, value=section["budget"]
            ).alignment = align_right
            ws.cell(
                row=current_row, column=4, value=section["cumulative"]
            ).alignment = align_right
            ws.cell(
                row=current_row, column=5, value=section["previous"]
            ).alignment = align_right
            ws.cell(
                row=current_row, column=6, value=section["current"]
            ).alignment = align_right

            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_section
                cell.font = font_bold
                cell.border = border_bottom_thick
            current_row += 1

            # Bills
            for bill in section["bills"]:
                ws.cell(row=current_row, column=1, value="")
                ws.cell(row=current_row, column=2, value=bill["name"])
                ws.cell(
                    row=current_row, column=3, value=bill["budget"]
                ).alignment = align_right
                ws.cell(
                    row=current_row, column=4, value=bill["cumulative"]
                ).alignment = align_right
                ws.cell(
                    row=current_row, column=5, value=bill["previous"]
                ).alignment = align_right
                ws.cell(
                    row=current_row, column=6, value=bill["current"]
                ).alignment = align_right

                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = border_bottom_light
                current_row += 1

            # Empty Row after section
            current_row += 1

        # Totals
        ws.cell(
            row=current_row, column=2, value="TOTAL WORK DONE TO DATE"
        ).font = font_bold
        ws.cell(row=current_row, column=2).alignment = align_right

        ws.cell(row=current_row, column=3, value=total_budget).alignment = align_right
        ws.cell(
            row=current_row, column=4, value=total_cumulative
        ).alignment = align_right
        ws.cell(row=current_row, column=5, value=total_previous).alignment = align_right
        ws.cell(row=current_row, column=6, value=total_current).alignment = align_right
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.font = font_bold
            cell.fill = fill_section
        current_row += 2

        # Contractual Special Items Section
        if payment_certificate.has_contractual_special_items:
            from app.BillOfQuantities.models.structure_models import LineItem

            all_line_items = LineItem.construct_payment_certificate(payment_certificate)
            addendum_items = all_line_items.filter(addendum=True, special_item=False)
            special_items = all_line_items.filter(addendum=False, special_item=True)

            ws.cell(
                row=current_row, column=2, value="CONTRACTUAL SPECIAL ITEMS"
            ).font = Font(bold=True, size=11)
            for col in range(2, 7):
                ws.cell(row=current_row, column=col).fill = fill_section
            current_row += 1

            # 1. Addendums
            if addendum_items.exists():
                ws.cell(
                    row=current_row, column=2, value="Addendum Line Items"
                ).font = font_bold
                current_row += 1
                for item in addendum_items:
                    ws.cell(row=current_row, column=2, value=item.description)
                    ws.cell(
                        row=current_row, column=3, value=item.total_price
                    ).alignment = align_right
                    ws.cell(
                        row=current_row, column=4, value=item.total_claimed
                    ).alignment = align_right
                    ws.cell(
                        row=current_row, column=5, value=item.previous_claimed
                    ).alignment = align_right
                    ws.cell(
                        row=current_row, column=6, value=item.current_claim
                    ).alignment = align_right
                    for col in range(1, 7):
                        ws.cell(
                            row=current_row, column=col
                        ).border = border_bottom_light
                    current_row += 1

                # Subtotal Addendum
                ws.cell(
                    row=current_row, column=2, value="Subtotal Addendum Items"
                ).alignment = align_right
                ws.cell(row=current_row, column=2).font = font_bold
                ws.cell(
                    row=current_row,
                    column=3,
                    value=payment_certificate.addendum_budget_total,
                ).alignment = align_right
                ws.cell(
                    row=current_row,
                    column=4,
                    value=payment_certificate.addendum_progressive_to_date,
                ).alignment = align_right
                ws.cell(
                    row=current_row,
                    column=5,
                    value=payment_certificate.addendum_progressive_previous,
                ).alignment = align_right
                ws.cell(
                    row=current_row,
                    column=6,
                    value=payment_certificate.addendum_current_claim_total,
                ).alignment = align_right
                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = font_bold
                    cell.border = border_bottom_thick
                current_row += 1

            # 2. Special Line Items
            if special_items.exists():
                ws.cell(
                    row=current_row, column=2, value="Special Line Items"
                ).font = font_bold
                current_row += 1
                for item in special_items:
                    ws.cell(row=current_row, column=2, value=item.description)
                    ws.cell(
                        row=current_row, column=3, value=item.total_price
                    ).alignment = align_right
                    ws.cell(
                        row=current_row, column=4, value=item.total_claimed
                    ).alignment = align_right
                    ws.cell(
                        row=current_row, column=5, value=item.previous_claimed
                    ).alignment = align_right
                    ws.cell(
                        row=current_row, column=6, value=item.current_claim
                    ).alignment = align_right
                    for col in range(1, 7):
                        ws.cell(
                            row=current_row, column=col
                        ).border = border_bottom_light
                    current_row += 1

                # Subtotal Special Items
                ws.cell(
                    row=current_row, column=2, value="Subtotal Special Items"
                ).alignment = align_right
                ws.cell(row=current_row, column=2).font = font_bold
                ws.cell(
                    row=current_row,
                    column=3,
                    value=payment_certificate.special_items_budget_total,
                ).alignment = align_right
                ws.cell(
                    row=current_row,
                    column=4,
                    value=payment_certificate.special_items_progressive_to_date,
                ).alignment = align_right
                ws.cell(
                    row=current_row,
                    column=5,
                    value=payment_certificate.special_items_progressive_previous,
                ).alignment = align_right
                ws.cell(
                    row=current_row,
                    column=6,
                    value=payment_certificate.special_items_current_claim_total,
                ).alignment = align_right
                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = font_bold
                    cell.border = border_bottom_thick
                current_row += 1

            # 3. Ledger Totals
            ledger_items = payment_certificate.get_ledger_summary_items()
            if ledger_items:
                for item in ledger_items:
                    ws.cell(row=current_row, column=2, value=item["description"])
                    ws.cell(row=current_row, column=3, value="").alignment = align_right
                    ws.cell(
                        row=current_row, column=4, value=item["total_amount"]
                    ).alignment = align_right
                    ws.cell(
                        row=current_row, column=5, value=item["previous_amount"]
                    ).alignment = align_right
                    ws.cell(
                        row=current_row, column=6, value=item["current_amount"]
                    ).alignment = align_right
                    for col in range(1, 7):
                        ws.cell(
                            row=current_row, column=col
                        ).border = border_bottom_light
                    current_row += 1

            # Grand Total Contractual Special Items
            ws.cell(
                row=current_row, column=2, value="TOTAL CONTRACTUAL SPECIAL ITEMS"
            ).alignment = align_right
            ws.cell(row=current_row, column=2).font = font_bold
            ws.cell(
                row=current_row,
                column=3,
                value=payment_certificate.addendum_budget_total
                + payment_certificate.special_items_budget_total,
            ).alignment = align_right
            ws.cell(
                row=current_row,
                column=4,
                value=payment_certificate.contractual_special_items_progressive_to_date,
            ).alignment = align_right
            ws.cell(
                row=current_row,
                column=5,
                value=payment_certificate.contractual_special_items_progressive_previous,
            ).alignment = align_right
            ws.cell(
                row=current_row,
                column=6,
                value=payment_certificate.contractual_special_items_current_claim_total,
            ).alignment = align_right
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.font = font_bold
                cell.fill = fill_section
                cell.border = border_bottom_thick
            current_row += 2

        # Grand Total
        ws.cell(row=current_row, column=2, value="GRAND TOTAL").font = font_bold
        ws.cell(
            row=current_row,
            column=3,
            value=payment_certificate.project.total_contract_value,
        ).alignment = align_right
        ws.cell(
            row=current_row,
            column=4,
            value=payment_certificate.grand_total_progressive_to_date,
        ).alignment = align_right
        ws.cell(
            row=current_row,
            column=5,
            value=payment_certificate.grand_total_progressive_previous,
        ).alignment = align_right
        ws.cell(
            row=current_row,
            column=6,
            value=payment_certificate.current_claim_total
            + payment_certificate.ledger_current_net_total,
        ).alignment = align_right
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.font = font_bold
            cell.fill = fill_footer
        current_row += 1

    # Set Column Widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 20

    # Format numbers
    for row in ws.iter_rows(min_row=5, max_row=current_row):
        for cell in row[2:6]:  # columns C to F
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = "#,##0.00"

    return wb
