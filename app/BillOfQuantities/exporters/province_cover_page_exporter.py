import datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_province_cover_page_to_xlsx(province, projects):
    """
    Export the aggregated cover page report for a province to XLSX format.
    Mirrors the styling, colors, and layout of cover_page_exporter.py.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Province Cover Page"

    # Aggregated Values
    orig_val = Decimal("0.00")
    amends_val = Decimal("0.00")
    sub_total_contract = Decimal("0.00")
    vat_val_contract = Decimal("0.00")
    total_contract = Decimal("0.00")

    work_progressive_previous = Decimal("0.00")
    contract_current_claim_total = Decimal("0.00")
    addendum_current_claim_total = Decimal("0.00")
    work_progressive_to_date = Decimal("0.00")
    advance_payment = Decimal("0.00")
    retention = Decimal("0.00")
    material_on_site = Decimal("0.00")
    other_specify = Decimal("0.00")

    project_count = len(projects)
    cert_count = 0

    for project in projects:
        orig_val += project.original_contract_value or Decimal("0.00")
        amends_val += project.addendum_contract_value or Decimal("0.00")

        # Get latest approved payment certificate
        cert = (
            project.payment_certificates.filter(status="APPROVED")
            .order_by("-certificate_number")
            .first()
        )

        if cert:
            cert_count += 1
            work_progressive_previous += cert.work_progressive_previous or Decimal(
                "0.00"
            )
            contract_current_claim_total += (
                cert.contract_current_claim_total or Decimal("0.00")
            )
            addendum_current_claim_total += (
                cert.addendum_current_claim_total or Decimal("0.00")
            )
            work_progressive_to_date += cert.work_progressive_to_date or Decimal("0.00")

            advance_payment += (
                cert.get_advance_payment_total() + cert.previous_advance_payment_total
            )
            retention += cert.get_retention_total() + cert.previous_retention_total
            material_on_site += (
                cert.get_materials_on_site_total()
                + cert.previous_materials_on_site_total
            )

            # Special items (other)
            totals_by_type = cert.get_special_item_totals_by_type()
            prev_totals_by_type = cert.previous_special_item_totals_by_type
            other_current = totals_by_type.get("OTHER", Decimal("0.00"))
            other_prev = prev_totals_by_type.get("OTHER", Decimal("0.00"))
            other_specify += other_current + other_prev

    sub_total_contract = orig_val + amends_val
    # If any project has VAT enabled, let's include it in contract/progressive calculations
    # To be precise, let's calculate VAT per project and sum it up
    for project in projects:
        sub = (project.original_contract_value or Decimal("0.00")) + (
            project.addendum_contract_value or Decimal("0.00")
        )
        if project.vat:
            vat_val_contract += sub * Decimal("0.15")
    total_contract = sub_total_contract + vat_val_contract

    # Progressive / Certified Totals
    progressive_to_date = (
        work_progressive_to_date
        + advance_payment
        + retention
        + material_on_site
        + other_specify
    )
    progressive_previous = Decimal("0.00")
    # Sum progressive_previous directly from the latest certificates
    for project in projects:
        cert = (
            project.payment_certificates.filter(status="APPROVED")
            .order_by("-certificate_number")
            .first()
        )
        if cert:
            ap_prev = cert.previous_advance_payment_total
            ret_prev = cert.previous_retention_total
            mat_prev = cert.previous_materials_on_site_total
            prev_totals_by_type = cert.previous_special_item_totals_by_type
            other_prev = prev_totals_by_type.get("OTHER", Decimal("0.00"))

            progressive_previous += (
                (cert.progressive_previous or Decimal("0.00"))
                + ap_prev
                + ret_prev
                + mat_prev
                + other_prev
            )

    current_claim_total = progressive_to_date - progressive_previous

    # Calculate VAT on payment per project and sum
    vat_val_payment = Decimal("0.00")
    for project in projects:
        cert = (
            project.payment_certificates.filter(status="APPROVED")
            .order_by("-certificate_number")
            .first()
        )
        if cert and project.vat:
            p_prog_to_date = (
                (cert.work_progressive_to_date or Decimal("0.00"))
                + (
                    cert.get_advance_payment_total()
                    + cert.previous_advance_payment_total
                )
                + (cert.get_retention_total() + cert.previous_retention_total)
                + (
                    cert.get_materials_on_site_total()
                    + cert.previous_materials_on_site_total
                )
            )
            p_prev = (
                (cert.work_progressive_previous or Decimal("0.00"))
                + cert.previous_advance_payment_total
                + cert.previous_retention_total
                + cert.previous_materials_on_site_total
            )
            p_current = p_prog_to_date - p_prev
            vat_val_payment += p_current * Decimal("0.15")

    total_certified = current_claim_total + vat_val_payment

    # Precise Colors and Styles
    c_black = "FF111111"
    c_grey_text = "FF717171"
    c_white = "FFFFFFFF"

    font_bold = Font(bold=True, color=c_black)
    font_normal = Font(bold=False, color=c_black)
    font_title = Font(bold=True, size=14, color=c_black)
    font_white_bold = Font(bold=True, color=c_white)
    _font_italic_small = Font(italic=True, size=9, color=c_grey_text)

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    fill_light_grey = PatternFill(
        start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid"
    )
    fill_white = PatternFill(
        start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid"
    )
    fill_dark_blue = PatternFill(
        start_color="FF1B3A6B", end_color="FF1B3A6B", fill_type="solid"
    )
    fill_gold = PatternFill(
        start_color="FFC8963E", end_color="FFC8963E", fill_type="solid"
    )
    fill_separator = PatternFill(
        start_color="FF003300", end_color="FF003300", fill_type="solid"
    )

    border_thin_top_bottom = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    border_thin_bottom = Border(bottom=Side(style="thin"))

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 25

    # Row Heights
    ws.row_dimensions[1].height = 55.5
    ws.row_dimensions[2].height = 13.5
    ws.row_dimensions[3].height = 3.0

    # Initialize all cells to white fill
    for row in range(1, 60):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill_white
            ws.cell(row=row, column=col).font = font_normal

    # Row 1: Logo & Title
    ws.cell(row=1, column=1, value="[ LOGO ]").font = font_bold
    ws.cell(row=1, column=1).alignment = align_center

    title_cell = ws.cell(row=1, column=3, value="PROVINCE COVER PAGE")
    title_cell.font = font_title
    title_cell.alignment = align_center
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)

    today_str = datetime.date.today().strftime("%d %b %Y")
    cert_cell = ws.cell(row=1, column=7, value=f"As of Date:\n{today_str}")
    cert_cell.alignment = Alignment(
        wrap_text=True, horizontal="right", vertical="center"
    )
    cert_cell.font = Font(color=c_grey_text)

    # Row 1 Borders
    for col in range(1, 8):
        ws.cell(row=1, column=col).border = border_thin_top_bottom
        ws.cell(row=1, column=col).fill = fill_light_grey

    # Row 2: Province Info
    province_info = f"Province: {province.name} | Total Projects: {project_count}"
    ws.cell(row=2, column=1, value=province_info).font = Font(
        italic=True, color=c_grey_text
    )
    ws.cell(row=2, column=1).alignment = align_center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    # Row 2 Borders
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = border_thin_bottom
        ws.cell(row=2, column=col).fill = fill_light_grey

    # Row 3: Thick separator line
    for col in range(1, 8):
        ws.cell(row=3, column=col).fill = fill_separator

    # Dynamic Info Rows starting at Row 5
    current_row = 5

    # Section A info
    info_rows = [
        ("Province Name", province.name),
        ("Total Active Projects", f"{project_count}"),
        ("Projects with Approved Certificates", f"{cert_count} of {project_count}"),
    ]
    for label, val in info_rows:
        ws.row_dimensions[current_row].height = 18.0
        ws.cell(row=current_row, column=1, value=label).font = Font(color=c_grey_text)
        ws.cell(row=current_row, column=1).alignment = align_right
        ws.cell(row=current_row, column=2, value=val).font = font_bold
        current_row += 1

    current_row += 1  # Leave an empty row

    # Section B Title
    ws.row_dimensions[current_row].height = 18.0
    ws.cell(row=current_row, column=1, value="CONTRACT VALUE").font = font_bold
    for col in range(1, 8):
        ws.cell(row=current_row, column=col).fill = fill_light_grey
        ws.cell(row=current_row, column=col).border = border_thin_top_bottom
    ws.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=7
    )
    current_row += 1

    # Section B Columns
    ws.row_dimensions[current_row].height = 15.75
    ws.cell(row=current_row, column=1, value="Description").font = Font(
        color=c_grey_text
    )
    ws.cell(row=current_row, column=7, value="Value (R)").font = Font(color=c_grey_text)
    ws.cell(row=current_row, column=7).alignment = align_right
    current_row += 1

    # Section B Values
    b_rows = [
        ("Original Contract Value", orig_val, False),
        ("Approved Variations", amends_val, False),
        ("Sub-Total Contract Value", sub_total_contract, True),
        ("VAT (15%)", vat_val_contract, False),
        ("Total Contract Value", total_contract, True),
    ]

    for label, val, is_bold in b_rows:
        ws.row_dimensions[current_row].height = 18.0
        c1 = ws.cell(row=current_row, column=1, value=label)
        c2 = ws.cell(row=current_row, column=7, value=val)
        c2.number_format = "R #,##0.00;[Red]-R #,##0.00;R 0.00"
        c2.alignment = align_right

        if is_bold:
            c1.font = font_bold
            c2.font = font_bold
            if label == "Total Contract Value":
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).fill = fill_dark_blue
                c1.font = font_white_bold
                c2.font = font_white_bold
        current_row += 1

    current_row += 1  # Leave an empty row

    # Section C Title
    ws.row_dimensions[current_row].height = 18.0
    ws.cell(row=current_row, column=1, value="PROGRESSIVE VALUATIONS").font = font_bold
    for col in range(1, 8):
        ws.cell(row=current_row, column=col).fill = fill_light_grey
        ws.cell(row=current_row, column=col).border = border_thin_top_bottom
    ws.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=7
    )
    current_row += 1

    # Section C Columns
    ws.row_dimensions[current_row].height = 15.75
    ws.cell(row=current_row, column=1, value="Description").font = Font(
        color=c_grey_text
    )
    ws.cell(row=current_row, column=7, value="Value (R)").font = Font(color=c_grey_text)
    ws.cell(row=current_row, column=7).alignment = align_right
    current_row += 1

    c_rows = [
        ("Work Progressive Previous", work_progressive_previous, False),
        ("Contract Current Claim", contract_current_claim_total, False),
        ("Addendum Current Claim", addendum_current_claim_total, False),
        ("Work Progressive to Date", work_progressive_to_date, True),
        (
            "Less: Advance Payment",
            -advance_payment if advance_payment > 0 else advance_payment,
            False,
        ),
        ("Less: Retention", -retention if retention > 0 else retention, False),
        ("Materials on Site", material_on_site, False),
        ("Other Specify", other_specify, False),
        ("Progressive to Date", progressive_to_date, True),
        ("Progressive Previous", progressive_previous, False),
        ("Current Claim Total", current_claim_total, True),
        ("VAT (15%)", vat_val_payment, False),
        ("Total Certified (Gross Payment)", total_certified, True),
    ]

    for label, val, is_bold in c_rows:
        ws.row_dimensions[current_row].height = 18.0
        c1 = ws.cell(row=current_row, column=1, value=label)
        c2 = ws.cell(row=current_row, column=7, value=val)
        c2.number_format = "R #,##0.00;[Red]-R #,##0.00;R 0.00"
        c2.alignment = align_right

        if is_bold:
            c1.font = font_bold
            c2.font = font_bold
            if label == "Total Certified (Gross Payment)":
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).fill = fill_gold
                c1.font = font_white_bold
                c2.font = font_white_bold
        current_row += 1

    return wb
