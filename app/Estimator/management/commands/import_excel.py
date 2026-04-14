from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand

from app.Estimator.importers import (
    PlantCostImporter,
    PlantSpecImporter,
    PreliminaryCostImporter,
    PreliminarySpecImporter,
)
from app.Estimator.models import (
    BOQItem,
    ProjectLabourCrew,
    ProjectLabourSpecification,
    ProjectMaterial,
    ProjectSpecification,
    ProjectSpecificationComponent,
    ProjectTradeCode,
    SystemLabourCrew,
    SystemLabourSpecification,
    SystemMaterial,
    SystemSpecification,
    SystemSpecificationComponent,
    SystemTradeCode,
)


class ExcelImporter:
    """Reusable importer for Resources Estimator Excel files.

    Can be called from the management command or from a web view.
    When project is set, imports into Project* models. Otherwise imports into System* models.
    """

    def __init__(self, file_path, output=None, project=None):
        self.file_path = file_path
        self.output = output
        self.project = project
        self.results = {}
        self.sheet_names = []

    def log(self, msg):
        if self.output:
            self.output.write(msg + "\n")

    def _find_sheet(self, wb, keywords):
        """Find a worksheet by fuzzy keyword matching. Returns None if not found."""
        for name in wb.sheetnames:
            normalised = name.lower().strip()
            for kw in keywords:
                if kw in normalised:
                    return wb[name]
        return None

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        self.sheet_names = list(wb.sheetnames)
        self.log(f"  Sheets found: {wb.sheetnames}")

        ws = self._find_sheet(wb, ["trade code", "tradecode"])
        if ws:
            self.results["trade_codes"] = self._import_trade_codes(ws)

        ws = self._find_sheet(
            wb, ["materials cost", "material cost", "materials code", "material code"]
        )
        if ws:
            self.results["materials"] = self._import_materials(ws)

        ws = self._find_sheet(
            wb,
            [
                "material specification",
                "material specifications",
                "specification code",
                "material spec",
            ],
        )
        if ws:
            self.results["specifications"] = self._import_material_specifications(ws)

        ws = self._find_sheet(wb, ["labour cost", "labour costs", "labor cost"])
        if ws:
            self.results["labour_crews"] = self._import_labour_costs(ws)

        ws = self._find_sheet(
            wb, ["labour specification", "labour spec", "labor specification"]
        )
        if ws:
            self.results["labour_specs"] = self._import_labour_specifications(ws)

        # Plant/preliminary sheets are delegated to the standalone importer
        # classes which handle both the master hierarchical layout and the
        # downloaded-template flat layout.
        self.results["plant_costs"] = self._run_sheet_importer(
            PlantCostImporter, "Plant Costs"
        )
        self.results["plant_specs"] = self._run_sheet_importer(
            PlantSpecImporter, "Plant Specifications"
        )
        self.results["preliminary_costs"] = self._run_sheet_importer(
            PreliminaryCostImporter, "Preliminary Costs"
        )
        self.results["preliminary_specs"] = self._run_sheet_importer(
            PreliminarySpecImporter, "Preliminary Specifications"
        )

        ws = self._find_sheet(wb, ["output boq", "output bill"])
        if ws:
            self.results["boq_items"] = self._import_boq_items(ws)

        wb.close()
        return self.results

    def _safe_decimal(self, value):
        if value is None:
            return None
        if isinstance(value, str) and (value.startswith("#") or not value.strip()):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    def _safe_str(self, value):
        if value is None:
            return ""
        s = str(value).strip()
        if s.startswith("#"):
            return ""
        return s

    # ── Trade Codes ────────────────────────────────────────────

    def _import_trade_codes(self, ws):
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (row + (None,) * 4)[:4]
            prefix, trade_name, _, _ = vals
            if not prefix:
                continue
            prefix_str = str(prefix).strip()
            if self.project:
                ProjectTradeCode.objects.update_or_create(
                    project=self.project,
                    prefix=prefix_str,
                    defaults={"trade_name": str(trade_name or "").strip()},
                )
            else:
                SystemTradeCode.objects.update_or_create(
                    prefix=prefix_str,
                    defaults={"trade_name": str(trade_name or "").strip()},
                )
            count += 1
        self.log(f"  Trade Codes: {count}")
        return count

    # ── Materials Cost ─────────────────────────────────────────

    def _import_materials(self, ws):
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (row + (None,) * 6)[:6]
            trade_name, mat_code, unit, rate, variety, spec = vals
            if not mat_code:
                continue

            mat_code_str = self._safe_str(mat_code)
            defaults = {
                "trade_name": self._safe_str(trade_name),
                "unit": self._safe_str(unit),
                "market_rate": self._safe_decimal(rate) or Decimal("0"),
                "material_variety": self._safe_str(variety),
                "market_spec": self._safe_str(spec),
            }

            if self.project:
                ProjectMaterial.objects.update_or_create(
                    project=self.project, material_code=mat_code_str, defaults=defaults
                )
            else:
                SystemMaterial.objects.update_or_create(
                    material_code=mat_code_str, defaults=defaults
                )
            count += 1
        self.log(f"  Materials: {count}")
        return count

    # ── Material Specification ─────────────────────────────────

    def _find_material(self, code_str):
        if not code_str:
            return None
        code_str = self._safe_str(code_str)
        if not code_str:
            return None
        if self.project:
            return ProjectMaterial.objects.filter(
                project=self.project, material_code=code_str
            ).first()
        return SystemMaterial.objects.filter(material_code=code_str).first()

    def _find_trade_code(self, trade_code_str):
        if not trade_code_str:
            return None
        trade_code_str = str(trade_code_str).strip()
        if self.project:
            for tc in ProjectTradeCode.objects.filter(project=self.project):
                if tc.trade_code == trade_code_str:
                    return tc
            return None
        for tc in SystemTradeCode.objects.all():
            if tc.trade_code == trade_code_str:
                return tc
        return None

    def _import_material_specifications(self, ws):
        spec_model = ProjectSpecification if self.project else SystemSpecification
        comp_model = (
            ProjectSpecificationComponent
            if self.project
            else SystemSpecificationComponent
        )

        if self.project:
            spec_model.objects.filter(project=self.project).delete()
        else:
            spec_model.objects.all().delete()
        count = 0

        for row in ws.iter_rows(min_row=4, values_only=True):
            vals = (row + (None,) * 22)[:22]
            section = vals[0]
            trade_code_str = vals[1]
            boq_unit = vals[2]
            spec_code = vals[3]

            if not spec_code:
                continue

            trade = self._find_trade_code(trade_code_str)

            create_kwargs = {
                "section": self._safe_str(section),
                "trade_code": trade,
                "unit_label": self._safe_str(boq_unit) or "m3",
                "name": self._safe_str(spec_code),
            }
            if self.project:
                create_kwargs["project"] = self.project
            else:
                create_kwargs["boq_quantity"] = self._safe_decimal(vals[17]) or Decimal(
                    "0"
                )

            spec = spec_model.objects.create(**create_kwargs)

            for i in range(4):
                mat_code = vals[4 + i]
                qty_val = vals[8 + i]

                if not mat_code:
                    continue

                mat = self._find_material(mat_code)
                comp_model.objects.create(
                    specification=spec,
                    material=mat,
                    label=self._safe_str(mat_code),
                    qty_per_unit=self._safe_decimal(qty_val) or Decimal("0"),
                    sort_order=i,
                )

            count += 1

        self.log(f"  Material Specifications: {count}")
        return count

    # ── Labour Costs ───────────────────────────────────────────

    def _import_labour_costs(self, ws):
        model = ProjectLabourCrew if self.project else SystemLabourCrew
        if self.project:
            model.objects.filter(project=self.project).delete()
        else:
            model.objects.all().delete()
        count = 0

        for row in ws.iter_rows(min_row=3, values_only=True):
            vals = (row + (None,) * 11)[:11]
            crew_type = vals[1]
            if not crew_type:
                continue

            create_kwargs = {
                "crew_type": self._safe_str(crew_type),
                "crew_size": int(vals[2] or 0),
                "skilled": int(vals[3] or 0),
                "semi_skilled": int(vals[4] or 0),
                "general": int(vals[5] or 0),
                "daily_production": self._safe_decimal(vals[6]) or Decimal("0"),
                "skilled_rate": self._safe_decimal(vals[8]) or Decimal("0"),
                "semi_skilled_rate": self._safe_decimal(vals[9]) or Decimal("0"),
                "general_rate": self._safe_decimal(vals[10]) or Decimal("0"),
            }
            if self.project:
                create_kwargs["project"] = self.project
            model.objects.create(**create_kwargs)
            count += 1

        self.log(f"  Labour Crews: {count}")
        return count

    # ── Labour Specification ───────────────────────────────────

    def _import_labour_specifications(self, ws):
        model = (
            ProjectLabourSpecification if self.project else SystemLabourSpecification
        )
        crew_model = ProjectLabourCrew if self.project else SystemLabourCrew
        if self.project:
            model.objects.filter(project=self.project).delete()
        else:
            model.objects.all().delete()
        count = 0

        for row in ws.iter_rows(min_row=3, values_only=True):
            vals = (row + (None,) * 15)[:15]
            labour_spec = vals[2]
            if not labour_spec:
                continue

            crew_type_str = self._safe_str(vals[4])
            if self.project:
                crew = (
                    crew_model.objects.filter(
                        project=self.project, crew_type=crew_type_str
                    ).first()
                    if crew_type_str
                    else None
                )
            else:
                crew = (
                    crew_model.objects.filter(crew_type=crew_type_str).first()
                    if crew_type_str
                    else None
                )

            create_kwargs = {
                "section": self._safe_str(vals[0]),
                "trade_name": self._safe_str(vals[1]),
                "name": self._safe_str(labour_spec),
                "unit": self._safe_str(vals[3]),
                "crew": crew,
                "daily_production": self._safe_decimal(vals[5]) or Decimal("0"),
                "team_mix": self._safe_decimal(vals[6]) or Decimal("1"),
                "site_factor": self._safe_decimal(vals[7]) or Decimal("1"),
                "tools_factor": self._safe_decimal(vals[8]) or Decimal("1"),
                "leadership_factor": self._safe_decimal(vals[9]) or Decimal("1"),
            }
            if self.project:
                create_kwargs["project"] = self.project
            else:
                create_kwargs["boq_quantity"] = self._safe_decimal(vals[13]) or Decimal(
                    "0"
                )
            model.objects.create(**create_kwargs)
            count += 1

        self.log(f"  Labour Specifications: {count}")
        return count

    # ── Plant / Preliminary delegation ─────────────────────────

    def _run_sheet_importer(self, importer_cls, label):
        """Run a standalone importer and log the result.

        Returns the created+updated total to match the legacy return shape.
        """
        importer = importer_cls(self.file_path, project=self.project)
        result = importer.run()
        created = result.get("created", 0)
        updated = result.get("updated", 0)
        sheet_used = result.get("sheet_used")
        suffix = f" [sheet: {sheet_used}]" if sheet_used else ""
        self.log(f"  {label}: {created} created, {updated} updated{suffix}")
        return created + updated

    # ── Output BoQ ─────────────────────────────────────────────

    def _find_specification(self, spec_name_str):
        if not spec_name_str:
            return None
        if self.project:
            return ProjectSpecification.objects.filter(
                project=self.project, name=spec_name_str
            ).first()
        return SystemSpecification.objects.filter(name=spec_name_str).first()

    def _find_labour_specification(self, name_str):
        if not name_str:
            return None
        if self.project:
            return ProjectLabourSpecification.objects.filter(
                project=self.project, name=name_str
            ).first()
        return SystemLabourSpecification.objects.filter(name=name_str).first()

    def _import_boq_items(self, ws):
        # BOQItem is a single model with a nullable project FK — scope the
        # wipe so a system-library import never touches project-scoped rows.
        if self.project:
            BOQItem.objects.filter(project=self.project).delete()
        else:
            BOQItem.objects.filter(project__isnull=True).delete()
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (row + (None,) * 21)[:21]
            section = vals[0]
            desc = vals[8]

            if not section and not desc:
                continue

            is_header = bool(desc and not vals[10] and not vals[11])

            trade = self._find_trade_code(vals[2])

            spec_str = self._safe_str(vals[4])
            spec = self._find_specification(spec_str)

            labour_str = self._safe_str(vals[5])
            labour_spec = self._find_labour_specification(labour_str)

            material = None
            if not spec and spec_str:
                material = self._find_material(spec_str)

            create_kwargs = {
                "section": self._safe_str(section),
                "bill_no": self._safe_str(vals[1]),
                "trade_code": trade,
                "specification": spec,
                "labour_specification": labour_spec,
                "material": material,
                "item_no": self._safe_str(vals[6]),
                "pay_ref": self._safe_str(vals[7]),
                "description": self._safe_str(desc),
                "unit": self._safe_str(vals[9]),
                "contract_quantity": self._safe_decimal(vals[10]),
                "contract_rate": self._safe_decimal(vals[11]),
                "progress_quantity": self._safe_decimal(vals[13]),
                "forecast_quantity": self._safe_decimal(vals[14]),
                "is_section_header": is_header,
            }
            if self.project:
                create_kwargs["project"] = self.project
            BOQItem.objects.create(**create_kwargs)
            count += 1
        self.log(f"  BoQ Items: {count}")
        return count


class SystemSpecImporter:
    """Importer for System Material Specification library."""

    def __init__(self, file_path, output=None):
        self.file_path = file_path
        self.output = output
        self.results = {}

    def log(self, msg):
        if self.output:
            self.output.write(msg + "\n")

    def _safe_str(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def _safe_decimal(self, value):
        if value is None:
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    def run(self):
        from app.Estimator.models import (
            SystemMaterial,
            SystemMaterialSpec,
            SystemMaterialSpecComponent,
        )

        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = wb.active

        specs_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (row + (None,) * 4)[:4]
            spec_name, unit, mat_code, qty = vals

            spec_name = self._safe_str(spec_name)
            unit = self._safe_str(unit)

            if not spec_name:
                continue

            key = (spec_name, unit or "m3")
            if key not in specs_data:
                specs_data[key] = []

            mat_code_str = self._safe_str(mat_code)
            if mat_code_str:
                specs_data[key].append(
                    {
                        "material_code": mat_code_str,
                        "qty_per_unit": self._safe_decimal(qty) or Decimal("0"),
                    }
                )

        created = 0
        updated = 0

        for (name, unit), components in specs_data.items():
            spec, was_created = SystemMaterialSpec.objects.get_or_create(
                name=name, defaults={"unit": unit}
            )
            if not was_created:
                spec.unit = unit
                spec.save()
                spec.system_spec_components.all().delete()
                updated += 1
            else:
                created += 1

            for i, comp in enumerate(components):
                mat = SystemMaterial.objects.filter(
                    material_code=comp["material_code"]
                ).first()
                SystemMaterialSpecComponent.objects.create(
                    spec=spec,
                    material=mat,
                    label=comp["material_code"],
                    qty_per_unit=comp["qty_per_unit"],
                    sort_order=i,
                )

        self.results["system_specs"] = {"created": created, "updated": updated}
        wb.close()
        return self.results


class Command(BaseCommand):
    help = "Import data from Resources Estimator Excel file"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to the Excel file")

    def handle(self, *args, **options):
        importer = ExcelImporter(options["file"], output=self.stdout)
        self.stdout.write("Importing full Resources Estimator Excel...")
        importer.run()
        self.stdout.write(self.style.SUCCESS("Import completed successfully."))
