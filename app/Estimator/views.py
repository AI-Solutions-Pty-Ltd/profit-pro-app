import json
import os
import tempfile
from decimal import Decimal
from pathlib import Path

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db import models
from django.db.models import DecimalField, F, Sum, Value
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import FormView, ListView, TemplateView
from django.views.generic.base import ContextMixin

from app.Account.models import Municipality, Province
from app.BillOfQuantities.models.structure_models import LineItem
from app.Project.models import Project

from .calculations import (
    calculate_pct_of_total,
    calculate_variance,
    format_num,
)
from .forms import (
    ContractorLabourCrewForm,
    ContractorLabourSpecificationForm,
    ContractorMaterialForm,
    ContractorPlantCostForm,
    ContractorPlantSpecificationForm,
    ContractorPreliminaryCostForm,
    ContractorPreliminarySpecificationForm,
    ContractorSpecificationComponentFormSet,
    ContractorSpecificationForm,
    ContractorTradeCodeForm,
    ExcelImportForm,
    LabourCrewForm,
    LabourSpecificationForm,
    MaterialForm,
    PlantCostForm,
    PlantSpecificationForm,
    PreliminaryCostForm,
    PreliminarySpecificationForm,
    ProjectAssumptionsForm,
    SpecificationComponentFormSet,
    SpecificationForm,
    SystemLabourCrewForm,
    SystemLabourSpecificationForm,
    SystemMaterialForm,
    SystemMunicipalityForm,
    SystemPlantCostForm,
    SystemPlantSpecificationForm,
    SystemPreliminaryCostForm,
    SystemPreliminarySpecificationForm,
    SystemProvinceForm,
    SystemSpecificationComponentFormSet,
    SystemSpecificationForm,
    SystemTradeCodeForm,
)
from .models import (
    BOQItem,
    ContractorItemLibraryEntry,
    ContractorLabourCrew,
    ContractorLabourSpecification,
    ContractorMaterial,
    ContractorPlantCost,
    ContractorPlantSpecification,
    ContractorPlantSpecificationComponent,
    ContractorPreliminaryCost,
    ContractorPreliminarySpecification,
    ContractorSpecification,
    ContractorSpecificationComponent,
    ContractorTradeCode,
    ProjectAssumptions,
    ProjectItemLibraryEntry,
    ProjectLabourCrew,
    ProjectLabourSpecification,
    ProjectMaterial,
    ProjectPlantCost,
    ProjectPlantSpecification,
    ProjectPlantSpecificationComponent,
    ProjectPreliminaryCost,
    ProjectPreliminarySpecification,
    ProjectSpecification,
    ProjectSpecificationComponent,
    ProjectTradeCode,
    SystemItemLibraryEntry,
    SystemLabourCrew,
    SystemLabourSpecification,
    SystemMaterial,
    SystemPlantCost,
    SystemPlantSpecification,
    SystemPlantSpecificationComponent,
    SystemPreliminaryCost,
    SystemPreliminarySpecification,
    SystemSpecification,
    SystemSpecificationComponent,
    SystemTradeCode,
    sync_boq_from_lineitems,
)


class ProjectEstimatorMixin(ContextMixin):
    """Mixin that loads the project from URL kwargs and adds it to context."""

    def get_project(self):
        if not hasattr(self, "_project"):
            kwargs = getattr(self, "kwargs", {})
            self._project = get_object_or_404(Project, pk=kwargs.get("project_pk"))
        return self._project

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["project"] = self.get_project()
        context["project_pk"] = getattr(self, "kwargs", {}).get("project_pk")
        return context


def _pagination_query_params(request):
    """Return GET params (excluding `page`) urlencoded for pagination links."""
    params = request.GET.copy()
    params.pop("page", None)
    return params.urlencode()


def _spec_datalist_context(qs, unit_field="unit"):
    """Distinct existing values for the section / trade_name / unit datalists.

    `qs` should be a queryset already scoped to the relevant tier (project,
    contractor, or system). `unit_field` is "unit_label" for material specs
    and "unit" for labour/plant/prelim specs.
    """

    def _values(field):
        if not hasattr(qs.model, field):
            return []
        return list(
            qs.exclude(**{field: ""})
            .values_list(field, flat=True)
            .distinct()
            .order_by(field)
        )

    return {
        "dl_sections": _values("section"),
        "dl_trade_names": _values("trade_name"),
        "dl_units": _values(unit_field),
    }


def _handle_clear_action(request, queryset, *, label="entries"):
    """Handle a `clear_all` POST. Deletes everything in `queryset`.

    Returns True if the request was a clear action (caller should redirect),
    False otherwise.
    """
    if request.POST.get("action") != "clear_all":
        return False
    count = queryset.count()
    queryset.delete()
    messages.success(request, f"Cleared {count} {label}.")
    return True


def _handle_bulk_action(request, queryset, *, allow_toggle_active=False):
    """Handle bulk_delete / bulk_activate / bulk_deactivate POSTs.

    Returns True if the request was a bulk action (caller should redirect),
    False otherwise.
    """
    action = request.POST.get("action", "")
    if not action.startswith("bulk_"):
        return False
    ids = [i for i in request.POST.getlist("ids") if i]
    if not ids:
        messages.warning(request, "No items selected.")
        return True
    qs = queryset.filter(pk__in=ids)
    if action == "bulk_delete":
        count = qs.count()
        qs.delete()
        messages.success(request, f"Deleted {count} item(s).")
        return True
    if allow_toggle_active and action == "bulk_activate":
        count = qs.update(is_active=True)
        messages.success(request, f"Activated {count} item(s).")
        return True
    if allow_toggle_active and action == "bulk_deactivate":
        count = qs.update(is_active=False)
        messages.success(request, f"Deactivated {count} item(s).")
        return True
    messages.error(request, f"Unknown bulk action: {action}")
    return True


def _flash_sync_result(request, result, entity_label):
    """Flash a success/error message for a project-side contractor sync."""
    if result.get("skipped_no_contractor"):
        messages.error(
            request,
            f"Cannot sync {entity_label}: this project has no contractor "
            "assigned. Set one in project setup.",
        )
        return
    messages.success(
        request,
        f"{entity_label} synced from contractor library — "
        f"{result['updated']} updated, {result['created']} new.",
    )


class ProjectAssumptionsView(ProjectEstimatorMixin, TemplateView):
    template_name = "estimator/project_assumptions.html"

    def get_assumptions(self):
        obj, _ = ProjectAssumptions.objects.get_or_create(project=self.get_project())
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ProjectAssumptionsForm(instance=self.get_assumptions())
        return context

    def post(self, request, *args, **kwargs):
        assumptions = self.get_assumptions()
        old_values = {
            "material_markup_pct": assumptions.material_markup_pct,
            "labour_markup_pct": assumptions.labour_markup_pct,
            "transport_pct": assumptions.transport_pct,
        }
        form = ProjectAssumptionsForm(request.POST, instance=assumptions)
        if form.is_valid():
            form.save()
            new = form.instance
            project = self.get_project()
            propagated = 0
            for field in ("material_markup_pct", "labour_markup_pct", "transport_pct"):
                old_value = old_values[field]
                new_value = getattr(new, field)
                if new_value == old_value:
                    continue
                propagated += BOQItem.objects.filter(
                    project=project, **{field: old_value}
                ).update(**{field: new_value})
            msg = "Project assumptions saved."
            if propagated:
                msg += (
                    f" Propagated to {propagated} default markup field"
                    f"{'s' if propagated != 1 else ''} on BoQ items "
                    "(rows with manually edited markups left unchanged)."
                )
            messages.success(request, msg)
            return redirect(
                "estimator:project_assumptions", project_pk=self.kwargs["project_pk"]
            )
        context = self.get_context_data()
        context["form"] = form
        return self.render_to_response(context)


class ApplyAssumptionsView(ProjectEstimatorMixin, View):
    """Push current project assumption markups to all BOQItems in the project."""

    def post(self, request, project_pk):
        project = self.get_project()
        assumptions = ProjectAssumptions.objects.filter(project=project).first()
        if not assumptions:
            messages.error(request, "No project assumptions to apply.")
            return redirect("estimator:project_assumptions", project_pk=project_pk)
        updated = BOQItem.objects.filter(project=project).update(
            material_markup_pct=assumptions.material_markup_pct,
            labour_markup_pct=assumptions.labour_markup_pct,
            transport_pct=assumptions.transport_pct,
        )
        messages.success(
            request,
            f"Applied assumptions to {updated} BoQ items.",
        )
        return redirect("estimator:project_assumptions", project_pk=project_pk)


class DashboardView(ProjectEstimatorMixin, ListView):
    model = BOQItem
    template_name = "estimator/dashboard.html"
    context_object_name = "items"
    paginate_by = 100

    def get_queryset(self):
        qs = (
            BOQItem.objects.filter(project=self.get_project())
            .select_related(
                "trade_code",
                "specification",
                "labour_specification",
                "labour_specification__crew",
                "plant_specification",
                "preliminary_specification",
                "material",
                "library_entry",
                "project__estimator_assumptions",
            )
            .prefetch_related(
                "specification__spec_components__material",
            )
        )

        # Apply filters from query params
        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)

        bill_no = self.request.GET.get("bill_no")
        if bill_no:
            qs = qs.filter(bill_no=bill_no)

        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)

        mat_spec = self.request.GET.get("mat_spec")
        if mat_spec == "none":
            qs = qs.filter(specification__isnull=True, material__isnull=True)
        elif mat_spec == "has_spec":
            qs = qs.filter(specification__isnull=False)
        elif mat_spec == "has_material":
            qs = qs.filter(material__isnull=False)
        elif mat_spec:
            qs = qs.filter(specification__name=mat_spec)

        lab_plant_spec = self.request.GET.get("lab_plant_spec")
        if lab_plant_spec == "none":
            qs = qs.filter(
                labour_specification__isnull=True, plant_specification__isnull=True
            )
        elif lab_plant_spec:
            qs = qs.filter(
                models.Q(labour_specification__name=lab_plant_spec)
                | models.Q(plant_specification__name=lab_plant_spec)
            )

        prelim_spec = self.request.GET.get("prelim_spec")
        if prelim_spec == "none":
            qs = qs.filter(preliminary_specification__isnull=True)
        elif prelim_spec:
            qs = qs.filter(preliminary_specification__name=prelim_spec)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()

        summary_items = (
            BOQItem.objects.filter(project=project, is_section_header=False)
            .select_related(
                "specification",
                "labour_specification",
                "labour_specification__crew",
                "plant_specification",
                "preliminary_specification",
                "material",
                "project__estimator_assumptions",
            )
            .prefetch_related("specification__spec_components__material")
        )

        total_contract = Decimal("0")
        total_material = Decimal("0")
        total_labour = Decimal("0")
        total_plant = Decimal("0")
        total_preliminary = Decimal("0")
        total_progress = Decimal("0")
        total_forecast = Decimal("0")
        total_markup = Decimal("0")
        total_transport = Decimal("0")

        for item in summary_items:
            if item.contract_amount:
                total_contract += item.contract_amount
            if item.new_materials_amount:
                total_material += item.new_materials_amount
            if item.new_labour_amount:
                total_labour += item.new_labour_amount
            if item.new_plant_amount:
                total_plant += item.new_plant_amount
            if item.new_preliminary_amount:
                total_preliminary += item.new_preliminary_amount
            if item.progress_amount:
                total_progress += item.progress_amount
            if item.forecast_amount:
                total_forecast += item.forecast_amount
            if item.markup_amount:
                total_markup += item.markup_amount
            if item.transport_amount:
                total_transport += item.transport_amount

        context["total_contract_amount"] = total_contract
        context["total_material_amount"] = total_material
        context["total_labour_amount"] = total_labour
        context["total_plant_amount"] = total_plant
        context["total_preliminary_amount"] = total_preliminary
        context["total_progress_amount"] = total_progress
        context["total_forecast_amount"] = total_forecast
        context["total_markup_amount"] = total_markup
        context["total_transport_amount"] = total_transport

        # Filter options (scoped to project)
        project_items = BOQItem.objects.filter(project=project)
        context["sections"] = (
            project_items.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["bill_nos"] = (
            project_items.exclude(bill_no="")
            .values_list("bill_no", flat=True)
            .distinct()
            .order_by("bill_no")
        )
        context["trade_codes"] = ProjectTradeCode.objects.filter(project=project)

        # Spec names for dropdowns: active specs + any inactive spec currently
        # linked on a BoQ row (so existing data still renders correctly).
        def _names_active_plus_linked(spec_model):
            active = set(
                spec_model.objects.filter(project=project, is_active=True)
                .exclude(name="")
                .values_list("name", flat=True)
            )
            linked_inactive = set(
                spec_model.objects.filter(
                    project=project, is_active=False, boq_items__isnull=False
                )
                .exclude(name="")
                .values_list("name", flat=True)
                .distinct()
            )
            return sorted(active | linked_inactive)

        context["spec_names"] = _names_active_plus_linked(ProjectSpecification)
        context["labour_spec_names"] = _names_active_plus_linked(
            ProjectLabourSpecification
        )
        context["plant_spec_names"] = _names_active_plus_linked(
            ProjectPlantSpecification
        )
        context["labour_plant_spec_names"] = sorted(
            set(context["labour_spec_names"]) | set(context["plant_spec_names"])
        )
        context["prelim_spec_names"] = _names_active_plus_linked(
            ProjectPreliminarySpecification
        )

        # Item Library entries for the per-row picker (only those with an
        # item_code set — others can't be picked by code).
        context["library_entries_for_picker"] = list(
            ProjectItemLibraryEntry.objects.filter(project=project)
            .exclude(item_code="")
            .order_by("item_code")
            .values("id", "item_code", "description")
        )

        # Current filter values
        context["f_section"] = self.request.GET.get("section", "")
        context["f_bill_no"] = self.request.GET.get("bill_no", "")
        context["f_trade_code"] = self.request.GET.get("trade_code", "")
        context["f_mat_spec"] = self.request.GET.get("mat_spec", "")
        context["f_lab_plant_spec"] = self.request.GET.get("lab_plant_spec", "")
        context["f_prelim_spec"] = self.request.GET.get("prelim_spec", "")

        # Project assumptions (wastage)
        assumptions, _ = ProjectAssumptions.objects.get_or_create(project=project)
        context["wastage_pct"] = assumptions.wastage_pct

        context["query_params"] = _pagination_query_params(self.request)

        return context


class BaselineBoqView(ProjectEstimatorMixin, ListView):
    """Reads baseline BoQ data from BillOfQuantities LineItem for the project."""

    model = LineItem
    template_name = "estimator/baseline_boq.html"
    context_object_name = "items"

    def get_queryset(self):
        project = self.get_project()
        qs = (
            LineItem.objects.filter(project=project)
            .select_related("bill", "bill__structure", "structure", "package")
            .order_by("row_index")
        )

        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(bill__name=section)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = context["items"]

        total_contract = Decimal("0")
        for item in items:
            if item.is_work and item.total_price:
                total_contract += item.total_price

        context["total_contract_amount"] = total_contract
        context["item_count"] = items.count() if hasattr(items, "count") else len(items)

        # Filter options
        project = self.get_project()
        context["sections"] = (
            LineItem.objects.filter(project=project)
            .exclude(bill__isnull=True)
            .values_list("bill__name", flat=True)
            .distinct()
            .order_by("bill__name")
        )

        context["f_section"] = self.request.GET.get("section", "")
        context["boq_item_count"] = BOQItem.objects.filter(project=project).count()

        return context


class SyncBoqView(ProjectEstimatorMixin, View):
    """Sync BOQItems from BillOfQuantities LineItems for this project."""

    def get_project(self):
        if not hasattr(self, "_project"):
            self._project = get_object_or_404(Project, pk=self.kwargs["project_pk"])
        return self._project

    def post(self, request, project_pk):
        project = self.get_project()
        created, updated, deleted = sync_boq_from_lineitems(project)
        messages.success(
            request,
            f"BoQ synced: {created} created, {updated} updated, {deleted} deleted",
        )
        return redirect(
            reverse("estimator:dashboard", kwargs={"project_pk": project_pk})
        )


class AutofillBoqFromLibraryView(ProjectEstimatorMixin, View):
    """Bulk-fill BoQ rows from matching Item Library entries.

    POST `mode=reset_and_rerun` first clears specs on every previously
    auto-filled row, then runs the fill. Any other value (or missing) only
    fills rows that currently have no specs.
    """

    def post(self, request, project_pk):
        from .services import (
            autofill_boq_from_library,
            reset_boq_autofill_trackers,
        )

        project = self.get_project()
        reset_count = 0
        if request.POST.get("mode") == "reset_and_rerun":
            reset_count = reset_boq_autofill_trackers(project)

        result = autofill_boq_from_library(project)
        parts = []
        if reset_count:
            parts.append(f"{reset_count} reset")
        parts.append(f"{result['filled']} filled")
        if result["skipped_already_set"]:
            parts.append(f"{result['skipped_already_set']} already had specs")
        if result["ambiguous"]:
            parts.append(f"{result['ambiguous']} ambiguous (skipped)")
        if result["no_match"]:
            parts.append(f"{result['no_match']} no match")
        messages.success(
            request, "Autofill from Item Library — " + ", ".join(parts) + "."
        )
        return redirect(
            reverse("estimator:dashboard", kwargs={"project_pk": project_pk})
        )


@method_decorator(csrf_exempt, name="dispatch")
class SaveBoqRowToLibraryView(View):
    """AJAX: upsert a single BoQ row's spec mapping into the project Item Library."""

    def post(self, request, project_pk, pk):
        from .services import save_boq_item_to_library

        item = get_object_or_404(BOQItem, pk=pk, project_id=project_pk)
        if item.is_section_header:
            return JsonResponse(
                {"error": "Section headers cannot be saved to library"}, status=400
            )
        try:
            data = json.loads(request.body) if request.body else {}
        except json.JSONDecodeError:
            data = {}
        result = save_boq_item_to_library(item, item_code=data.get("item_code"))
        return JsonResponse(
            {
                "ok": True,
                "created": result["created"],
                "entry_id": result["entry_id"],
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class ApplyLibraryEntryToBoqView(View):
    """AJAX: apply a ProjectItemLibraryEntry's spec FKs to a single BoQ row.

    Overwrites library_entry + the four spec FKs. Trade code, component,
    description, and unit on the row are left untouched.
    """

    def post(self, request, project_pk, pk):
        from .services import apply_library_entry_to_boq_item

        item = get_object_or_404(BOQItem, pk=pk, project_id=project_pk)
        if item.is_section_header:
            return JsonResponse(
                {"error": "Section headers cannot use library entries"}, status=400
            )
        try:
            data = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        entry_id = data.get("entry_id")
        if not entry_id:
            return JsonResponse({"error": "entry_id required"}, status=400)
        entry = get_object_or_404(
            ProjectItemLibraryEntry, pk=entry_id, project_id=project_pk
        )
        apply_library_entry_to_boq_item(item, entry)

        def fmt(val):
            return None if val is None else format_num(val)

        labour_plant_name = ""
        if item.labour_specification:
            labour_plant_name = item.labour_specification.name
        elif item.plant_specification:
            labour_plant_name = item.plant_specification.name

        return JsonResponse(
            {
                "ok": True,
                "entry_id": entry.id,
                "specs": {
                    "specification": item.specification.name
                    if item.specification
                    else "",
                    "labour_plant_specification": labour_plant_name,
                    "preliminary_specification": (
                        item.preliminary_specification.name
                        if item.preliminary_specification
                        else ""
                    ),
                },
                "new_materials_rate": fmt(item.new_materials_rate),
                "new_labour_rate": fmt(item.new_labour_rate),
                "new_plant_rate": fmt(item.new_plant_rate),
                "new_preliminary_rate": fmt(item.new_preliminary_rate),
                "baseline_new_price": fmt(item.baseline_new_price),
                "progress_amount": fmt(item.progress_amount),
                "forecast_amount": fmt(item.forecast_amount),
                "new_materials_amount": fmt(item.new_materials_amount),
                "new_labour_amount": fmt(item.new_labour_amount),
                "new_plant_amount": fmt(item.new_plant_amount),
                "new_preliminary_amount": fmt(item.new_preliminary_amount),
            }
        )


class SpecificationListView(ProjectEstimatorMixin, TemplateView):
    """Material calculator — driven by Output BoQ.

    Groups BOQItems by (section, specification) and sums contract_quantity
    per group. Looks up components and rates from the ProjectSpecification.
    """

    template_name = "estimator/specification_list.html"

    def _build_calculator_rows(
        self, project, section_filter=None, trade_code_filter=None, name_filter=None
    ):
        qs = (
            BOQItem.objects.filter(
                project=project,
                specification__isnull=False,
                specification__is_active=True,
                is_section_header=False,
            )
            .select_related(
                "specification__trade_code",
            )
            .prefetch_related(
                "specification__spec_components__material",
            )
        )

        if section_filter:
            qs = qs.filter(section=section_filter)
        if name_filter:
            qs = qs.filter(specification__name=name_filter)
        if trade_code_filter:
            qs = qs.filter(specification__trade_code_id=trade_code_filter)

        # Group by (section, specification_id)
        from collections import OrderedDict

        groups = OrderedDict()
        for boq in qs.order_by("section", "specification__name"):
            key = (boq.section, getattr(boq, "specification_id", None))
            if key not in groups:
                groups[key] = {
                    "section": boq.section,
                    "spec": boq.specification,
                    "boq_qty": Decimal("0"),
                }
            if boq.contract_quantity:
                groups[key]["boq_qty"] += boq.contract_quantity

        # Build rows with component totals
        rows = []
        for group in groups.values():
            spec = group["spec"]
            boq_qty = group["boq_qty"]
            component_totals = []
            for sc in spec.spec_components.all():
                component_totals.append(
                    {
                        "id": sc.id,
                        "label": sc.label,
                        "qty_per_unit": sc.qty_per_unit,
                        "total_quantity": boq_qty * sc.qty_per_unit
                        if boq_qty
                        else Decimal("0"),
                        "unit": sc.material.unit if sc.material else "",
                    }
                )
            rows.append(
                {
                    "section": group["section"],
                    "spec": spec,
                    "boq_qty": boq_qty,
                    "rate_per_unit": spec.rate_per_unit,
                    "component_totals": component_totals,
                }
            )
        return rows

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()

        f_section = self.request.GET.get("section", "")
        f_trade_code = self.request.GET.get("trade_code", "")
        f_name = self.request.GET.get("name", "")

        context["specs"] = self._build_calculator_rows(
            project,
            section_filter=f_section or None,
            trade_code_filter=f_trade_code or None,
            name_filter=f_name or None,
        )

        context["spec_form"] = context.get(
            "spec_form", SpecificationForm(project=project)
        )
        context["component_formset"] = context.get(
            "component_formset", SpecificationComponentFormSet()
        )

        # Filter options from Output BoQ
        boq_with_spec = BOQItem.objects.filter(
            project=project,
            specification__isnull=False,
            specification__is_active=True,
            is_section_header=False,
        )
        context["sections"] = (
            boq_with_spec.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_codes"] = ProjectTradeCode.objects.filter(
            pk__in=boq_with_spec.exclude(specification__trade_code__isnull=True)
            .values_list("specification__trade_code_id", flat=True)
            .distinct()
        )
        context["names"] = (
            boq_with_spec.values_list("specification__name", flat=True)
            .distinct()
            .order_by("specification__name")
        )

        context["f_section"] = f_section
        context["f_trade_code"] = f_trade_code
        context["f_name"] = f_name

        return context

    def post(self, request, *args, **kwargs):
        project = self.get_project()
        spec_form = SpecificationForm(request.POST, project=project)
        component_formset = SpecificationComponentFormSet(request.POST)
        if spec_form.is_valid():
            spec = spec_form.save(commit=False)
            spec.project = project
            spec.save()
            component_formset = SpecificationComponentFormSet(
                request.POST, instance=spec
            )
            if component_formset.is_valid():
                component_formset.save()
                return redirect(
                    reverse(
                        "estimator:specifications",
                        kwargs={"project_pk": self.kwargs["project_pk"]},
                    )
                )
            else:
                spec.delete()
        return self.render_to_response(
            self.get_context_data(
                spec_form=spec_form, component_formset=component_formset
            )
        )


class MaterialSpecListView(ProjectEstimatorMixin, ListView):
    """Simple material specifications view — Name, Trade Code, Unit, Components."""

    model = ProjectSpecification
    template_name = "estimator/material_spec_list.html"
    context_object_name = "specs"
    paginate_by = 50

    def get_queryset(self):
        project = self.get_project()
        qs = (
            ProjectSpecification.objects.filter(project=project)
            .select_related("trade_code")
            .prefetch_related("spec_components__material")
            .annotate(_baseline_boq_qty=models.Sum("boq_items__contract_quantity"))
        )

        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)

        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)

        name = self.request.GET.get("name")
        if name:
            qs = qs.filter(name=name)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()
        context["spec_form"] = context.get(
            "spec_form", SpecificationForm(project=project)
        )
        context["component_formset"] = context.get(
            "component_formset", SpecificationComponentFormSet()
        )

        project_specs = ProjectSpecification.objects.filter(project=project)
        context["sections"] = (
            project_specs.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_codes"] = ProjectTradeCode.objects.filter(project=project)
        context["names"] = (
            project_specs.values_list("name", flat=True).distinct().order_by("name")
        )
        context["materials"] = ProjectMaterial.objects.filter(project=project).order_by(
            "material_code"
        )

        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_code"] = self.request.GET.get("trade_code", "")
        context["f_name"] = self.request.GET.get("name", "")
        context["query_params"] = _pagination_query_params(self.request)

        context.update(
            _spec_datalist_context(
                ProjectSpecification.objects.filter(project=project),
                unit_field="unit_label",
            )
        )

        return context

    def post(self, request, *args, **kwargs):
        project = self.get_project()
        action = request.POST.get("action")
        qs = ProjectSpecification.objects.filter(project=project)
        if _handle_clear_action(request, qs, label="material specs"):
            return redirect(
                reverse(
                    "estimator:material_specs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if _handle_bulk_action(request, qs, allow_toggle_active=True):
            return redirect(
                reverse(
                    "estimator:material_specs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if action == "sync_system":
            from .services import sync_material_specs_from_contractor

            result = sync_material_specs_from_contractor(project)
            _flash_sync_result(request, result, "Material specs")
            return redirect(
                reverse(
                    "estimator:material_specs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        spec_form = SpecificationForm(request.POST, project=project)
        component_formset = SpecificationComponentFormSet(request.POST)
        if spec_form.is_valid():
            spec = spec_form.save(commit=False)
            spec.project = project
            spec.save()
            component_formset = SpecificationComponentFormSet(
                request.POST, instance=spec
            )
            if component_formset.is_valid():
                component_formset.save()
                return redirect(
                    reverse(
                        "estimator:material_specs",
                        kwargs={"project_pk": self.kwargs["project_pk"]},
                    )
                )
            else:
                spec.delete()
        self.object_list = self.get_queryset()
        return self.render_to_response(
            self.get_context_data(
                spec_form=spec_form, component_formset=component_formset
            )
        )


class MaterialsListView(ProjectEstimatorMixin, ListView):
    model = ProjectMaterial
    template_name = "estimator/materials_list.html"
    context_object_name = "materials"

    def get_queryset(self):
        qs = ProjectMaterial.objects.filter(project=self.get_project())
        unit = self.request.GET.get("unit", "").strip()
        if unit:
            qs = qs.filter(unit=unit)
        trade_name = self.request.GET.get("trade_name", "").strip()
        if trade_name:
            qs = qs.filter(trade_name=trade_name)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", MaterialForm(project=self.get_project()))
        project_materials = ProjectMaterial.objects.filter(project=self.get_project())
        context["units"] = (
            project_materials.exclude(unit="")
            .values_list("unit", flat=True)
            .distinct()
            .order_by("unit")
        )
        context["trade_names"] = (
            project_materials.exclude(trade_name="")
            .values_list("trade_name", flat=True)
            .distinct()
            .order_by("trade_name")
        )
        context["f_q"] = self.request.GET.get("q", "")
        context["f_unit"] = self.request.GET.get("unit", "")
        context["f_trade_name"] = self.request.GET.get("trade_name", "")
        return context

    def post(self, request, *args, **kwargs):
        project_pk = self.kwargs["project_pk"]
        action = request.POST.get("action")
        qs = ProjectMaterial.objects.filter(project=self.get_project())

        if _handle_clear_action(request, qs, label="materials"):
            return redirect(
                reverse("estimator:materials", kwargs={"project_pk": project_pk})
            )

        if _handle_bulk_action(request, qs):
            return redirect(
                reverse("estimator:materials", kwargs={"project_pk": project_pk})
            )

        if action == "sync_system":
            from .services import sync_materials_from_contractor

            result = sync_materials_from_contractor(self.get_project())
            _flash_sync_result(request, result, "Material costs")
            return redirect(
                reverse("estimator:materials", kwargs={"project_pk": project_pk})
            )

        # Default: add material form
        form = MaterialForm(request.POST, project=self.get_project())
        if form.is_valid():
            obj = form.save(commit=False)
            obj.project = self.get_project()
            obj.save()
            return redirect(
                reverse("estimator:materials", kwargs={"project_pk": project_pk})
            )
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


class LabourCostListView(ProjectEstimatorMixin, ListView):
    model = ProjectLabourCrew
    template_name = "estimator/labour_costs_list.html"
    context_object_name = "crews"

    def get_queryset(self):
        return ProjectLabourCrew.objects.filter(project=self.get_project())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", LabourCrewForm())
        return context

    def post(self, request, *args, **kwargs):
        project_pk = self.kwargs["project_pk"]
        action = request.POST.get("action")
        qs = ProjectLabourCrew.objects.filter(project=self.get_project())

        if _handle_clear_action(request, qs, label="labour crews"):
            return redirect(
                reverse("estimator:labour_costs", kwargs={"project_pk": project_pk})
            )

        if _handle_bulk_action(request, qs):
            return redirect(
                reverse("estimator:labour_costs", kwargs={"project_pk": project_pk})
            )

        if action == "sync_system":
            from .services import sync_labour_costs_from_contractor

            result = sync_labour_costs_from_contractor(self.get_project())
            _flash_sync_result(request, result, "Labour costs")
            return redirect(
                reverse("estimator:labour_costs", kwargs={"project_pk": project_pk})
            )

        # Default: add crew form
        form = LabourCrewForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.project = self.get_project()
            obj.save()
            return redirect(
                reverse("estimator:labour_costs", kwargs={"project_pk": project_pk})
            )
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


class LabourSpecificationListView(ProjectEstimatorMixin, TemplateView):
    """Labour calculator — driven by Output BoQ.

    Groups BOQItems by (section, labour_specification) and sums
    contract_quantity per group.
    """

    template_name = "estimator/labour_specification_list.html"

    def _build_calculator_rows(
        self, project, section_filter=None, trade_name_filter=None, name_filter=None
    ):
        qs = BOQItem.objects.filter(
            project=project,
            labour_specification__isnull=False,
            labour_specification__is_active=True,
            is_section_header=False,
        ).select_related(
            "labour_specification__crew",
        )

        if section_filter:
            qs = qs.filter(section=section_filter)
        if name_filter:
            qs = qs.filter(labour_specification__name=name_filter)
        if trade_name_filter:
            qs = qs.filter(labour_specification__trade_name=trade_name_filter)

        from collections import OrderedDict

        groups = OrderedDict()
        for boq in qs.order_by("section", "labour_specification__name"):
            key = (boq.section, getattr(boq, "labour_specification_id", None))
            if key not in groups:
                groups[key] = {
                    "section": boq.section,
                    "ls": boq.labour_specification,
                    "boq_qty": Decimal("0"),
                }
            if boq.contract_quantity:
                groups[key]["boq_qty"] += boq.contract_quantity

        rows = []
        for group in groups.values():
            ls = group["ls"]
            boq_qty = group["boq_qty"]
            rows.append(
                {
                    "section": group["section"],
                    "ls": ls,
                    "boq_qty": boq_qty,
                    "daily_output": ls.daily_output,
                    "daily_cost": ls.daily_cost,
                    "rate_per_unit": ls.rate_per_unit,
                    "total_cost": boq_qty * ls.rate_per_unit
                    if boq_qty
                    else Decimal("0"),
                }
            )
        return rows

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()

        f_section = self.request.GET.get("section", "")
        f_trade_name = self.request.GET.get("trade_name", "")
        f_name = self.request.GET.get("name", "")

        context["labour_specs"] = self._build_calculator_rows(
            project,
            section_filter=f_section or None,
            trade_name_filter=f_trade_name or None,
            name_filter=f_name or None,
        )

        context["form"] = context.get("form", LabourSpecificationForm(project=project))

        # Filter options from Output BoQ
        boq_with_lspec = BOQItem.objects.filter(
            project=project,
            labour_specification__isnull=False,
            labour_specification__is_active=True,
            is_section_header=False,
        )
        context["sections"] = (
            boq_with_lspec.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_names"] = (
            boq_with_lspec.exclude(labour_specification__trade_name="")
            .values_list("labour_specification__trade_name", flat=True)
            .distinct()
            .order_by("labour_specification__trade_name")
        )
        context["names"] = (
            boq_with_lspec.values_list("labour_specification__name", flat=True)
            .distinct()
            .order_by("labour_specification__name")
        )

        context["f_section"] = f_section
        context["f_trade_name"] = f_trade_name
        context["f_name"] = f_name

        return context

    def post(self, request, *args, **kwargs):
        project = self.get_project()
        form = LabourSpecificationForm(request.POST, project=project)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.project = project
            obj.save()
            return redirect(
                reverse(
                    "estimator:labour_specs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        return self.render_to_response(self.get_context_data(form=form))


class TradeCodeListView(ProjectEstimatorMixin, ListView):
    model = ProjectTradeCode
    template_name = "estimator/trade_code_list.html"
    context_object_name = "trade_codes"

    def get_queryset(self):
        return ProjectTradeCode.objects.filter(project=self.get_project())

    def post(self, request, *args, **kwargs):
        project_pk = self.kwargs["project_pk"]
        action = request.POST.get("action")
        qs = ProjectTradeCode.objects.filter(project=self.get_project())
        if _handle_clear_action(request, qs, label="trade codes"):
            return redirect(
                reverse("estimator:trade_codes", kwargs={"project_pk": project_pk})
            )
        if _handle_bulk_action(request, qs):
            return redirect(
                reverse("estimator:trade_codes", kwargs={"project_pk": project_pk})
            )
        if action == "sync_system":
            from .services import sync_trade_codes_from_contractor

            result = sync_trade_codes_from_contractor(self.get_project())
            _flash_sync_result(request, result, "Trade codes")
            return redirect(
                reverse("estimator:trade_codes", kwargs={"project_pk": project_pk})
            )
        return redirect(
            reverse("estimator:trade_codes", kwargs={"project_pk": project_pk})
        )


class ExcelImportView(ProjectEstimatorMixin, FormView):
    template_name = "estimator/import_excel.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:import_excel", kwargs={"project_pk": self.kwargs["project_pk"]}
        )

    def form_valid(self, form):
        uploaded = form.cleaned_data["file"]

        # Write uploaded file to a temp file for openpyxl
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        try:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp.close()

            from .management.commands.import_excel import ExcelImporter

            importer = ExcelImporter(tmp.name, project=self.get_project())
            results = importer.run()

            parts = []
            labels = {
                "trade_codes": "Trade Codes",
                "materials": "Material Costs",
                "specifications": "Material Estimator",
                "labour_crews": "Labour Crews",
                "labour_specs": "Labour Specs",
                "plant_costs": "Plant Costs",
                "plant_specs": "Plant Specs",
                "preliminary_costs": "Preliminary Costs",
                "preliminary_specs": "Preliminary Specs",
                "boq_items": "Output BoQ Items",
            }
            for key, label in labels.items():
                if key in results:
                    parts.append(f"{label}: {results[key]}")

            skipped = [
                label
                for key, label in labels.items()
                if key not in results and key != "boq_items"
            ]
            msg = f"Import successful — {', '.join(parts)}"
            if skipped:
                msg += (
                    f". Skipped (sheet not found): {', '.join(skipped)}. "
                    f"Sheets in file: {importer.sheet_names}"
                )
            messages.success(self.request, msg)
        except Exception as e:
            messages.error(self.request, f"Import failed: {e}")
        finally:
            os.unlink(tmp.name)

        return redirect(self.get_success_url())


class ReportsIndexView(ProjectEstimatorMixin, TemplateView):
    template_name = "estimator/reports_index.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()
        items = BOQItem.objects.filter(project=project)

        totals = {
            "contract": Decimal("0"),
            "baseline": Decimal("0"),
            "progress": Decimal("0"),
            "forecast": Decimal("0"),
            "mat": {
                "baseline": Decimal("0"),
                "progress": Decimal("0"),
                "forecast": Decimal("0"),
            },
            "lab": {
                "baseline": Decimal("0"),
                "progress": Decimal("0"),
                "forecast": Decimal("0"),
            },
            "plt": {
                "baseline": Decimal("0"),
                "progress": Decimal("0"),
                "forecast": Decimal("0"),
            },
            "pre": {
                "baseline": Decimal("0"),
                "progress": Decimal("0"),
                "forecast": Decimal("0"),
            },
        }

        for item in items:
            # Total Project Amounts
            contract_amt = item.contract_amount
            baseline_price = item.baseline_new_price
            qty_c = item.contract_quantity or Decimal("0")
            qty_p = item.progress_quantity or Decimal("0")
            qty_f = item.forecast_quantity or Decimal("0")

            baseline_amt = baseline_price * qty_c if baseline_price else Decimal("0")
            progress_amt = item.progress_amount or Decimal("0")
            forecast_amt = item.forecast_amount or Decimal("0")

            if contract_amt:
                totals["contract"] += contract_amt
            if baseline_amt:
                totals["baseline"] += baseline_amt
            if progress_amt:
                totals["progress"] += progress_amt
            if forecast_amt:
                totals["forecast"] += forecast_amt

            # Resource Specific Amounts
            m_rate = item.new_materials_rate
            m_wastage = item._wastage_factor
            if m_rate:
                totals["mat"]["baseline"] += m_rate * m_wastage * qty_c
                totals["mat"]["progress"] += m_rate * m_wastage * qty_p
                totals["mat"]["forecast"] += m_rate * m_wastage * qty_f

            l_rate = item.new_labour_rate
            if l_rate:
                totals["lab"]["baseline"] += l_rate * qty_c
                totals["lab"]["progress"] += l_rate * qty_p
                totals["lab"]["forecast"] += l_rate * qty_f

            p_rate = item.new_plant_rate
            if p_rate:
                totals["plt"]["baseline"] += p_rate * qty_c
                totals["plt"]["progress"] += p_rate * qty_p
                totals["plt"]["forecast"] += p_rate * qty_f

            pr_rate = item.new_preliminary_rate
            if pr_rate:
                totals["pre"]["baseline"] += pr_rate * qty_c
                totals["pre"]["progress"] += pr_rate * qty_p
                totals["pre"]["forecast"] += pr_rate * qty_f

        def get_summary(label, amt_a, amt_b, report_url_name):
            var, pct = calculate_variance(amt_a, amt_b)
            return {
                "label": label,
                "amount_a": amt_a,
                "amount_b": amt_b,
                "variance": var,
                "variance_pct": pct,
                "url": reverse(
                    f"estimator:{report_url_name}",
                    kwargs={"project_pk": project.pk},
                ),
            }

        context["summaries"] = [
            get_summary(
                "Baseline Assessment",
                totals["contract"],
                totals["baseline"],
                "report_baseline_assessment",
            ),
            get_summary(
                "Progress Assessment",
                totals["baseline"],
                totals["progress"],
                "report_progress_assessment",
            ),
            get_summary(
                "Forecast Assessment",
                totals["baseline"],
                totals["forecast"],
                "report_forecast_assessment",
            ),
        ]

        def get_resource_summaries(resource_totals, res_type, list_prefix):
            base = resource_totals["baseline"]
            prog = resource_totals["progress"]
            fore = resource_totals["forecast"]

            var_p, pct_p = calculate_variance(base, prog)
            var_f, pct_f = calculate_variance(base, fore)

            return {
                "baseline": {
                    "label": f"{res_type} Baseline",
                    "amount": base,
                    "url": reverse(
                        f"estimator:report_{list_prefix}_list_baseline",
                        kwargs={"project_pk": project.pk},
                    ),
                },
                "progress": {
                    "label": f"{res_type} Progress",
                    "amount_a": base,
                    "amount_b": prog,
                    "variance": var_p,
                    "variance_pct": pct_p,
                    "url": reverse(
                        f"estimator:report_{list_prefix}_list_progress",
                        kwargs={"project_pk": project.pk},
                    ),
                },
                "forecast": {
                    "label": f"{res_type} Forecast",
                    "amount_a": base,
                    "amount_b": fore,
                    "variance": var_f,
                    "variance_pct": pct_f,
                    "url": reverse(
                        f"estimator:report_{list_prefix}_list_forecast",
                        kwargs={"project_pk": project.pk},
                    ),
                },
            }

        context["mat_summary"] = get_resource_summaries(
            totals["mat"], "Material", "material"
        )
        context["lab_summary"] = get_resource_summaries(
            totals["lab"], "Labour", "labour"
        )
        context["plt_summary"] = get_resource_summaries(totals["plt"], "Plant", "plant")
        context["pre_summary"] = get_resource_summaries(
            totals["pre"], "Prelim", "prelim"
        )

        return context


class PricedBoqReportView(ProjectEstimatorMixin, ListView):
    model = BOQItem
    template_name = "estimator/reports/priced_boq.html"
    context_object_name = "items"

    REPORT_CONFIGS = {
        "baseline_assessment": {
            "title": "Baseline Assessment",
            "subtitle": "Compares new and old amount based on original quantities",
            "amount_a_label": "Baseline Amount",
            "amount_b_label": "New Baseline Amount",
            "key_rates_only": False,
        },
        "progress_assessment": {
            "title": "Progress Assessment",
            "subtitle": "Compares original amount and new amount based on progress quantities",
            "amount_a_label": "Baseline Amount",
            "amount_b_label": "Progress Amount",
            "key_rates_only": False,
        },
        "forecast_assessment": {
            "title": "Forecast Assessment",
            "subtitle": "Compares baseline BoQ with new rates and forecast quantities",
            "amount_a_label": "Baseline Amount",
            "amount_b_label": "Forecast Amount",
            "key_rates_only": False,
        },
        "key_rates_assessment": {
            "title": "Key Rates Assessment",
            "subtitle": "Compares new and old amount based on original quantities for priced items",
            "amount_a_label": "Baseline Amount",
            "amount_b_label": "New Baseline Amount",
            "key_rates_only": True,
        },
    }

    def _get_config(self):
        return self.REPORT_CONFIGS[self.kwargs["report_type"]]

    def _get_amounts(self, item, report_type):
        """Return (amount_a, amount_b) for a given item and report type."""
        contract_amt = item.contract_amount
        baseline_amt = (
            item.baseline_new_price * item.contract_quantity
            if item.baseline_new_price and item.contract_quantity
            else None
        )

        if report_type == "progress_assessment":
            return (baseline_amt, item.progress_amount)
        elif report_type == "forecast_assessment":
            return (baseline_amt, item.forecast_amount)
        else:  # baseline_assessment and key_rates_assessment
            return (contract_amt, baseline_amt)

    def get_queryset(self):
        qs = (
            BOQItem.objects.filter(project=self.get_project())
            .select_related(
                "trade_code",
                "specification",
                "labour_specification",
                "labour_specification__crew",
                "plant_specification",
                "preliminary_specification",
                "material",
                "project__estimator_assumptions",
            )
            .prefetch_related(
                "specification__spec_components__material",
            )
            .filter(is_section_header=False)
        )

        config = self._get_config()
        if config["key_rates_only"]:
            qs = qs.filter(
                models.Q(specification__isnull=False) | models.Q(material__isnull=False)
            )

        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)

        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        config = self._get_config()
        report_type = self.kwargs["report_type"]

        context["report_title"] = config["title"]
        context["report_subtitle"] = config["subtitle"]
        context["amount_a_label"] = config["amount_a_label"]
        context["amount_b_label"] = config["amount_b_label"]

        total_a = Decimal("0")
        total_b = Decimal("0")
        report_rows = []
        # Group rows by section so we can show per-section subtotals and a
        # section-level variance comparison. Insertion order follows the
        # queryset (BoQ order), matching the flat detail table.
        section_map: dict[str, dict] = {}

        for item in context["items"]:
            amount_a, amount_b = self._get_amounts(item, report_type)
            variance_amt, variance_pct = calculate_variance(amount_a, amount_b)

            mat_rate = item.new_materials_rate
            lab_rate = item.new_labour_rate
            plant_rate = item.new_plant_rate
            prelim_rate = item.new_preliminary_rate
            bnp = item.baseline_new_price
            if bnp and bnp > 0:
                mat_pct = (Decimal(str(mat_rate or 0)) / bnp) * Decimal("100")
                lab_pct = (Decimal(str(lab_rate or 0)) / bnp) * Decimal("100")
                plant_pct = (Decimal(str(plant_rate or 0)) / bnp) * Decimal("100")
                prelim_pct = (Decimal(str(prelim_rate or 0)) / bnp) * Decimal("100")
            else:
                mat_pct = None
                lab_pct = None
                plant_pct = None
                prelim_pct = None

            if amount_a:
                total_a += amount_a
            if amount_b:
                total_b += amount_b

            row = {
                "section": item.section,
                "bill_no": item.bill_no,
                "description": item.description,
                "amount_a": amount_a,
                "amount_b": amount_b,
                "variance_amount": variance_amt,
                "variance_pct": variance_pct,
                "materials_pct": mat_pct,
                "labour_pct": lab_pct,
                "plant_pct": plant_pct,
                "preliminary_pct": prelim_pct,
            }
            report_rows.append(row)

            section_key = item.section or "—"
            group = section_map.get(section_key)
            if group is None:
                group = {
                    "section": section_key,
                    "rows": [],
                    "total_a": Decimal("0"),
                    "total_b": Decimal("0"),
                }
                section_map[section_key] = group
            group["rows"].append(row)
            if amount_a:
                group["total_a"] += amount_a
            if amount_b:
                group["total_b"] += amount_b

        # Finalise per-section variance.
        section_groups = list(section_map.values())
        for group in section_groups:
            var_amt, var_pct = calculate_variance(group["total_a"], group["total_b"])
            group["variance_amount"] = var_amt
            group["variance_pct"] = var_pct

        # Section comparison panel: sort by magnitude of variance so the
        # sections carrying the most risk surface first, and size a bar
        # relative to the largest swing for an at-a-glance read.
        max_abs_variance = max(
            (
                abs(g["variance_amount"])
                for g in section_groups
                if g["variance_amount"] is not None
            ),
            default=Decimal("0"),
        )
        section_summary = sorted(
            section_groups,
            key=lambda g: abs(g["variance_amount"] or Decimal("0")),
            reverse=True,
        )
        for group in section_summary:
            var_amt = group["variance_amount"]
            if var_amt is not None and max_abs_variance > 0:
                group["bar_pct"] = int(abs(var_amt) / max_abs_variance * 100)
            else:
                group["bar_pct"] = 0

        total_variance, total_variance_pct = calculate_variance(total_a, total_b)
        context["report_rows"] = report_rows
        context["section_groups"] = section_groups
        context["section_summary"] = section_summary
        context["total_a"] = total_a
        context["total_b"] = total_b
        context["total_variance"] = total_variance
        context["total_variance_pct"] = total_variance_pct

        # Filter options
        project = self.get_project()
        project_items = BOQItem.objects.filter(project=project)
        context["sections"] = (
            project_items.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_codes"] = ProjectTradeCode.objects.filter(project=project)
        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_code"] = self.request.GET.get("trade_code", "")

        return context


class MaterialListReportView(ProjectEstimatorMixin, ListView):
    model = BOQItem
    template_name = "estimator/reports/material_list.html"
    context_object_name = "items"

    VARIANT_CONFIGS = {
        "baseline": {
            "title_new": "Baseline Material List (New Rates)",
            "title_contract": "Baseline Material List (Contract Rates)",
            "qty_field": "contract_quantity",
        },
        "progress": {
            "title_new": "Progress Material List (New Rates)",
            "title_contract": "Progress Material List (Contract Rates)",
            "qty_field": "progress_quantity",
        },
        "forecast": {
            "title_new": "Forecast Material List (New Rates)",
            "title_contract": "Forecast Material List (Contract Rates)",
            "qty_field": "forecast_quantity",
        },
    }

    def _get_config(self):
        return self.VARIANT_CONFIGS[self.kwargs["variant"]]

    def _get_rate_type(self):
        return self.kwargs.get("rate_type", "new")

    def get_queryset(self):
        qs = (
            BOQItem.objects.filter(project=self.get_project())
            .select_related(
                "trade_code",
                "specification",
                "material",
                "project__estimator_assumptions",
            )
            .prefetch_related(
                "specification__spec_components__material",
            )
            .filter(
                is_section_header=False,
            )
            .filter(
                models.Q(specification__isnull=False) | models.Q(material__isnull=False)
            )
        )

        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)

        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        config = self._get_config()
        qty_field = config["qty_field"]
        rate_type = self._get_rate_type()

        title_key = "title_contract" if rate_type == "contract" else "title_new"
        context["report_title"] = config[title_key]
        context["spec_label"] = "Material Spec"
        context["rate_type"] = rate_type

        if rate_type == "contract":
            context["parent_template"] = (
                "estimator/base_baseline_estimator_materials.html"
            )
        else:
            context["parent_template"] = "estimator/base_materials_estimator.html"
        variant = self.kwargs["variant"]
        list_url_name = (
            f"estimator:report_material_list_{variant}_contract"
            if rate_type == "contract"
            else f"estimator:report_material_list_{variant}"
        )
        context["toggle_spec_url"] = reverse(
            list_url_name, kwargs={"project_pk": self.get_project().pk}
        )
        context["toggle_alt_url"] = reverse(
            f"estimator:report_material_components_{variant}",
            kwargs={"project_pk": self.get_project().pk},
        )
        context["toggle_alt_label"] = "Component"
        context["toggle_active"] = "spec"

        grand_total = Decimal("0")
        aggregated: dict[tuple, dict] = {}
        trade_totals: dict[str, Decimal] = {}

        for item in context["items"]:
            raw_quantity = getattr(item, qty_field) or Decimal("0")
            if not raw_quantity:
                continue
            wastage_quantity = raw_quantity * item._wastage_factor
            if rate_type == "contract":
                rate = item.contract_rate
                amount = rate * raw_quantity if rate else None
            else:
                rate = item.new_materials_rate
                amount = wastage_quantity * rate if rate else None

            if item.specification:
                key = ("spec", item.specification_id)
                material_name = item.specification.name
                unit = item.unit
            elif item.material:
                key = ("mat", item.material_id)
                material_name = item.material.material_code
                unit = item.material.unit
            else:
                continue

            row = aggregated.setdefault(
                key,
                {
                    "material_name": material_name,
                    "unit": unit,
                    "quantity": Decimal("0"),
                    "wastage_quantity": Decimal("0"),
                    "amount": Decimal("0"),
                },
            )
            row["quantity"] += raw_quantity
            row["wastage_quantity"] += wastage_quantity
            if amount:
                row["amount"] += amount
                grand_total += amount
                t = str(item.trade_code) if item.trade_code else "Unassigned"
                trade_totals[t] = trade_totals.get(t, Decimal("0")) + amount

        report_rows = []
        for row in aggregated.values():
            qty_for_rate = (
                row["wastage_quantity"] if rate_type != "contract" else row["quantity"]
            )
            row["rate"] = row["amount"] / qty_for_rate if qty_for_rate else None
            report_rows.append(row)
        report_rows.sort(key=lambda r: r["amount"], reverse=True)

        for row in report_rows:
            row["pct_of_total"] = calculate_pct_of_total(row["amount"], grand_total)

        context["report_rows"] = report_rows
        context["grand_total"] = grand_total
        context["total_quantity"] = sum(
            (r["quantity"] for r in report_rows), Decimal("0")
        )
        context["total_wastage_quantity"] = sum(
            (r["wastage_quantity"] for r in report_rows), Decimal("0")
        )
        context["show_wastage"] = True
        context["show_rate"] = True

        material_totals: dict[str, Decimal] = {}
        for row in report_rows:
            if row["amount"]:
                m = row["material_name"] or "Unknown"
                material_totals[m] = (
                    material_totals.get(m, Decimal("0")) + row["amount"]
                )

        # Sort trades by amount descending
        sorted_trades = sorted(trade_totals.items(), key=lambda x: x[1], reverse=True)
        context["chart_trade_labels"] = json.dumps([t[0] for t in sorted_trades])
        context["chart_trade_values"] = json.dumps([float(t[1]) for t in sorted_trades])

        # Top 10 materials by amount
        sorted_materials = sorted(
            material_totals.items(), key=lambda x: x[1], reverse=True
        )[:10]
        context["chart_material_labels"] = json.dumps([m[0] for m in sorted_materials])
        context["chart_material_values"] = json.dumps(
            [float(m[1]) for m in sorted_materials]
        )

        # Filter options
        project = self.get_project()
        project_items = BOQItem.objects.filter(project=project)
        context["sections"] = (
            project_items.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_codes"] = ProjectTradeCode.objects.filter(project=project)
        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_code"] = self.request.GET.get("trade_code", "")

        return context


class LabourListReportView(ProjectEstimatorMixin, ListView):
    model = BOQItem
    template_name = "estimator/reports/labour_list.html"
    context_object_name = "items"

    VARIANT_CONFIGS = {
        "baseline": {
            "title_new": "Baseline Labour List (New Rates)",
            "title_contract": "Baseline Labour List (Contract Rates)",
            "qty_field": "contract_quantity",
        },
        "progress": {
            "title_new": "Progress Labour List (New Rates)",
            "title_contract": "Progress Labour List (Contract Rates)",
            "qty_field": "progress_quantity",
        },
        "forecast": {
            "title_new": "Forecast Labour List (New Rates)",
            "title_contract": "Forecast Labour List (Contract Rates)",
            "qty_field": "forecast_quantity",
        },
    }

    def _get_config(self):
        return self.VARIANT_CONFIGS[self.kwargs["variant"]]

    def _get_rate_type(self):
        return self.kwargs.get("rate_type", "new")

    def get_queryset(self):
        qs = (
            BOQItem.objects.filter(project=self.get_project())
            .select_related(
                "trade_code",
                "labour_specification",
                "labour_specification__crew",
            )
            .filter(
                is_section_header=False,
                labour_specification__isnull=False,
            )
        )

        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)

        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        config = self._get_config()
        qty_field = config["qty_field"]
        rate_type = self._get_rate_type()

        title_key = "title_contract" if rate_type == "contract" else "title_new"
        context["report_title"] = config[title_key]
        context["rate_type"] = rate_type

        if rate_type == "contract":
            context["parent_template"] = "estimator/base_baseline_estimator_labour.html"
        else:
            context["parent_template"] = "estimator/base_labour_estimator.html"
        variant = self.kwargs["variant"]
        list_url_name = (
            f"estimator:report_labour_list_{variant}_contract"
            if rate_type == "contract"
            else f"estimator:report_labour_list_{variant}"
        )
        context["toggle_spec_url"] = reverse(
            list_url_name, kwargs={"project_pk": self.get_project().pk}
        )
        context["toggle_alt_url"] = reverse(
            f"estimator:report_labour_skills_{variant}",
            kwargs={"project_pk": self.get_project().pk},
        )
        context["toggle_alt_label"] = "Skill"
        context["toggle_active"] = "spec"

        grand_total = Decimal("0")
        aggregated: dict[int, dict] = {}

        for item in context["items"]:
            ls = item.labour_specification
            if ls is None:
                continue
            crew = ls.crew
            quantity = getattr(item, qty_field) or Decimal("0")
            if not quantity:
                continue
            if rate_type == "contract":
                rate = item.contract_rate
                amount = rate * quantity if rate else None
            else:
                rate = item.new_labour_rate
                amount = quantity * rate if rate else None

            row = aggregated.setdefault(
                ls.pk,
                {
                    "spec_name": ls.name,
                    "crew_type": crew.crew_type if crew else "-",
                    "no_of_crews": 1,
                    "quantity": Decimal("0"),
                    "amount": Decimal("0"),
                    "skilled": crew.skilled if crew else 0,
                    "semi_skilled": crew.semi_skilled if crew else 0,
                    "general": crew.general if crew else 0,
                },
            )
            row["quantity"] += quantity
            if amount:
                row["amount"] += amount
                grand_total += amount

        report_rows = []
        for row in aggregated.values():
            row["crew_rate_per_unit"] = (
                row["amount"] / row["quantity"] if row["quantity"] else None
            )
            report_rows.append(row)
        report_rows.sort(key=lambda r: r["amount"], reverse=True)

        for row in report_rows:
            row["pct_of_total"] = calculate_pct_of_total(row["amount"], grand_total)

        context["report_rows"] = report_rows
        context["grand_total"] = grand_total
        context["total_no_of_crews"] = sum((r["no_of_crews"] for r in report_rows), 0)
        context["total_quantity"] = sum(
            (r["quantity"] for r in report_rows), Decimal("0")
        )
        context["total_skilled"] = sum((r["skilled"] for r in report_rows), 0)
        context["total_semi_skilled"] = sum((r["semi_skilled"] for r in report_rows), 0)
        context["total_general"] = sum((r["general"] for r in report_rows), 0)

        # Chart data: cost by crew type
        crew_totals = {}
        crew_composition = {}
        for row in report_rows:
            ct = row["crew_type"]
            if row["amount"]:
                crew_totals[ct] = crew_totals.get(ct, Decimal("0")) + row["amount"]
            if ct not in crew_composition:
                crew_composition[ct] = {
                    "skilled": row["skilled"],
                    "semi_skilled": row["semi_skilled"],
                    "general": row["general"],
                }

        sorted_crews = sorted(crew_totals.items(), key=lambda x: x[1], reverse=True)
        context["chart_crew_labels"] = json.dumps([c[0] for c in sorted_crews])
        context["chart_crew_values"] = json.dumps([float(c[1]) for c in sorted_crews])

        # Crew composition for stacked bar
        comp_labels = sorted(crew_composition.keys())
        context["chart_comp_labels"] = json.dumps(comp_labels)
        context["chart_comp_skilled"] = json.dumps(
            [crew_composition[c]["skilled"] for c in comp_labels]
        )
        context["chart_comp_semi"] = json.dumps(
            [crew_composition[c]["semi_skilled"] for c in comp_labels]
        )
        context["chart_comp_general"] = json.dumps(
            [crew_composition[c]["general"] for c in comp_labels]
        )

        # Filter options
        project = self.get_project()
        project_items = BOQItem.objects.filter(project=project)
        context["sections"] = (
            project_items.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_codes"] = ProjectTradeCode.objects.filter(project=project)
        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_code"] = self.request.GET.get("trade_code", "")

        return context


class _SimpleSpecListReportView(ProjectEstimatorMixin, ListView):
    """Shared base: Plant/Prelim report pages mirroring MaterialListReportView.

    Concrete subclasses set spec_field, spec_rate_attr, parent_template_new,
    parent_template_contract, title_noun.
    """

    model = BOQItem
    template_name = "estimator/reports/material_list.html"
    context_object_name = "items"

    VARIANT_CONFIGS = {
        "baseline": {"qty_field": "contract_quantity"},
        "progress": {"qty_field": "progress_quantity"},
        "forecast": {"qty_field": "forecast_quantity"},
    }

    spec_field = ""
    spec_rate_attr = ""
    parent_template_new = ""
    parent_template_contract = ""
    title_noun = ""
    # First-column header label (e.g. "Plant Spec", "Prelim Spec").
    spec_label = "Spec"
    # Optional URL-name prefixes to enable the "By Spec / By Component" toggle.
    # Subclass sets both to enable; leave empty to omit (e.g. preliminary).
    spec_url_prefix = ""
    component_url_prefix = ""
    component_label = "Component"
    # Wastage applies to materials only; rate column hidden for preliminaries.
    show_wastage = False
    show_rate = True

    def _get_rate_type(self):
        return self.kwargs.get("rate_type", "new")

    def get_queryset(self):
        qs = (
            BOQItem.objects.filter(project=self.get_project())
            .select_related("trade_code", self.spec_field)
            .filter(is_section_header=False)
            .filter(**{f"{self.spec_field}__isnull": False})
        )
        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)
        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        variant = self.kwargs["variant"]
        qty_field = self.VARIANT_CONFIGS[variant]["qty_field"]
        rate_type = self._get_rate_type()
        variant_title = variant.capitalize()
        rate_suffix = "Contract Rates" if rate_type == "contract" else "New Rates"
        context["report_title"] = (
            f"{variant_title} {self.title_noun} List ({rate_suffix})"
        )
        context["spec_label"] = self.spec_label
        context["rate_type"] = rate_type
        context["parent_template"] = (
            self.parent_template_contract
            if rate_type == "contract"
            else self.parent_template_new
        )
        if self.spec_url_prefix and self.component_url_prefix:
            list_url_name = (
                f"{self.spec_url_prefix}{variant}_contract"
                if rate_type == "contract"
                else f"{self.spec_url_prefix}{variant}"
            )
            context["toggle_spec_url"] = reverse(
                list_url_name, kwargs={"project_pk": self.get_project().pk}
            )
            context["toggle_alt_url"] = reverse(
                f"{self.component_url_prefix}{variant}",
                kwargs={"project_pk": self.get_project().pk},
            )
            context["toggle_alt_label"] = self.component_label
            context["toggle_active"] = "spec"

        grand_total = Decimal("0")
        aggregated: dict[int, dict] = {}
        trade_totals: dict[str, Decimal] = {}

        for item in context["items"]:
            spec = getattr(item, self.spec_field)
            if spec is None:
                continue
            quantity = getattr(item, qty_field) or Decimal("0")
            if not quantity:
                continue
            if rate_type == "contract":
                rate = item.contract_rate
            else:
                rate = getattr(spec, self.spec_rate_attr)
            amount = rate * quantity if rate else None

            row = aggregated.setdefault(
                spec.pk,
                {
                    "material_name": spec.name,
                    "unit": item.unit,
                    "quantity": Decimal("0"),
                    "amount": Decimal("0"),
                },
            )
            row["quantity"] += quantity
            if amount:
                row["amount"] += amount
                grand_total += amount
                t = str(item.trade_code) if item.trade_code else "Unassigned"
                trade_totals[t] = trade_totals.get(t, Decimal("0")) + amount

        report_rows = []
        for row in aggregated.values():
            row["rate"] = row["amount"] / row["quantity"] if row["quantity"] else None
            report_rows.append(row)
        report_rows.sort(key=lambda r: r["amount"], reverse=True)

        for row in report_rows:
            row["pct_of_total"] = calculate_pct_of_total(row["amount"], grand_total)

        context["report_rows"] = report_rows
        context["grand_total"] = grand_total
        context["total_quantity"] = sum(
            (r["quantity"] for r in report_rows), Decimal("0")
        )
        context["show_wastage"] = self.show_wastage
        context["show_rate"] = self.show_rate

        name_totals: dict[str, Decimal] = {}
        for row in report_rows:
            if row["amount"]:
                m = row["material_name"] or "Unknown"
                name_totals[m] = name_totals.get(m, Decimal("0")) + row["amount"]

        sorted_trades = sorted(trade_totals.items(), key=lambda x: x[1], reverse=True)
        context["chart_trade_labels"] = json.dumps([t[0] for t in sorted_trades])
        context["chart_trade_values"] = json.dumps([float(t[1]) for t in sorted_trades])
        sorted_names = sorted(name_totals.items(), key=lambda x: x[1], reverse=True)[
            :10
        ]
        context["chart_material_labels"] = json.dumps([m[0] for m in sorted_names])
        context["chart_material_values"] = json.dumps(
            [float(m[1]) for m in sorted_names]
        )

        project = self.get_project()
        project_items = BOQItem.objects.filter(project=project)
        context["sections"] = (
            project_items.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_codes"] = ProjectTradeCode.objects.filter(project=project)
        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_code"] = self.request.GET.get("trade_code", "")
        return context


class PlantListReportView(_SimpleSpecListReportView):
    spec_field = "plant_specification"
    spec_rate_attr = "rate_per_unit"
    parent_template_new = "estimator/base_plant_estimator.html"
    parent_template_contract = "estimator/base_baseline_estimator_plant.html"
    title_noun = "Plant"
    spec_label = "Plant Spec"
    spec_url_prefix = "estimator:report_plant_list_"
    component_url_prefix = "estimator:report_plant_components_"
    component_label = "Component"


class PreliminaryListReportView(_SimpleSpecListReportView):
    spec_field = "preliminary_specification"
    spec_rate_attr = "amount"
    parent_template_new = "estimator/base_prelim_estimator.html"
    parent_template_contract = "estimator/base_baseline_estimator_prelim.html"
    title_noun = "Preliminary"
    spec_label = "Prelim Spec"
    show_rate = False


# ───────────────────────────────────────────────────────────────────
# "By Component" / "By Skill" project-aggregated reports
# ───────────────────────────────────────────────────────────────────


class _ComponentReportBase(ProjectEstimatorMixin, TemplateView):
    """Shared variant config for the project-aggregated component reports."""

    VARIANT_CONFIGS = {
        "baseline": {"qty_field": "contract_quantity", "title_prefix": "Baseline"},
        "progress": {"qty_field": "progress_quantity", "title_prefix": "Progress"},
        "forecast": {"qty_field": "forecast_quantity", "title_prefix": "Forecast"},
    }

    def _config(self):
        return self.VARIANT_CONFIGS[self.kwargs["variant"]]


class MaterialComponentReportView(_ComponentReportBase):
    """Aggregated material/component qtys needed across the whole project."""

    template_name = "estimator/reports/material_components_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        config = self._config()
        qty_field = config["qty_field"]
        project = self.get_project()
        variant = self.kwargs["variant"]

        # Project-level wastage factor (constant across all items) — applied
        # once after DB aggregation rather than per-row.
        try:
            wastage_pct = project.estimator_assumptions.wastage_pct or Decimal("0")
        except Exception:
            wastage_pct = Decimal("0")
        wastage_factor = Decimal("1") + wastage_pct / Decimal("100")

        # Per-item markup is (1 + (mat_markup + transport)/100). To avoid SQLite's
        # integer-division semantics on (100 + m + t)/100, we sum
        # (... × (100 + m + t)) inside SQL and divide by 100 in Python.
        spec_markup_x100 = (
            Value(Decimal("100"))
            + F("specification__boq_items__material_markup_pct")
            + F("specification__boq_items__transport_pct")
        )
        direct_markup_x100 = (
            Value(Decimal("100")) + F("material_markup_pct") + F("transport_pct")
        )

        # Spec branch: per-(component × boq_item) row, grouped by destination material.
        # NULL qty_field rows contribute NULL → ignored by Sum (matches original
        # `if not raw_qty: continue`). qty_per_unit=0 contributes 0 → matches
        # `if not comp.qty_per_unit: continue` for amount, and zero qty for quantity.
        spec_aggs = (
            ProjectSpecificationComponent.objects.filter(
                specification__boq_items__project=project,
                specification__boq_items__is_section_header=False,
                material__isnull=False,
            )
            .values(
                "material__material_code",
                "material__trade_name",
                "material__unit",
            )
            .annotate(
                sum_qty=Sum(
                    F("qty_per_unit") * F(f"specification__boq_items__{qty_field}"),
                    output_field=DecimalField(max_digits=24, decimal_places=6),
                ),
                sum_amount_x100=Sum(
                    F("qty_per_unit")
                    * F(f"specification__boq_items__{qty_field}")
                    * F("material__market_rate")
                    * spec_markup_x100,
                    output_field=DecimalField(max_digits=24, decimal_places=6),
                ),
            )
        )

        # Direct material branch: items with no specification but with a material.
        direct_aggs = (
            BOQItem.objects.filter(
                project=project,
                is_section_header=False,
                specification__isnull=True,
                material__isnull=False,
            )
            .values(
                "material__material_code",
                "material__trade_name",
                "material__unit",
            )
            .annotate(
                sum_qty=Sum(
                    F(qty_field),
                    output_field=DecimalField(max_digits=24, decimal_places=6),
                ),
                sum_amount_x100=Sum(
                    F(qty_field) * F("material__market_rate") * direct_markup_x100,
                    output_field=DecimalField(max_digits=24, decimal_places=6),
                ),
            )
        )

        # Merge by material_code, applying the project-wide wastage factor and
        # the deferred /100 from the markup expression.
        components: dict[str, dict] = {}

        def merge(agg):
            code = agg["material__material_code"]
            if code is None:
                return
            sum_qty = agg["sum_qty"]
            sum_amount_x100 = agg["sum_amount_x100"]
            # Skip materials with no contribution from this variant — mirrors the
            # original `if not raw_qty: continue` which never created a bucket.
            if not sum_qty and not sum_amount_x100:
                return
            row = components.setdefault(
                code,
                {
                    "material_code": code,
                    "trade_name": agg["material__trade_name"],
                    "unit": agg["material__unit"],
                    "quantity": Decimal("0"),
                    "amount": Decimal("0"),
                },
            )
            row["quantity"] += (sum_qty or Decimal("0")) * wastage_factor
            row["amount"] += (
                (sum_amount_x100 or Decimal("0")) / Decimal("100") * wastage_factor
            )

        for agg in spec_aggs:
            merge(agg)
        for agg in direct_aggs:
            merge(agg)

        report_rows = []
        grand_total = Decimal("0")
        for row in components.values():
            row["rate"] = (
                row["amount"] / row["quantity"] if row["quantity"] else Decimal("0")
            )
            grand_total += row["amount"]
            report_rows.append(row)
        report_rows.sort(key=lambda r: r["amount"], reverse=True)
        for row in report_rows:
            row["pct_of_total"] = calculate_pct_of_total(row["amount"], grand_total)

        context["report_title"] = (
            f"{config['title_prefix']} Materials by Component (Procurement)"
        )
        context["parent_template"] = "estimator/base_materials_estimator.html"
        context["report_rows"] = report_rows
        context["grand_total"] = grand_total
        context["total_quantity"] = sum(
            (r["quantity"] for r in report_rows), Decimal("0")
        )
        context["variant"] = variant
        context["project_pk"] = project.pk

        # Chart: top 10 components by cost
        top = report_rows[:10]
        context["chart_labels"] = json.dumps([r["material_code"] for r in top])
        context["chart_values"] = json.dumps([float(r["amount"]) for r in top])
        return context


class LabourSkillReportView(_ComponentReportBase):
    """Per labour-spec breakdown of skilled / semi-skilled / general person-days."""

    template_name = "estimator/reports/labour_skills_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        config = self._config()
        qty_field = config["qty_field"]
        project = self.get_project()
        variant = self.kwargs["variant"]

        items = BOQItem.objects.filter(
            project=project,
            is_section_header=False,
            labour_specification__isnull=False,
        ).select_related(
            "labour_specification",
            "labour_specification__crew",
        )

        # labour_specification_id -> aggregated row
        specs: dict[int, dict] = {}

        for item in items:
            ls = item.labour_specification
            if ls is None:
                continue
            crew = ls.crew
            raw_qty = getattr(item, qty_field) or Decimal("0")
            if not raw_qty:
                continue
            row = specs.setdefault(
                ls.pk,
                {
                    "spec_name": ls.name,
                    "unit": ls.unit,
                    "daily_output": ls.daily_output,
                    "crew_type": crew.crew_type if crew else "-",
                    "quantity": Decimal("0"),
                    "days_required": Decimal("0"),
                    "skilled": Decimal("0"),
                    "semi_skilled": Decimal("0"),
                    "general": Decimal("0"),
                    "cost": Decimal("0"),
                },
            )
            row["quantity"] += raw_qty
            output = ls.daily_output
            if output and output > 0 and crew:
                days = raw_qty / output
                markup = Decimal("1") + (
                    item.labour_markup_pct or Decimal("0")
                ) / Decimal("100")
                row["days_required"] += days
                row["skilled"] += days * Decimal(str(crew.skilled))
                row["semi_skilled"] += days * Decimal(str(crew.semi_skilled))
                row["general"] += days * Decimal(str(crew.general))
                row["cost"] += days * crew.crew_daily_cost * markup

        report_rows = sorted(specs.values(), key=lambda r: r["cost"], reverse=True)
        grand_total = sum((r["cost"] for r in report_rows), Decimal("0"))
        total_skilled = sum((r["skilled"] for r in report_rows), Decimal("0"))
        total_semi = sum((r["semi_skilled"] for r in report_rows), Decimal("0"))
        total_general = sum((r["general"] for r in report_rows), Decimal("0"))
        total_quantity = sum((r["quantity"] for r in report_rows), Decimal("0"))
        total_days = sum((r["days_required"] for r in report_rows), Decimal("0"))
        for row in report_rows:
            row["pct_of_total"] = calculate_pct_of_total(row["cost"], grand_total)

        context["report_title"] = f"{config['title_prefix']} Labour by Skill"
        context["parent_template"] = "estimator/base_labour_estimator.html"
        context["report_rows"] = report_rows
        context["grand_total"] = grand_total
        context["total_skilled"] = total_skilled
        context["total_semi_skilled"] = total_semi
        context["total_general"] = total_general
        context["total_quantity"] = total_quantity
        context["total_days"] = total_days
        context["variant"] = variant
        context["project_pk"] = project.pk

        context["chart_labels"] = json.dumps([r["spec_name"] for r in report_rows[:10]])
        context["chart_skilled"] = json.dumps(
            [float(r["skilled"]) for r in report_rows[:10]]
        )
        context["chart_semi"] = json.dumps(
            [float(r["semi_skilled"]) for r in report_rows[:10]]
        )
        context["chart_general"] = json.dumps(
            [float(r["general"]) for r in report_rows[:10]]
        )
        return context


class PlantComponentReportView(_ComponentReportBase):
    """Aggregated plant-type hours needed across the whole project."""

    template_name = "estimator/reports/plant_components_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        config = self._config()
        qty_field = config["qty_field"]
        project = self.get_project()
        variant = self.kwargs["variant"]

        # SUM(comp.hours * boq_item.<qty_field>) per plant_type, GROUP BY plant_type.
        # NULL qty rows contribute NULL → ignored by Sum, matching the original
        # `if not raw_qty: continue` semantics.
        aggs = (
            ProjectPlantSpecificationComponent.objects.filter(
                specification__boq_items__project=project,
                specification__boq_items__is_section_header=False,
                plant_type__isnull=False,
            )
            .values(
                "plant_type_id",
                "plant_type__name",
                "plant_type__hourly_rate",
            )
            .annotate(
                total_hours=Sum(
                    F("hours") * F(f"specification__boq_items__{qty_field}"),
                    output_field=DecimalField(max_digits=18, decimal_places=4),
                ),
            )
        )

        report_rows = []
        grand_total = Decimal("0")
        for agg in aggs:
            hours = agg["total_hours"]
            # Mirror the original `if not raw_qty: continue` — no contribution → no row.
            if not hours:
                continue
            rate = agg["plant_type__hourly_rate"] or Decimal("0")
            amount = hours * rate
            grand_total += amount
            report_rows.append(
                {
                    "name": agg["plant_type__name"],
                    "hours": hours,
                    "rate": rate,
                    "amount": amount,
                }
            )
        report_rows.sort(key=lambda r: r["amount"], reverse=True)
        for row in report_rows:
            row["pct_of_total"] = calculate_pct_of_total(row["amount"], grand_total)

        context["report_title"] = f"{config['title_prefix']} Plant by Component"
        context["parent_template"] = "estimator/base_plant_estimator.html"
        context["report_rows"] = report_rows
        context["grand_total"] = grand_total
        context["total_hours"] = sum((r["hours"] for r in report_rows), Decimal("0"))
        context["variant"] = variant
        context["project_pk"] = project.pk

        top = report_rows[:10]
        context["chart_labels"] = json.dumps([r["name"] for r in top])
        context["chart_values"] = json.dumps([float(r["amount"]) for r in top])
        return context


class _SimpleSpecCalculatorView(ProjectEstimatorMixin, TemplateView):
    """Shared base: Plant/Prelim calculator pages grouping BOQItems by spec."""

    spec_field = ""
    spec_rate_attr = ""
    parent_template = ""
    title_noun = ""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()
        f_section = self.request.GET.get("section", "")
        f_name = self.request.GET.get("name", "")

        qs = BOQItem.objects.filter(
            project=project,
            is_section_header=False,
            **{
                f"{self.spec_field}__isnull": False,
                f"{self.spec_field}__is_active": True,
            },
        ).select_related(self.spec_field)
        if f_section:
            qs = qs.filter(section=f_section)
        if f_name:
            qs = qs.filter(**{f"{self.spec_field}__name": f_name})

        from collections import OrderedDict

        groups = OrderedDict()
        spec_id_field = f"{self.spec_field}_id"
        order_field = f"{self.spec_field}__name"
        for boq in qs.order_by("section", order_field):
            spec = getattr(boq, self.spec_field)
            key = (boq.section, getattr(boq, spec_id_field, None))
            if key not in groups:
                groups[key] = {
                    "section": boq.section,
                    "spec": spec,
                    "boq_qty": Decimal("0"),
                }
            if boq.contract_quantity:
                groups[key]["boq_qty"] += boq.contract_quantity

        rows = []
        for group in groups.values():
            spec = group["spec"]
            boq_qty = group["boq_qty"]
            rate = getattr(spec, self.spec_rate_attr) or Decimal("0")
            rows.append(
                {
                    "section": group["section"],
                    "spec": spec,
                    "boq_qty": boq_qty,
                    "rate": rate,
                    "amount": rate * boq_qty if boq_qty else Decimal("0"),
                }
            )
        context["rows"] = rows
        context["parent_template"] = self.parent_template
        context["title_noun"] = self.title_noun

        boq_with = BOQItem.objects.filter(
            project=project,
            is_section_header=False,
            **{
                f"{self.spec_field}__isnull": False,
                f"{self.spec_field}__is_active": True,
            },
        )
        context["sections"] = (
            boq_with.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["names"] = (
            boq_with.values_list(f"{self.spec_field}__name", flat=True)
            .distinct()
            .order_by(f"{self.spec_field}__name")
        )
        context["f_section"] = f_section
        context["f_name"] = f_name
        return context


class PlantSpecificationListView(_SimpleSpecCalculatorView):
    template_name = "estimator/plant_specification_list.html"
    spec_field = "plant_specification"
    spec_rate_attr = "rate_per_unit"
    parent_template = "estimator/base_plant_estimator.html"
    title_noun = "Plant"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        for row in context["rows"]:
            spec = row["spec"]
            boq_qty = row["boq_qty"] or Decimal("0")
            totals = []
            if spec:
                for comp in spec.components.all():
                    if comp.plant_type is None:
                        continue
                    total_hours = (
                        boq_qty * comp.hours if boq_qty and comp.hours else Decimal("0")
                    )
                    totals.append(
                        {
                            "name": comp.plant_type.name,
                            "total_hours": total_hours,
                        }
                    )
            row["component_totals"] = totals
        return context


class PreliminarySpecificationListView(_SimpleSpecCalculatorView):
    template_name = "estimator/preliminary_specification_list.html"
    spec_field = "preliminary_specification"
    spec_rate_attr = "amount"
    parent_template = "estimator/base_prelim_estimator.html"
    title_noun = "Preliminary"


@method_decorator(csrf_exempt, name="dispatch")
class UpdateBoqItemView(View):
    """AJAX endpoint to update FK fields or decimal markup fields on a BOQItem."""

    ALLOWED_FIELDS = {
        "trade_code": (ProjectTradeCode, "trade_code"),
        "specification": (ProjectSpecification, "specification"),
        "labour_specification": (ProjectLabourSpecification, "labour_specification"),
        "plant_specification": (ProjectPlantSpecification, "plant_specification"),
        "preliminary_specification": (
            ProjectPreliminarySpecification,
            "preliminary_specification",
        ),
    }

    DECIMAL_FIELDS = {
        "material_markup_pct",
        "labour_markup_pct",
        "transport_pct",
        "progress_quantity",
        "forecast_quantity",
    }

    def post(self, request, project_pk, pk):
        item = get_object_or_404(BOQItem, pk=pk, project_id=project_pk)

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field in self.DECIMAL_FIELDS:
            try:
                setattr(item, field, Decimal(str(value or 0)))
            except Exception:
                return JsonResponse({"error": "Invalid value"}, status=400)
            item.save()
            return self._build_response(item)

        if field == "labour_plant_specification":
            if value is None or value == "" or value == 0:
                item.labour_specification = None
                item.plant_specification = None
            else:
                name = str(value)
                item.labour_specification = (
                    ProjectLabourSpecification.objects.filter(
                        name=name, section=item.section, project_id=project_pk
                    ).first()
                    or ProjectLabourSpecification.objects.filter(
                        name=name, project_id=project_pk
                    ).first()
                )
                item.plant_specification = (
                    ProjectPlantSpecification.objects.filter(
                        name=name, section=item.section, project_id=project_pk
                    ).first()
                    or ProjectPlantSpecification.objects.filter(
                        name=name, project_id=project_pk
                    ).first()
                )
            item.save()
            return self._build_response(item)

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        model_cls, attr_name = self.ALLOWED_FIELDS[field]

        if value is None or value == "" or value == 0:
            setattr(item, attr_name, None)
        else:
            try:
                # Try ID-based lookup first, scoped to project
                related_obj = model_cls.objects.get(
                    pk=int(value), project_id=project_pk
                )
                setattr(item, attr_name, related_obj)
            except (ValueError, TypeError):
                # Name-based lookup — find the spec matching the item's section
                try:
                    related_obj = model_cls.objects.filter(
                        name=str(value),
                        section=item.section,
                        project_id=project_pk,
                    ).first()
                    if not related_obj:
                        related_obj = model_cls.objects.filter(
                            name=str(value),
                            project_id=project_pk,
                        ).first()
                    if related_obj:
                        setattr(item, attr_name, related_obj)
                    else:
                        return JsonResponse(
                            {"error": f'{field} "{value}" not found'}, status=404
                        )
                except Exception:
                    return JsonResponse(
                        {"error": f'{field} "{value}" not found'}, status=404
                    )
            except model_cls.DoesNotExist:
                return JsonResponse(
                    {"error": f"{field} with id={value} not found"}, status=404
                )

        item.save()
        return self._build_response(item)

    @staticmethod
    def _build_response(item):
        def fmt(val):
            if val is None:
                return None
            return format_num(val)

        return JsonResponse(
            {
                "ok": True,
                "new_materials_rate": fmt(item.new_materials_rate),
                "new_labour_rate": fmt(item.new_labour_rate),
                "new_plant_rate": fmt(item.new_plant_rate),
                "new_preliminary_rate": fmt(item.new_preliminary_rate),
                "baseline_new_price": fmt(item.baseline_new_price),
                "progress_amount": fmt(item.progress_amount),
                "forecast_amount": fmt(item.forecast_amount),
                "new_materials_amount": fmt(item.new_materials_amount),
                "new_labour_amount": fmt(item.new_labour_amount),
                "new_plant_amount": fmt(item.new_plant_amount),
                "new_preliminary_amount": fmt(item.new_preliminary_amount),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class BulkUpdateBoqSpecsView(View):
    """AJAX endpoint to apply one or more spec FKs to many BOQItems at once.

    Payload: {"ids": [int, ...], "updates": {field: value, ...}}
    Allowed fields mirror UpdateBoqItemView; only fields present in `updates`
    are written. A value of "" (or null) clears that FK on all selected items.
    """

    FK_FIELDS = {
        "trade_code": ProjectTradeCode,
        "specification": ProjectSpecification,
        "preliminary_specification": ProjectPreliminarySpecification,
    }

    def post(self, request, project_pk):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        ids = data.get("ids") or []
        updates = data.get("updates") or {}

        if not isinstance(ids, list) or not ids:
            return JsonResponse({"error": "No items selected"}, status=400)
        if not isinstance(updates, dict) or not updates:
            return JsonResponse({"error": "No fields to update"}, status=400)

        items = list(
            BOQItem.objects.filter(
                pk__in=ids, project_id=project_pk, is_section_header=False
            )
        )
        if not items:
            return JsonResponse({"error": "No matching items"}, status=404)

        for item in items:
            for field, value in updates.items():
                cleared = value is None or value == ""

                if field == "labour_plant_specification":
                    if cleared:
                        item.labour_specification = None
                        item.plant_specification = None
                    else:
                        name = str(value)
                        item.labour_specification = (
                            ProjectLabourSpecification.objects.filter(
                                name=name, section=item.section, project_id=project_pk
                            ).first()
                            or ProjectLabourSpecification.objects.filter(
                                name=name, project_id=project_pk
                            ).first()
                        )
                        item.plant_specification = (
                            ProjectPlantSpecification.objects.filter(
                                name=name, section=item.section, project_id=project_pk
                            ).first()
                            or ProjectPlantSpecification.objects.filter(
                                name=name, project_id=project_pk
                            ).first()
                        )
                elif field in self.FK_FIELDS:
                    model_cls = self.FK_FIELDS[field]
                    if cleared:
                        setattr(item, field, None)
                    else:
                        related = None
                        try:
                            related = model_cls.objects.filter(
                                pk=int(value), project_id=project_pk
                            ).first()
                        except (TypeError, ValueError):
                            pass
                        if related is None:
                            related = (
                                model_cls.objects.filter(
                                    name=str(value),
                                    section=item.section,
                                    project_id=project_pk,
                                ).first()
                                or model_cls.objects.filter(
                                    name=str(value), project_id=project_pk
                                ).first()
                            )
                        setattr(item, field, related)
                else:
                    return JsonResponse(
                        {"error": f'Field "{field}" not allowed'}, status=400
                    )
            item.save()

        return JsonResponse({"ok": True, "updated": len(items)})


@method_decorator(csrf_exempt, name="dispatch")
class BulkMarkupUpdateView(View):
    """AJAX endpoint to bulk-update a markup field on all non-header BOQItems."""

    ALLOWED_FIELDS = {"material_markup_pct", "labour_markup_pct", "transport_pct"}

    def post(self, request, project_pk):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        try:
            decimal_value = Decimal(str(value or 0))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        BOQItem.objects.filter(project_id=project_pk, is_section_header=False).update(
            **{field: decimal_value}
        )
        return JsonResponse({"ok": True})


@method_decorator(csrf_exempt, name="dispatch")
class UpdateMaterialView(View):
    """AJAX endpoint to update pack_qty / pack_cost on a ProjectMaterial.

    market_rate is auto-recomputed by the model save() and returned in the
    response so the read-only Rate column updates without a page reload.
    """

    ALLOWED_FIELDS = {
        "pack_qty",
        "pack_cost",
        "trade_name",
        "material_code",
        "unit",
        "material_variety",
        "market_spec",
    }

    def post(self, request, project_pk, pk):
        item = get_object_or_404(ProjectMaterial, pk=pk, project_id=project_pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        try:
            if field in ("pack_qty", "pack_cost"):
                setattr(item, field, Decimal(str(value or 0)))
            else:
                setattr(item, field, str(value or ""))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse(
            {
                "ok": True,
                "pack_qty": format_num(item.pack_qty),
                "pack_cost": format_num(item.pack_cost),
                "market_rate": format_num(item.market_rate),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSpecComponentView(View):
    """AJAX endpoint to update fields on a ProjectSpecificationComponent."""

    ALLOWED_FIELDS = {"qty_per_unit", "material", "label"}

    def post(self, request, project_pk, pk):
        item = get_object_or_404(ProjectSpecificationComponent, pk=pk)
        if item.specification.project_id != int(project_pk):
            return JsonResponse({"error": "Not found"}, status=404)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        try:
            if field == "qty_per_unit":
                item.qty_per_unit = Decimal(str(value or 0))
            elif field == "label":
                item.label = (value or "").strip()
            else:  # material
                if value in (None, "", 0, "0"):
                    item.material = None
                else:
                    material = ProjectMaterial.objects.filter(
                        pk=int(value), project_id=project_pk
                    ).first()
                    if material is None:
                        return JsonResponse({"error": "Material not found"}, status=404)
                    item.material = material
                    item.label = material.material_code
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        spec = item.specification
        return JsonResponse(
            {
                "ok": True,
                "qty_per_unit": format_num(item.qty_per_unit),
                "label": item.label,
                "material_id": item.material_id,
                "spec_id": spec.id,
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class AddSpecComponentView(View):
    """AJAX endpoint to add a new component to a ProjectSpecification."""

    def post(self, request, project_pk, pk):
        spec = get_object_or_404(ProjectSpecification, pk=pk, project_id=project_pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}

        material_id = data.get("material")
        label = (data.get("label") or "").strip()
        qty_raw = data.get("qty_per_unit", 0)

        material = None
        if material_id not in (None, "", 0, "0"):
            try:
                material = ProjectMaterial.objects.filter(
                    pk=int(material_id), project_id=project_pk
                ).first()
            except (TypeError, ValueError):
                material = None

        if not label and material is not None:
            label = material.material_code

        try:
            qty = Decimal(str(qty_raw or 0))
        except Exception:
            qty = Decimal("0")

        next_order = (
            spec.spec_components.aggregate(models.Max("sort_order"))["sort_order__max"]
            or 0
        ) + 1
        comp = ProjectSpecificationComponent.objects.create(
            specification=spec,
            material=material,
            label=label,
            qty_per_unit=qty,
            sort_order=next_order,
        )
        return JsonResponse(
            {
                "ok": True,
                "component": {
                    "id": comp.id,
                    "material_id": comp.material_id,
                    "label": comp.label,
                    "qty_per_unit": format_num(comp.qty_per_unit),
                },
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteSpecComponentView(View):
    """AJAX endpoint to delete a ProjectSpecificationComponent."""

    def post(self, request, project_pk, pk):
        comp = get_object_or_404(
            ProjectSpecificationComponent,
            pk=pk,
            specification__project_id=project_pk,
        )
        spec = comp.specification
        comp.delete()
        return JsonResponse(
            {
                "ok": True,
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteSpecificationView(View):
    """AJAX endpoint to delete a ProjectSpecification (material spec)."""

    def post(self, request, project_pk, pk):
        spec = get_object_or_404(ProjectSpecification, pk=pk, project_id=project_pk)
        spec.delete()
        return JsonResponse({"ok": True})


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSpecificationView(View):
    """AJAX endpoint to update fields on a ProjectSpecification (material spec)."""

    ALLOWED_FIELDS = {
        "section": "str",
        "name": "str",
        "unit_label": "str",
        "is_active": "bool",
    }

    def post(self, request, project_pk, pk):
        item = get_object_or_404(ProjectSpecification, pk=pk, project_id=project_pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True})


class SystemMaterialSpecListView(ProjectEstimatorMixin, ListView):
    """List System Specifications with form to add new spec (project-scoped view)."""

    model = SystemSpecification
    template_name = "estimator/system_spec_list.html"
    context_object_name = "specs"

    def get_queryset(self):
        return (
            SystemSpecification.objects.select_related("trade_code")
            .prefetch_related("spec_components__material")
            .all()
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["spec_form"] = context.get("spec_form", SystemSpecificationForm())
        context["component_formset"] = context.get(
            "component_formset", SystemSpecificationComponentFormSet()
        )
        context["names"] = (
            SystemSpecification.objects.values_list("name", flat=True)
            .distinct()
            .order_by("name")
        )
        context["f_name"] = self.request.GET.get("name", "")
        return context

    def post(self, request, *args, **kwargs):
        spec_form = SystemSpecificationForm(request.POST)
        component_formset = SystemSpecificationComponentFormSet(request.POST)
        if spec_form.is_valid():
            spec = spec_form.save()
            component_formset = SystemSpecificationComponentFormSet(
                request.POST, instance=spec
            )
            if component_formset.is_valid():
                component_formset.save()
                return redirect(
                    reverse(
                        "estimator:system_specs",
                        kwargs={"project_pk": self.kwargs["project_pk"]},
                    )
                )
            else:
                spec.delete()
        self.object_list = self.get_queryset()
        return self.render_to_response(
            self.get_context_data(
                spec_form=spec_form, component_formset=component_formset
            )
        )


class UpdateSystemSpecComponentView(View):
    """AJAX endpoint to update qty_per_unit on a SystemSpecificationComponent."""

    def post(self, request, project_pk, pk):
        item = get_object_or_404(SystemSpecificationComponent, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field != "qty_per_unit":
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        try:
            item.qty_per_unit = Decimal(str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        spec = item.specification
        return JsonResponse(
            {
                "ok": True,
                "qty_per_unit": format_num(item.qty_per_unit),
                "spec_id": spec.id,
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


class SystemSpecUploadView(ProjectEstimatorMixin, FormView):
    """Upload a System Specs library from Excel template."""

    template_name = "estimator/system_spec_upload.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:system_specs", kwargs={"project_pk": self.kwargs["project_pk"]}
        )

    def form_valid(self, form):
        uploaded = form.cleaned_data["file"]

        # Write uploaded file to a temp file for openpyxl
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        try:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp.close()

            from .management.commands.import_excel import SystemSpecImporter

            importer = SystemSpecImporter(tmp.name)
            results = importer.run()

            specs_result = results.get("system_specs", {})
            created = specs_result.get("created", 0)
            updated = specs_result.get("updated", 0)

            msg_parts = [f"{created} created, {updated} updated"]
            messages.success(
                self.request,
                f"System Specs uploaded successfully — {', '.join(msg_parts)}",
            )
        except Exception as e:
            messages.error(self.request, f"Import failed: {str(e)}")
        finally:
            os.unlink(tmp.name)

        return redirect(self.get_success_url())


class DownloadSystemSpecTemplateView(View):
    """Download the System Specs template Excel file."""

    def get(self, request, project_pk):
        template_path = Path(__file__).parent / "data" / "SystemSpec_Template.xlsx"

        if not os.path.exists(template_path):
            messages.error(request, "Template file not found")
            return redirect(
                reverse("estimator:system_specs", kwargs={"project_pk": project_pk})
            )

        response = FileResponse(open(template_path, "rb"), as_attachment=True)
        response["Content-Disposition"] = (
            'attachment; filename="SystemSpec_Template.xlsx"'
        )
        response["Content-Type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return response


@method_decorator(csrf_exempt, name="dispatch")
class UpdateLabourSpecView(View):
    """AJAX endpoint to update fields on a ProjectLabourSpecification."""

    ALLOWED_FIELDS = {
        "section": "str",
        "trade_name": "str",
        "name": "str",
        "unit": "str",
        "crew": "fk",
        "daily_production": "decimal",
        "team_mix": "decimal",
        "site_factor": "decimal",
        "tools_factor": "decimal",
        "leadership_factor": "decimal",
        "is_active": "bool",
    }

    def post(self, request, project_pk, pk):
        item = get_object_or_404(
            ProjectLabourSpecification, pk=pk, project_id=project_pk
        )
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            elif field_type == "fk":
                if value == "" or value is None:
                    item.crew = None
                else:
                    item.crew = get_object_or_404(
                        ProjectLabourCrew, pk=int(value), project_id=project_pk
                    )
            elif field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()

        def fmt(val):
            if val is None:
                return None
            return format_num(val)

        return JsonResponse(
            {
                "ok": True,
                "daily_output": fmt(item.daily_output),
                "rate_per_unit": fmt(item.rate_per_unit),
                "daily_cost": fmt(item.daily_cost),
                "total_cost": fmt(item.total_cost),
                "crew_type": item.crew.crew_type if item.crew else None,
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class UpdateLabourCrewView(View):
    """AJAX endpoint to update fields on a ProjectLabourCrew."""

    ALLOWED_FIELDS = {
        "crew_type": "str",
        "skilled": "int",
        "semi_skilled": "int",
        "general": "int",
        "skilled_rate": "decimal",
        "semi_skilled_rate": "decimal",
        "general_rate": "decimal",
    }

    def post(self, request, project_pk, pk):
        item = get_object_or_404(ProjectLabourCrew, pk=pk, project_id=project_pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "int":
                setattr(item, field, int(value))
            elif field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse(
            {
                "ok": True,
                "crew_size": item.crew_size,
                "crew_daily_cost": format_num(item.crew_daily_cost),
            }
        )


# ── Simple Labour Spec Definition View ────────────────────────────


class LabourSpecDefListView(ProjectEstimatorMixin, ListView):
    """Simple labour specifications view — Section, Trade Name, Name, Unit, Crew."""

    model = ProjectLabourSpecification
    template_name = "estimator/labour_spec_def_list.html"
    context_object_name = "labour_specs"
    paginate_by = 50

    def get_queryset(self):
        project = self.get_project()
        qs = ProjectLabourSpecification.objects.filter(project=project).select_related(
            "crew"
        )

        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)

        trade_name = self.request.GET.get("trade_name")
        if trade_name:
            qs = qs.filter(trade_name=trade_name)

        name = self.request.GET.get("name")
        if name:
            qs = qs.filter(name=name)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()
        context["form"] = context.get("form", LabourSpecificationForm(project=project))

        project_lspecs = ProjectLabourSpecification.objects.filter(project=project)
        context["sections"] = (
            project_lspecs.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_names"] = (
            project_lspecs.exclude(trade_name="")
            .values_list("trade_name", flat=True)
            .distinct()
            .order_by("trade_name")
        )
        context["names"] = (
            project_lspecs.values_list("name", flat=True).distinct().order_by("name")
        )

        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_name"] = self.request.GET.get("trade_name", "")
        context["f_name"] = self.request.GET.get("name", "")
        context["crews"] = ProjectLabourCrew.objects.filter(project=project)
        context["query_params"] = _pagination_query_params(self.request)

        context.update(_spec_datalist_context(project_lspecs))

        return context

    def post(self, request, *args, **kwargs):
        project = self.get_project()
        action = request.POST.get("action")
        qs = ProjectLabourSpecification.objects.filter(project=project)
        if _handle_clear_action(request, qs, label="labour specs"):
            return redirect(
                reverse(
                    "estimator:labour_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if _handle_bulk_action(request, qs, allow_toggle_active=True):
            return redirect(
                reverse(
                    "estimator:labour_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if action == "sync_system":
            from .services import sync_labour_specs_from_contractor

            result = sync_labour_specs_from_contractor(project)
            _flash_sync_result(request, result, "Labour specs")
            return redirect(
                reverse(
                    "estimator:labour_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        form = LabourSpecificationForm(request.POST, project=project)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.project = project
            obj.save()
            return redirect(
                reverse(
                    "estimator:labour_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


# ── Plant Costs ─────────────────────────────────────────────────


class PlantCostListView(ProjectEstimatorMixin, ListView):
    model = ProjectPlantCost
    template_name = "estimator/plant_costs_list.html"
    context_object_name = "plants"

    def get_queryset(self):
        return ProjectPlantCost.objects.filter(project=self.get_project())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", PlantCostForm())
        return context

    def post(self, request, *args, **kwargs):
        project_pk = self.kwargs["project_pk"]
        action = request.POST.get("action")
        qs = ProjectPlantCost.objects.filter(project=self.get_project())

        if _handle_clear_action(request, qs, label="plant costs"):
            return redirect(
                reverse("estimator:plant_costs", kwargs={"project_pk": project_pk})
            )

        if _handle_bulk_action(request, qs):
            return redirect(
                reverse("estimator:plant_costs", kwargs={"project_pk": project_pk})
            )

        if action == "sync_system":
            from .services import sync_plant_costs_from_contractor

            result = sync_plant_costs_from_contractor(self.get_project())
            _flash_sync_result(request, result, "Plant costs")
            return redirect(
                reverse("estimator:plant_costs", kwargs={"project_pk": project_pk})
            )

        form = PlantCostForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.project = self.get_project()
            obj.save()
            return redirect(
                reverse("estimator:plant_costs", kwargs={"project_pk": project_pk})
            )
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdatePlantCostView(View):
    """AJAX endpoint to update fields on a ProjectPlantCost."""

    ALLOWED_FIELDS = {
        "name": "str",
        "hourly_production": "decimal",
        "hourly_rate": "decimal",
    }

    def post(self, request, project_pk, pk):
        item = get_object_or_404(ProjectPlantCost, pk=pk, project_id=project_pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True})


# ── Plant Spec Definitions ─────────────────────────────────────


class PlantSpecDefListView(ProjectEstimatorMixin, ListView):
    model = ProjectPlantSpecification
    template_name = "estimator/plant_spec_def_list.html"
    context_object_name = "plant_specs"
    paginate_by = 50

    def get_queryset(self):
        project = self.get_project()
        qs = ProjectPlantSpecification.objects.filter(project=project).prefetch_related(
            "components__plant_type"
        )
        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)
        trade_name = self.request.GET.get("trade_name")
        if trade_name:
            qs = qs.filter(trade_name=trade_name)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_project()
        context["form"] = context.get("form", PlantSpecificationForm(project=project))
        project_pspecs = ProjectPlantSpecification.objects.filter(project=project)
        context["sections"] = (
            project_pspecs.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_names"] = (
            project_pspecs.exclude(trade_name="")
            .values_list("trade_name", flat=True)
            .distinct()
            .order_by("trade_name")
        )
        context["plants"] = ProjectPlantCost.objects.filter(project=project)
        context["labour_specs"] = ProjectLabourSpecification.objects.filter(
            project=project, is_active=True
        ).order_by("name")
        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_name"] = self.request.GET.get("trade_name", "")
        context["query_params"] = _pagination_query_params(self.request)
        context.update(_spec_datalist_context(project_pspecs))
        return context

    def post(self, request, *args, **kwargs):
        project = self.get_project()
        action = request.POST.get("action")
        qs = ProjectPlantSpecification.objects.filter(project=project)
        if _handle_clear_action(request, qs, label="plant specs"):
            return redirect(
                reverse(
                    "estimator:plant_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if _handle_bulk_action(request, qs, allow_toggle_active=True):
            return redirect(
                reverse(
                    "estimator:plant_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if action == "sync_system":
            from .services import sync_plant_specs_from_contractor

            result = sync_plant_specs_from_contractor(project)
            _flash_sync_result(request, result, "Plant specs")
            return redirect(
                reverse(
                    "estimator:plant_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        form = PlantSpecificationForm(request.POST, project=project)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.project = project
            obj.save()
            return redirect(
                reverse(
                    "estimator:plant_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdatePlantSpecView(View):
    """AJAX endpoint to update fields on a ProjectPlantSpecification."""

    ALLOWED_FIELDS = {
        "section": "str",
        "trade_name": "str",
        "name": "str",
        "unit": "str",
        "daily_production": "decimal",
        "operator_factor": "decimal",
        "site_factor": "decimal",
        "is_active": "bool",
    }

    def post(self, request, project_pk, pk):
        item = get_object_or_404(
            ProjectPlantSpecification, pk=pk, project_id=project_pk
        )
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            elif field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse(
            {
                "ok": True,
                "daily_output": str(item.daily_output),
                "daily_cost": str(item.daily_cost),
                "rate_per_unit": str(item.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class UpdatePlantSpecComponentView(View):
    """AJAX endpoint to update plant_type or hours on a ProjectPlantSpecificationComponent."""

    ALLOWED_FIELDS = {"plant_type", "hours"}

    def post(self, request, project_pk, pk):
        comp = get_object_or_404(
            ProjectPlantSpecificationComponent,
            pk=pk,
            specification__project_id=project_pk,
        )
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        try:
            if field == "hours":
                comp.hours = Decimal(str(value or 0))
            else:
                if value in (None, "", 0, "0"):
                    comp.plant_type = None
                else:
                    comp.plant_type = ProjectPlantCost.objects.filter(
                        pk=int(value), project_id=project_pk
                    ).first()
                    if comp.plant_type is None:
                        return JsonResponse(
                            {"error": "Plant type not found"}, status=404
                        )
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        comp.save()
        spec = comp.specification
        return JsonResponse(
            {
                "ok": True,
                "daily_cost": str(spec.daily_cost),
                "rate_per_unit": str(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class AddPlantSpecComponentView(View):
    """AJAX endpoint to add a new component to a ProjectPlantSpecification."""

    def post(self, request, project_pk, pk):
        spec = get_object_or_404(
            ProjectPlantSpecification, pk=pk, project_id=project_pk
        )
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}

        plant_type_id = data.get("plant_type")
        hours_raw = data.get("hours", 0)

        plant_type = None
        if plant_type_id not in (None, "", 0, "0"):
            try:
                plant_type = ProjectPlantCost.objects.filter(
                    pk=int(plant_type_id), project_id=project_pk
                ).first()
            except (TypeError, ValueError):
                plant_type = None

        try:
            hours = Decimal(str(hours_raw or 0))
        except Exception:
            hours = Decimal("0")

        next_order = (
            spec.components.aggregate(models.Max("sort_order"))["sort_order__max"] or 0
        ) + 1
        comp = ProjectPlantSpecificationComponent.objects.create(
            specification=spec,
            plant_type=plant_type,
            hours=hours,
            sort_order=next_order,
        )
        return JsonResponse(
            {
                "ok": True,
                "component": {
                    "id": comp.id,
                    "plant_type_id": comp.plant_type_id,
                    "plant_type_name": comp.plant_type.name if comp.plant_type else "",
                    "hours": str(comp.hours),
                },
                "daily_cost": str(spec.daily_cost),
                "rate_per_unit": str(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeletePlantSpecComponentView(View):
    """AJAX endpoint to delete a ProjectPlantSpecificationComponent."""

    def post(self, request, project_pk, pk):
        comp = get_object_or_404(
            ProjectPlantSpecificationComponent,
            pk=pk,
            specification__project_id=project_pk,
        )
        spec = comp.specification
        comp.delete()
        return JsonResponse(
            {
                "ok": True,
                "daily_cost": str(spec.daily_cost),
                "rate_per_unit": str(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeletePlantSpecificationView(View):
    """AJAX endpoint to delete a ProjectPlantSpecification."""

    def post(self, request, project_pk, pk):
        spec = get_object_or_404(
            ProjectPlantSpecification, pk=pk, project_id=project_pk
        )
        spec.delete()
        return JsonResponse({"ok": True})


# ── Preliminary Costs ──────────────────────────────────────────


class PreliminaryCostListView(ProjectEstimatorMixin, ListView):
    model = ProjectPreliminaryCost
    template_name = "estimator/preliminary_costs_list.html"
    context_object_name = "preliminaries"

    def get_queryset(self):
        qs = ProjectPreliminaryCost.objects.filter(project=self.get_project())
        ptype = self.request.GET.get("preliminary_type", "").strip()
        if ptype:
            qs = qs.filter(preliminary_type=ptype)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", PreliminaryCostForm())
        context["preliminary_types"] = ProjectPreliminaryCost._meta.get_field(
            "preliminary_type"
        ).choices
        context["f_q"] = self.request.GET.get("q", "")
        context["f_preliminary_type"] = self.request.GET.get("preliminary_type", "")
        return context

    def post(self, request, *args, **kwargs):
        project_pk = self.kwargs["project_pk"]
        action = request.POST.get("action")
        qs = ProjectPreliminaryCost.objects.filter(project=self.get_project())

        if _handle_clear_action(request, qs, label="preliminary costs"):
            return redirect(
                reverse(
                    "estimator:preliminary_costs",
                    kwargs={"project_pk": project_pk},
                )
            )

        if _handle_bulk_action(request, qs):
            return redirect(
                reverse(
                    "estimator:preliminary_costs",
                    kwargs={"project_pk": project_pk},
                )
            )

        if action == "sync_system":
            from .services import sync_preliminary_costs_from_contractor

            result = sync_preliminary_costs_from_contractor(self.get_project())
            _flash_sync_result(request, result, "Preliminary costs")
            return redirect(
                reverse(
                    "estimator:preliminary_costs",
                    kwargs={"project_pk": project_pk},
                )
            )

        form = PreliminaryCostForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.project = self.get_project()
            obj.save()
            return redirect(
                reverse(
                    "estimator:preliminary_costs",
                    kwargs={"project_pk": project_pk},
                )
            )
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdatePreliminaryCostView(View):
    """AJAX endpoint to update fields on a ProjectPreliminaryCost."""

    ALLOWED_FIELDS = {
        "name": "str",
        "preliminary_type": "str",
        "sum_value": "decimal",
        "amount": "decimal",
        "number_per_month": "decimal",
        "monthly_rate": "decimal",
        "months": "decimal",
    }

    def post(self, request, project_pk, pk):
        item = get_object_or_404(ProjectPreliminaryCost, pk=pk, project_id=project_pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse(
            {
                "ok": True,
                "computed_amount": str(item.computed_amount),
            }
        )


# ── Preliminary Spec Definitions ───────────────────────────────


class PreliminarySpecDefListView(ProjectEstimatorMixin, ListView):
    model = ProjectPreliminarySpecification
    template_name = "estimator/preliminary_spec_def_list.html"
    context_object_name = "preliminary_specs"

    def get_queryset(self):
        return ProjectPreliminarySpecification.objects.filter(
            project=self.get_project()
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get(
            "form", PreliminarySpecificationForm(project=self.get_project())
        )
        context["preliminary_type_choices"] = (
            SystemPreliminaryCost.PRELIMINARY_TYPE_CHOICES
        )
        context.update(
            _spec_datalist_context(
                ProjectPreliminarySpecification.objects.filter(
                    project=self.get_project()
                )
            )
        )
        return context

    def post(self, request, *args, **kwargs):
        project = self.get_project()
        action = request.POST.get("action")
        qs = ProjectPreliminarySpecification.objects.filter(project=project)
        if _handle_clear_action(request, qs, label="preliminary specs"):
            return redirect(
                reverse(
                    "estimator:preliminary_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if _handle_bulk_action(request, qs, allow_toggle_active=True):
            return redirect(
                reverse(
                    "estimator:preliminary_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if action == "sync_system":
            from .services import sync_preliminary_specs_from_contractor

            result = sync_preliminary_specs_from_contractor(project)
            _flash_sync_result(request, result, "Preliminary specs")
            return redirect(
                reverse(
                    "estimator:preliminary_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        form = PreliminarySpecificationForm(request.POST, project=project)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.project = project
            obj.save()
            return redirect(
                reverse(
                    "estimator:preliminary_spec_defs",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdatePreliminarySpecView(View):
    """AJAX endpoint to update fields on a ProjectPreliminarySpecification."""

    ALLOWED_FIELDS = {
        "section": "str",
        "trade_name": "str",
        "name": "str",
        "unit": "str",
        "preliminary_type": "str",
        "is_active": "bool",
    }

    def post(self, request, project_pk, pk):
        item = get_object_or_404(
            ProjectPreliminarySpecification, pk=pk, project_id=project_pk
        )
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            elif field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True, "amount": str(item.amount)})


# ── Upload / Download Template Views ──────────────────────────────


def _handle_upload(
    request, importer_class, success_url, entity_name, project=None, company=None
):
    """Shared logic for all upload views."""
    uploaded = request.FILES.get("file")
    if not uploaded:
        messages.error(request, "No file provided.")
        return redirect(success_url)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        for chunk in uploaded.chunks():
            tmp.write(chunk)
        tmp.close()

        importer = importer_class(tmp.name, project=project, company=company)
        result = importer.run()

        created = result.get("created", 0)
        updated = result.get("updated", 0)
        msg = f"{entity_name} uploaded — {created} created, {updated} updated"
        sheet_used = result.get("sheet_used")
        if sheet_used is not None and result.get("fell_back"):
            msg += f" [sheet used: '{sheet_used}' (fallback — no sheet name matched)]"
        messages.success(request, msg)
    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
    finally:
        os.unlink(tmp.name)

    return redirect(success_url)


def _generate_template(headers, filename):
    """Generate an Excel template styled to match the master workbook.

    Returns HttpResponse (not FileResponse/BytesIO-streaming) because some
    WSGI/proxy setups strip Content-Length from streaming responses and the
    client ends up with a 500.
    """
    import io

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    header_font = Font(name="Aptos Narrow", size=11, bold=True)
    data_font = Font(name="Aptos Narrow", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    medium = Side(style="medium", color="000000")
    thin = Side(style="thin", color="000000")
    last_col = len(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center
        cell.border = Border(
            top=medium,
            bottom=medium,
            left=medium if col_idx == 1 else thin,
            right=medium if col_idx == last_col else thin,
        )
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 14)

    # Reserve 20 empty data rows framed by the same outer border so the user
    # sees a clean table outline matching the uploaded sheets.
    data_row_count = 20
    for r in range(2, 2 + data_row_count):
        is_last_row = r == 1 + data_row_count
        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = data_font
            cell.alignment = center
            cell.border = Border(
                top=thin,
                bottom=medium if is_last_row else thin,
                left=medium if col_idx == 1 else thin,
                right=medium if col_idx == last_col else thin,
            )

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    data = buffer.getvalue()

    response = HttpResponse(
        data,
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Content-Length"] = str(len(data))
    return response


class CloneFromProjectMixin:
    """Adds a "clone from another project (same contractor)" option to
    project-level upload pages.

    The clone copies the *entire* estimator library from the chosen source
    project (replacing this project's existing library data — not its BoQ
    items). Source projects are restricted to the same contractor.
    """

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        project = self.get_project()
        if project.contractor_id:
            ctx["clone_source_projects"] = (
                Project.objects.filter(contractor_id=project.contractor_id)
                .exclude(pk=project.pk)
                .order_by("name")
            )
        else:
            ctx["clone_source_projects"] = Project.objects.none()
        return ctx

    def post(self, request, *args, **kwargs):
        if request.POST.get("action") == "clone_project":
            from .services import clone_from_project

            project = self.get_project()
            source_pk = request.POST.get("source_project")
            if not source_pk:
                messages.error(request, "Please select a project to clone from.")
            elif not project.contractor_id:
                messages.error(
                    request,
                    "This project has no contractor assigned, so there are "
                    "no related projects to clone from.",
                )
            else:
                source = get_object_or_404(
                    Project,
                    pk=source_pk,
                    contractor_id=project.contractor_id,
                )
                try:
                    result = clone_from_project(project, source)
                    messages.success(
                        request,
                        f"Cloned the full estimator library from "
                        f"'{source.name}' — "
                        f"{result.get('trade_codes', 0)} trade codes, "
                        f"{result.get('materials', 0)} materials, "
                        f"{result.get('labour_crews', 0)} labour crews, "
                        f"{result.get('specifications', 0)} specifications, "
                        f"{result.get('labour_specs', 0)} labour specs, "
                        f"{result.get('plant_specs', 0)} plant specs, "
                        f"{result.get('preliminary_specs', 0)} prelim specs.",
                    )
                except Exception as e:
                    messages.error(request, f"Clone from project failed: {e}")
            return redirect(self.get_success_url())
        return super().post(request, *args, **kwargs)


# ── Trade Codes Upload/Download ───────────────────────────────────


class TradeCodeUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:trade_codes", kwargs={"project_pk": self.kwargs["project_pk"]}
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base_baseline_costs.html"
        ctx["upload_title"] = "Upload Trade Codes"
        ctx["upload_description"] = (
            "Upload trade code prefixes and names from an Excel template."
        )
        ctx["download_url_name"] = "estimator:download_trade_code_template"
        ctx["columns"] = [
            ("Prefix", "Trade code prefix (e.g. CFR, PLB) — required, unique key"),
            ("Trade Name", "Trade name (e.g. Concrete, Plumbing)"),
        ]
        ctx["notes"] = [
            "Prefix is the unique key — existing trade codes with the same prefix will be updated.",
        ]
        return ctx

    def form_valid(self, form):
        from .importers import TradeCodeImporter

        return _handle_upload(
            self.request,
            TradeCodeImporter,
            self.get_success_url(),
            "Trade Codes",
            project=self.get_project(),
        )


class DownloadTradeCodeTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(["Prefix", "Trade Name"], "TradeCode_Template.xlsx")


# ── Material Costs Upload/Download ────────────────────────────────


class MaterialCostUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:materials", kwargs={"project_pk": self.kwargs["project_pk"]}
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base_baseline_costs.html"
        ctx["upload_title"] = "Upload Material Costs"
        ctx["upload_description"] = (
            "Upload material pricing data from an Excel template."
        )
        ctx["download_url_name"] = "estimator:download_material_cost_template"
        ctx["columns"] = [
            ("Trade Name", "Trade group (e.g. CFR-Concrete)"),
            ("Material Code", "Unique material code (required)"),
            ("Unit", "Consumption unit (e.g. Bag, m3, brick, L)"),
            ("Pack Qty", "Pack size that the price covers (e.g. 1000 for 1000 bricks)"),
            ("Pack Cost", "Cost for the whole pack (e.g. 2800 for 1000 bricks)"),
            ("Material Variety", "Variety description"),
            ("Market Spec", "Market specification"),
        ]
        ctx["notes"] = [
            "Material Code is the unique key — existing materials with the same code will be updated.",
            "Effective rate per unit is auto-computed as Pack Cost ÷ Pack Qty.",
            "Pack Qty defaults to 1 if blank — older templates with a single 'Market Rate' column are still supported.",
        ]
        return ctx

    def form_valid(self, form):
        from .importers import MaterialCostImporter

        return _handle_upload(
            self.request,
            MaterialCostImporter,
            self.get_success_url(),
            "Material Costs",
            project=self.get_project(),
        )


class DownloadMaterialCostTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(
            [
                "Trade Name",
                "Material Code",
                "Unit",
                "Pack Qty",
                "Pack Cost",
                "Material Variety",
                "Market Spec",
            ],
            "MaterialCost_Template.xlsx",
        )


# ── Labour Costs Upload/Download ─────────────────────────────────


class LabourCostUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:labour_costs", kwargs={"project_pk": self.kwargs["project_pk"]}
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base_baseline_costs.html"
        ctx["upload_title"] = "Upload Labour Costs"
        ctx["upload_description"] = "Upload labour crew compositions and daily rates."
        ctx["download_url_name"] = "estimator:download_labour_cost_template"
        ctx["columns"] = [
            ("Crew Type", "Unique crew identifier (required)"),
            (
                "Crew Size",
                "Ignored on import — auto-computed as Skilled + Semi Skilled + General",
            ),
            ("Skilled", "Number of skilled workers"),
            ("Semi Skilled", "Number of semi-skilled workers"),
            ("General", "Number of general workers"),
            ("Skilled Rate", "Daily rate for skilled workers (R)"),
            ("Semi Skilled Rate", "Daily rate for semi-skilled workers (R)"),
            ("General Rate", "Daily rate for general workers (R)"),
        ]
        ctx["notes"] = [
            "Crew Type is the unique key — existing crews with the same type will be updated.",
            "Crew Size is auto-computed from Skilled + Semi Skilled + General; the column is kept for backwards compatibility but its value is ignored.",
            "Worker counts must be whole numbers. Rates must be numeric.",
        ]
        return ctx

    def form_valid(self, form):
        from .importers import LabourCostImporter

        return _handle_upload(
            self.request,
            LabourCostImporter,
            self.get_success_url(),
            "Labour Costs",
            project=self.get_project(),
        )


class DownloadLabourCostTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(
            [
                "Crew Type",
                "Crew Size",
                "Skilled",
                "Semi Skilled",
                "General",
                "Skilled Rate",
                "Semi Skilled Rate",
                "General Rate",
            ],
            "LabourCost_Template.xlsx",
        )


# ── Material Specs Upload/Download ───────────────────────────────


class MaterialSpecUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:material_specs", kwargs={"project_pk": self.kwargs["project_pk"]}
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base_materials_estimator.html"
        ctx["upload_title"] = "Upload Material Specifications"
        ctx["upload_description"] = (
            "Upload material specification definitions with component breakdowns."
        )
        ctx["download_url_name"] = "estimator:download_material_spec_template"
        ctx["columns"] = [
            (
                "Spec Name",
                "Specification name (required) — rows with the same name form one spec",
            ),
            ("Section", "Section grouping"),
            ("Trade Code Prefix", "Trade code prefix to link"),
            ("Unit", "Unit of measurement (e.g. m3)"),
            ("Material Code", "Material code for this component"),
            ("Label", "Component label (defaults to material code)"),
            ("Qty per Unit", "Quantity of material per unit of spec"),
        ]
        ctx["notes"] = [
            "Multi-row format: multiple rows with the same Spec Name define one specification's components.",
            "Section, Trade Code Prefix, and Unit are taken from the first row of each spec.",
            "All referenced materials must exist in Material Costs before uploading.",
            "Existing specs with the same name will have their components replaced.",
        ]
        return ctx

    def form_valid(self, form):
        from .importers import MaterialSpecImporter

        return _handle_upload(
            self.request,
            MaterialSpecImporter,
            self.get_success_url(),
            "Material Specifications",
            project=self.get_project(),
        )


class DownloadMaterialSpecTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(
            [
                "Spec Name",
                "Section",
                "Trade Code Prefix",
                "Unit",
                "Material Code",
                "Label",
                "Qty per Unit",
            ],
            "MaterialSpec_Template.xlsx",
        )


# ── Labour Specs Upload/Download ─────────────────────────────────


class LabourSpecUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:labour_spec_defs",
            kwargs={"project_pk": self.kwargs["project_pk"]},
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base_labour_estimator.html"
        ctx["upload_title"] = "Upload Labour Specifications"
        ctx["upload_description"] = (
            "Upload labour specification definitions with crew assignments and factors."
        )
        ctx["download_url_name"] = "estimator:download_labour_spec_template"
        ctx["columns"] = [
            ("Section", "Section grouping"),
            ("Trade Name", "Trade name"),
            ("Name", "Specification name (required, unique key)"),
            ("Unit", "Unit of measurement (e.g. m3)"),
            ("Crew Type", "Crew type to link (must exist in Labour Costs)"),
            ("Daily Production", "Base daily production rate"),
            ("Team Mix", "Team mix factor (default 1)"),
            ("Site Factor", "Site condition factor (default 1)"),
            ("Tools Factor", "Tools & equipment factor (default 1)"),
            ("Leadership Factor", "Leadership factor (default 1)"),
        ]
        ctx["notes"] = [
            "Name is the unique key — existing specs with the same name will be updated.",
            "All referenced Crew Types must exist in Labour Costs before uploading.",
            "Factor fields default to 1 if left blank.",
        ]
        return ctx

    def form_valid(self, form):
        from .importers import LabourSpecImporter

        return _handle_upload(
            self.request,
            LabourSpecImporter,
            self.get_success_url(),
            "Labour Specifications",
            project=self.get_project(),
        )


class DownloadLabourSpecTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(
            [
                "Section",
                "Trade Name",
                "Name",
                "Unit",
                "Crew Type",
                "Daily Production",
                "Team Mix",
                "Site Factor",
                "Tools Factor",
                "Leadership Factor",
            ],
            "LabourSpec_Template.xlsx",
        )


# ── Initialize Estimator ─────────────────────────────────────────


class InitializeEstimatorView(ProjectEstimatorMixin, View):
    """Clone system library into project-scoped records."""

    def post(self, request, project_pk):
        from .services import initialize_project_estimator

        project = self.get_project()
        try:
            result = initialize_project_estimator(project)
            messages.success(
                request,
                (
                    f"Estimator initialized — "
                    f"{result['trade_codes']} trade codes, "
                    f"{result['materials']} materials, "
                    f"{result['labour_crews']} labour crews, "
                    f"{result['specifications']} specifications, "
                    f"{result['labour_specs']} labour specs"
                ),
            )
        except Exception as e:
            messages.error(request, f"Initialization failed: {e}")
        return redirect(
            reverse("estimator:dashboard", kwargs={"project_pk": project_pk})
        )


# ── Plant Costs Upload/Download ───────────────────────────────────


class PlantCostUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:plant_costs", kwargs={"project_pk": self.kwargs["project_pk"]}
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base_baseline_costs.html"
        ctx["upload_title"] = "Upload Plant Costs"
        ctx["upload_description"] = (
            "Upload plant & equipment costs with hourly production and rates."
        )
        ctx["download_url_name"] = "estimator:download_plant_cost_template"
        ctx["columns"] = [
            ("Plant & Equipment", "Name of plant/equipment"),
            ("Hourly Production", "Production output per hour"),
            ("Hourly Rate", "Cost per hour (R)"),
        ]
        return ctx

    def form_valid(self, form):
        from .importers import PlantCostImporter

        return _handle_upload(
            self.request,
            PlantCostImporter,
            self.get_success_url(),
            "Plant Costs",
            project=self.get_project(),
        )


class DownloadPlantCostTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(
            ["Plant & Equipment", "Hourly Production", "Hourly Rate"],
            "PlantCost_Template.xlsx",
        )


# ── Plant Specs Upload/Download ───────────────────────────────────


class PlantSpecUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:plant_spec_defs",
            kwargs={"project_pk": self.kwargs["project_pk"]},
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base_plant_estimator.html"
        ctx["upload_title"] = "Upload Plant Specifications"
        ctx["upload_description"] = (
            "Upload plant specification definitions with production factors."
        )
        ctx["download_url_name"] = "estimator:download_plant_spec_template"
        ctx["columns"] = [
            ("Section", "Section group"),
            ("Trade Name", "Trade name"),
            ("Plant Specification", "Specification name"),
            ("Unit", "Unit of measure"),
            ("Plant Type", "Name of plant (matches Plant Costs)"),
            ("Daily Production", "Daily production output"),
            ("Operator", "Operator factor"),
            ("Site", "Site factor"),
        ]
        return ctx

    def form_valid(self, form):
        from .importers import PlantSpecImporter

        return _handle_upload(
            self.request,
            PlantSpecImporter,
            self.get_success_url(),
            "Plant Specifications",
            project=self.get_project(),
        )


class DownloadPlantSpecTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(
            [
                "Section",
                "Trade Name",
                "Plant Specification",
                "Unit",
                "Plant Type",
                "Daily Production",
                "Operator",
                "Site",
            ],
            "PlantSpec_Template.xlsx",
        )


# ── Preliminary Costs Upload/Download ─────────────────────────────


class PreliminaryCostUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:preliminary_costs",
            kwargs={"project_pk": self.kwargs["project_pk"]},
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base_baseline_costs.html"
        ctx["upload_title"] = "Upload Preliminary Costs"
        ctx["upload_description"] = (
            "Upload preliminary costs with type, amounts and time-based rates."
        )
        ctx["download_url_name"] = "estimator:download_preliminary_cost_template"
        ctx["columns"] = [
            ("Preliminary Type", "Type code (e.g. fixed_contractual, time_facilities)"),
            ("Name", "Name of the preliminary item"),
            ("Sum", "Sum/lump sum value"),
            ("Amount", "Amount (fixed items)"),
            ("Number/Month", "Quantity per month (time items)"),
            ("Monthly Rate", "Monthly rate (time items)"),
            ("Months", "Duration in months (time items)"),
        ]
        return ctx

    def form_valid(self, form):
        from .importers import PreliminaryCostImporter

        return _handle_upload(
            self.request,
            PreliminaryCostImporter,
            self.get_success_url(),
            "Preliminary Costs",
            project=self.get_project(),
        )


class DownloadPreliminaryCostTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(
            [
                "Preliminary Type",
                "Name",
                "Sum",
                "Amount",
                "Number/Month",
                "Monthly Rate",
                "Months",
            ],
            "PreliminaryCost_Template.xlsx",
        )


# ── Preliminary Specs Upload/Download ─────────────────────────────


class PreliminarySpecUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:preliminary_spec_defs",
            kwargs={"project_pk": self.kwargs["project_pk"]},
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base_prelim_estimator.html"
        ctx["upload_title"] = "Upload Preliminary Specifications"
        ctx["upload_description"] = "Upload preliminary specification definitions."
        ctx["download_url_name"] = "estimator:download_preliminary_spec_template"
        ctx["columns"] = [
            ("Section", "Section group"),
            ("Trade Name", "Trade name"),
            ("Name", "Specification name"),
            ("Unit", "Unit of measure"),
            ("Amount", "Amount (R)"),
        ]
        return ctx

    def form_valid(self, form):
        from .importers import PreliminarySpecImporter

        return _handle_upload(
            self.request,
            PreliminarySpecImporter,
            self.get_success_url(),
            "Preliminary Specifications",
            project=self.get_project(),
        )


class DownloadPreliminarySpecTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(
            ["Section", "Trade Name", "Name", "Unit", "Preliminary Type"],
            "PreliminarySpec_Template.xlsx",
        )


# ══════════════════════════════════════════════════════════════════════
# System Library Views
# ══════════════════════════════════════════════════════════════════════


class SystemLibraryMixin(UserPassesTestMixin, ContextMixin):
    """Mixin for system library views: requires staff."""

    def test_func(self):
        request = getattr(self, "request", None)
        return request and request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_system_view"] = True
        return context


# ── System Trade Codes ────────────────────────────────────────────────


class SystemTradeCodeListView(SystemLibraryMixin, ListView):
    model = SystemTradeCode
    template_name = "estimator/system/trade_code_list.html"
    context_object_name = "trade_codes"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemTradeCodeForm())
        context["f_q"] = self.request.GET.get("q", "")
        return context

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request, SystemTradeCode.objects.all(), label="trade codes"
        ):
            return redirect(reverse("estimator:sys_trade_codes"))
        if _handle_bulk_action(request, SystemTradeCode.objects.all()):
            return redirect(reverse("estimator:sys_trade_codes"))
        form = SystemTradeCodeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(reverse("estimator:sys_trade_codes"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


class SystemTradeCodeUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Trade Codes"
        context["upload_description"] = (
            "Upload trade code prefixes and names from an Excel template."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = "estimator:sys_download_trade_code_template"
        return context

    def form_valid(self, form):
        from .importers import TradeCodeImporter

        return _handle_upload(
            self.request, TradeCodeImporter, "estimator:sys_trade_codes", "Trade Codes"
        )


class DownloadSystemTradeCodeTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            ["Prefix", "Trade Name"],
            "system_trade_codes_template.xlsx",
        )


# ── System Municipalities ─────────────────────────────────────────────


class SystemMunicipalityListView(SystemLibraryMixin, ListView):
    model = Municipality
    template_name = "estimator/system/municipality_list.html"
    context_object_name = "municipalities"
    paginate_by = 50

    def get_queryset(self):
        qs = Municipality.objects.all()
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(
                models.Q(province__name__icontains=q)
                | models.Q(municipality_name__icontains=q)
                | models.Q(code__icontains=q)
                | models.Q(district__icontains=q)
            )
        province = self.request.GET.get("province", "").strip()
        if province:
            qs = qs.filter(province_id=province)
        return qs.order_by("province__name", "municipality_name")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemMunicipalityForm())
        context["provinces"] = Province.objects.all().order_by("name")
        context["f_q"] = self.request.GET.get("q", "")
        context["f_province"] = self.request.GET.get("province", "")
        context["query_params"] = _pagination_query_params(self.request)
        return context

    def post(self, request, *args, **kwargs):
        if "load_defaults" in request.POST:
            province_codes = {
                "Western Cape": "WC",
                "Northern Cape": "NC",
                "Eastern Cape": "EC",
                "Free State": "FS",
                "Gauteng": "GP",
                "Kwa-Zulu Natal": "KZN",
                "Limpopo": "LP",
                "Mpumalanga": "MP",
                "North-West": "NW",
            }
            defaults = [
                # Western Cape
                {
                    "province": "Western Cape",
                    "municipality": "Breede Valley Local Municipality",
                    "code": "WC025",
                    "district": "Cape Winelands",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Drakenstein Local Municipality",
                    "code": "WC023",
                    "district": "Cape Winelands",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Langeberg Local Municipality",
                    "code": "WC026",
                    "district": "Cape Winelands",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Stellenbosch Local Municipality",
                    "code": "WC024",
                    "district": "Cape Winelands",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Witzenberg Local Municipality",
                    "code": "WC022",
                    "district": "Cape Winelands",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Beaufort West Local Municipality",
                    "code": "WC053",
                    "district": "Central Karoo",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Laingsburg Local Municipality",
                    "code": "WC051",
                    "district": "Central Karoo",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Prince Albert Local Municipality",
                    "code": "WC052",
                    "district": "Central Karoo",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Bitou Local Municipality",
                    "code": "WC047",
                    "district": "Garden Route",
                },
                {
                    "province": "Western Cape",
                    "municipality": "George Local Municipality",
                    "code": "WC044",
                    "district": "Garden Route",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Hessequa Local Municipality",
                    "code": "WC042",
                    "district": "Garden Route",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Kannaland Local Municipality",
                    "code": "WC041",
                    "district": "Garden Route",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Knysna Local Municipality",
                    "code": "WC048",
                    "district": "Garden Route",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Mossel Bay Local Municipality",
                    "code": "WC043",
                    "district": "Garden Route",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Oudtshoorn Local Municipality",
                    "code": "WC045",
                    "district": "Garden Route",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Cape Agulhas Local Municipality",
                    "code": "WC033",
                    "district": "Overberg",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Overstrand Local Municipality",
                    "code": "WC032",
                    "district": "Overberg",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Swellendam Local Municipality",
                    "code": "WC034",
                    "district": "Overberg",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Theewaterskloof Local Municipality",
                    "code": "WC031",
                    "district": "Overberg",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Bergrivier Local Municipality",
                    "code": "WC013",
                    "district": "West Coast",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Cederberg Local Municipality",
                    "code": "WC012",
                    "district": "West Coast",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Matzikama Local Municipality",
                    "code": "WC011",
                    "district": "West Coast",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Saldanha Bay Local Municipality",
                    "code": "WC014",
                    "district": "West Coast",
                },
                {
                    "province": "Western Cape",
                    "municipality": "Swartland Local Municipality",
                    "code": "WC015",
                    "district": "West Coast",
                },
                {
                    "province": "Western Cape",
                    "municipality": "City of Cape Town Metropolitan Municipality",
                    "code": "CPT",
                    "district": "Metropolitan",
                },
                # Northern Cape
                {
                    "province": "Northern Cape",
                    "municipality": "Dikgatlong Local Municipality",
                    "code": "NC092",
                    "district": "Frances Baard",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Magareng Local Municipality",
                    "code": "NC093",
                    "district": "Frances Baard",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Phokwane Local Municipality",
                    "code": "NC094",
                    "district": "Frances Baard",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Sol Plaatje Local Municipality",
                    "code": "NC091",
                    "district": "Frances Baard",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Ga-Segonyana Local Municipality",
                    "code": "NC452",
                    "district": "John Taolo Gaetsewe",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Gamagara Local Municipality",
                    "code": "NC453",
                    "district": "John Taolo Gaetsewe",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Joe Morolong Local Municipality",
                    "code": "NC451",
                    "district": "John Taolo Gaetsewe",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Hantam Local Municipality",
                    "code": "NC065",
                    "district": "Namakwa",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Kamiesberg Local Municipality",
                    "code": "NC064",
                    "district": "Namakwa",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Karoo Hoogland Local Municipality",
                    "code": "NC066",
                    "district": "Namakwa",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Khâi-Ma Local Municipality",
                    "code": "NC067",
                    "district": "Namakwa",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Nama Khoi Local Municipality",
                    "code": "NC062",
                    "district": "Namakwa",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Richtersveld Local Municipality",
                    "code": "NC061",
                    "district": "Namakwa",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Emthanjeni Local Municipality",
                    "code": "NC073",
                    "district": "Pixley ka Seme",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Kareeberg Local Municipality",
                    "code": "NC074",
                    "district": "Pixley ka Seme",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Renosterberg Local Municipality",
                    "code": "NC075",
                    "district": "Pixley ka Seme",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Siyancuma Local Municipality",
                    "code": "NC078",
                    "district": "Pixley ka Seme",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Siyathemba Local Municipality",
                    "code": "NC077",
                    "district": "Pixley ka Seme",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Thembelihle Local Municipality",
                    "code": "NC076",
                    "district": "Pixley ka Seme",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Ubuntu Local Municipality",
                    "code": "NC071",
                    "district": "Pixley ka Seme",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Umsobomvu Local Municipality",
                    "code": "NC072",
                    "district": "Pixley ka Seme",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "!Kheis Local Municipality",
                    "code": "NC084",
                    "district": "ZF Mgcawu",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Dawid Kruiper Local Municipality",
                    "code": "NC087",
                    "district": "ZF Mgcawu",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Kai !Garib Local Municipality",
                    "code": "NC082",
                    "district": "ZF Mgcawu",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Kgatelopele Local Municipality",
                    "code": "NC086",
                    "district": "ZF Mgcawu",
                },
                {
                    "province": "Northern Cape",
                    "municipality": "Tsantsabane Local Municipality",
                    "code": "NC085",
                    "district": "ZF Mgcawu",
                },
                # Eastern Cape
                {
                    "province": "Eastern Cape",
                    "municipality": "Matatiele Local Municipality",
                    "code": "EC441",
                    "district": "Alfred Nzo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Ntabankulu Local Municipality",
                    "code": "EC444",
                    "district": "Alfred Nzo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Umzimvubu Local Municipality",
                    "code": "EC442",
                    "district": "Alfred Nzo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Winnie Madikizela-Mandela Local Municipality",
                    "code": "EC443",
                    "district": "Alfred Nzo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Amahlathi Local Municipality",
                    "code": "EC124",
                    "district": "Amathole",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Great Kei Local Municipality",
                    "code": "EC123",
                    "district": "Amathole",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Mbhashe Local Municipality",
                    "code": "EC121",
                    "district": "Amathole",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Mnquma Local Municipality",
                    "code": "EC122",
                    "district": "Amathole",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Ngqushwa Local Municipality",
                    "code": "EC126",
                    "district": "Amathole",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Raymond Mhlaba Local Municipality",
                    "code": "EC129",
                    "district": "Amathole",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Dr AB Xuma Local Municipality",
                    "code": "EC137",
                    "district": "Chris Hani",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Emalahleni Local Municipality",
                    "code": "EC136",
                    "district": "Chris Hani",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Enoch Mgijima Local Municipality",
                    "code": "EC139",
                    "district": "Chris Hani",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Intsika Yethu Local Municipality",
                    "code": "EC135",
                    "district": "Chris Hani",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Inxuba Yethemba Local Municipality",
                    "code": "EC131",
                    "district": "Chris Hani",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Sakhisizwe Local Municipality",
                    "code": "EC138",
                    "district": "Chris Hani",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Elundini Local Municipality",
                    "code": "EC141",
                    "district": "Joe Gqabi",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Senqu Local Municipality",
                    "code": "EC142",
                    "district": "Joe Gqabi",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Walter Sisulu Local Municipality",
                    "code": "EC145",
                    "district": "Joe Gqabi",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Ingquza Hill Local Municipality",
                    "code": "EC153",
                    "district": "OR Tambo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "King Sabata Dalindyebo Local Municipality",
                    "code": "EC157",
                    "district": "OR Tambo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Kumkani Mhlontlo Local Municipality",
                    "code": "EC156",
                    "district": "OR Tambo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Nyandeni Local Municipality",
                    "code": "EC155",
                    "district": "OR Tambo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Port St Johns Local Municipality",
                    "code": "EC154",
                    "district": "OR Tambo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Blue Crane Route Local Municipality",
                    "code": "EC102",
                    "district": "Sarah Baartman",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Dr Beyers Naudé Local Municipality",
                    "code": "EC101",
                    "district": "Sarah Baartman",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Kou-Kamma Local Municipality",
                    "code": "EC109",
                    "district": "Sarah Baartman",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Kouga Local Municipality",
                    "code": "EC108",
                    "district": "Sarah Baartman",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Makana Local Municipality",
                    "code": "EC104",
                    "district": "Sarah Baartman",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Ndlambe Local Municipality",
                    "code": "EC105",
                    "district": "Sarah Baartman",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Sundays River Valley Local Municipality",
                    "code": "EC106",
                    "district": "Sarah Baartman",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Alfred Nzo District Municipality",
                    "code": "DC44",
                    "district": "Alfred Nzo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Amathole District Municipality",
                    "code": "DC12",
                    "district": "Amathole",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Buffalo City Metropolitan Municipality",
                    "code": "BUF",
                    "district": "Metropolitan",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Chris Hani District Municipality",
                    "code": "DC13",
                    "district": "Chris Hani",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Joe Gqabi District Municipality",
                    "code": "DC14",
                    "district": "Joe Gqabi",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Nelson Mandela Bay Metropolitan Municipality",
                    "code": "NMA",
                    "district": "Metropolitan",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "OR Tambo District Municipality",
                    "code": "DC15",
                    "district": "OR Tambo",
                },
                {
                    "province": "Eastern Cape",
                    "municipality": "Sarah Baartman District Municipality",
                    "code": "DC10",
                    "district": "Sarah Baartman",
                },
                # Free State
                {
                    "province": "Free State",
                    "municipality": "Fezile Dabi District Municipality",
                    "code": "DC20",
                    "district": "Fezile Dabi",
                },
                {
                    "province": "Free State",
                    "municipality": "Lejweleputswa District Municipality",
                    "code": "DC18",
                    "district": "Lejweleputswa",
                },
                {
                    "province": "Free State",
                    "municipality": "Mangaung Metropolitan Municipality",
                    "code": "MAN",
                    "district": "Metropolitan",
                },
                {
                    "province": "Free State",
                    "municipality": "Thabo Mofutsanyana District Municipality",
                    "code": "DC19",
                    "district": "Thabo Mofutsanyana",
                },
                {
                    "province": "Free State",
                    "municipality": "Xhariep District Municipality",
                    "code": "DC16",
                    "district": "Xhariep",
                },
                {
                    "province": "Free State",
                    "municipality": "Mafube Local Municipality",
                    "code": "FS205",
                    "district": "Fezile Dabi",
                },
                {
                    "province": "Free State",
                    "municipality": "Metsimaholo Local Municipality",
                    "code": "FS204",
                    "district": "Fezile Dabi",
                },
                {
                    "province": "Free State",
                    "municipality": "Moqhaka Local Municipality",
                    "code": "FS201",
                    "district": "Fezile Dabi",
                },
                {
                    "province": "Free State",
                    "municipality": "Ngwathe Local Municipality",
                    "code": "FS203",
                    "district": "Fezile Dabi",
                },
                {
                    "province": "Free State",
                    "municipality": "Masilonyana Local Municipality",
                    "code": "FS181",
                    "district": "Lejweleputswa",
                },
                {
                    "province": "Free State",
                    "municipality": "Matjhabeng Local Municipality",
                    "code": "FS184",
                    "district": "Lejweleputswa",
                },
                {
                    "province": "Free State",
                    "municipality": "Nala Local Municipality",
                    "code": "FS185",
                    "district": "Lejweleputswa",
                },
                {
                    "province": "Free State",
                    "municipality": "Tokologo Local Municipality",
                    "code": "FS182",
                    "district": "Lejweleputswa",
                },
                {
                    "province": "Free State",
                    "municipality": "Tswelopele Local Municipality",
                    "code": "FS183",
                    "district": "Lejweleputswa",
                },
                {
                    "province": "Free State",
                    "municipality": "Dihlabeng Local Municipality",
                    "code": "FS192",
                    "district": "Thabo Mofutsanyana",
                },
                {
                    "province": "Free State",
                    "municipality": "Maluti-a-Phofung Local Municipality",
                    "code": "FS194",
                    "district": "Thabo Mofutsanyana",
                },
                {
                    "province": "Free State",
                    "municipality": "Mantsopa Local Municipality",
                    "code": "FS196",
                    "district": "Thabo Mofutsanyana",
                },
                {
                    "province": "Free State",
                    "municipality": "Nketoana Local Municipality",
                    "code": "FS193",
                    "district": "Thabo Mofutsanyana",
                },
                {
                    "province": "Free State",
                    "municipality": "Phumelela Local Municipality",
                    "code": "FS195",
                    "district": "Thabo Mofutsanyana",
                },
                {
                    "province": "Free State",
                    "municipality": "Setsoto Local Municipality",
                    "code": "FS191",
                    "district": "Thabo Mofutsanyana",
                },
                {
                    "province": "Free State",
                    "municipality": "Kopanong Local Municipality",
                    "code": "FS162",
                    "district": "Xhariep",
                },
                {
                    "province": "Free State",
                    "municipality": "Letsemeng Local Municipality",
                    "code": "FS161",
                    "district": "Xhariep",
                },
                {
                    "province": "Free State",
                    "municipality": "Mohokare Local Municipality",
                    "code": "FS163",
                    "district": "Xhariep",
                },
                # Gauteng
                {
                    "province": "Gauteng",
                    "municipality": "Emfuleni Local Municipality",
                    "code": "GT421",
                    "district": "Sedibeng",
                },
                {
                    "province": "Gauteng",
                    "municipality": "Midvaal Local Municipality",
                    "code": "GT422",
                    "district": "Sedibeng",
                },
                {
                    "province": "Gauteng",
                    "municipality": "Lesedi Local Municipality",
                    "code": "GT423",
                    "district": "Sedibeng",
                },
                {
                    "province": "Gauteng",
                    "municipality": "Mogale City Local Municipality",
                    "code": "GT481",
                    "district": "West Rand",
                },
                {
                    "province": "Gauteng",
                    "municipality": "Merafong City Local Municipality",
                    "code": "GT484",
                    "district": "West Rand",
                },
                {
                    "province": "Gauteng",
                    "municipality": "Rand West City Local Municipality",
                    "code": "GT485",
                    "district": "West Rand",
                },
                {
                    "province": "Gauteng",
                    "municipality": "City of Ekurhuleni Metropolitan Municipality",
                    "code": "EKU",
                    "district": "Metropolitan",
                },
                {
                    "province": "Gauteng",
                    "municipality": "City of Johannesburg Metropolitan Municipality",
                    "code": "JHB",
                    "district": "Metropolitan",
                },
                {
                    "province": "Gauteng",
                    "municipality": "City of Tshwane Metropolitan Municipality",
                    "code": "TSH",
                    "district": "Metropolitan",
                },
                {
                    "province": "Gauteng",
                    "municipality": "Sedibeng District Municipality",
                    "code": "DC42",
                    "district": "Sedibeng",
                },
                {
                    "province": "Gauteng",
                    "municipality": "West Rand District Municipality",
                    "code": "DC48",
                    "district": "West Rand",
                },
                # Kwa-Zulu Natal
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Amajuba District Municipality",
                    "code": "DC25",
                    "district": "Amajuba",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "eThekwini Metropolitan Municipality",
                    "code": "ETH",
                    "district": "Metropolitan",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Harry Gwala District Municipality",
                    "code": "DC43",
                    "district": "Harry Gwala",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "iLembe District Municipality",
                    "code": "DC29",
                    "district": "iLembe",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "King Cetshwayo District Municipality",
                    "code": "DC28",
                    "district": "King Cetshwayo",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Ugu District Municipality",
                    "code": "DC21",
                    "district": "Ugu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMgungundlovu District Municipality",
                    "code": "DC22",
                    "district": "uMgungundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMkhanyakude District Municipality",
                    "code": "DC27",
                    "district": "Umkhanyakude",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMzinyathi District Municipality",
                    "code": "DC24",
                    "district": "Umzinyathi",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uThukela District Municipality",
                    "code": "DC23",
                    "district": "Uthukela",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Zululand District Municipality",
                    "code": "DC26",
                    "district": "Zululand",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Dannhauser Local Municipality",
                    "code": "KZN254",
                    "district": "Amajuba",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "eMadlangeni Local Municipality",
                    "code": "KZN253",
                    "district": "Amajuba",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Newcastle Local Municipality",
                    "code": "KZN252",
                    "district": "Amajuba",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Dr Nkosazana Dlamini Zuma Local Municipality",
                    "code": "KZN436",
                    "district": "Harry Gwala",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Greater Kokstad Local Municipality",
                    "code": "KZN433",
                    "district": "Harry Gwala",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Johannes Phumani Phungula Local Municipality",
                    "code": "KZN434",
                    "district": "Harry Gwala",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Umzimkhulu Local Municipality",
                    "code": "KZN435",
                    "district": "Harry Gwala",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "KwaDukuza Local Municipality",
                    "code": "KZN292",
                    "district": "iLembe",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Mandeni Local Municipality",
                    "code": "KZN291",
                    "district": "iLembe",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Maphumulo Local Municipality",
                    "code": "KZN294",
                    "district": "iLembe",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Ndwedwe Local Municipality",
                    "code": "KZN293",
                    "district": "iLembe",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Mthonjaneni Local Municipality",
                    "code": "KZN285",
                    "district": "King Cetshwayo",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Nkandla Local Municipality",
                    "code": "KZN286",
                    "district": "King Cetshwayo",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMfolozi Local Municipality",
                    "code": "KZN281",
                    "district": "King Cetshwayo",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMhlathuze Local Municipality",
                    "code": "KZN282",
                    "district": "King Cetshwayo",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMlalazi Local Municipality",
                    "code": "KZN284",
                    "district": "King Cetshwayo",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Ray Nkonyeni Local Municipality",
                    "code": "KZN216",
                    "district": "Ugu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMdoni Local Municipality",
                    "code": "KZN212",
                    "district": "Ugu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMuziwabantu Local Municipality",
                    "code": "KZN214",
                    "district": "Ugu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Umzumbe Local Municipality",
                    "code": "KZN213",
                    "district": "Ugu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Impendle Local Municipality",
                    "code": "KZN224",
                    "district": "uMgungundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Mkhambathini Local Municipality",
                    "code": "KZN226",
                    "district": "uMgungundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Mpofana Local Municipality",
                    "code": "KZN223",
                    "district": "uMgungundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Msunduzi Local Municipality",
                    "code": "KZN225",
                    "district": "uMgungundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Richmond Local Municipality",
                    "code": "KZN227",
                    "district": "uMgungundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMngeni Local Municipality",
                    "code": "KZN222",
                    "district": "uMgngundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Richmond Local Municipality",
                    "code": "KZN227",
                    "district": "uMgungundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMngeni Local Municipality",
                    "code": "KZN222",
                    "district": "uMgungundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMshwathi Local Municipality",
                    "code": "KZN221",
                    "district": "uMgungundlovu",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Big Five Hlabisa Local Municipality",
                    "code": "KZN276",
                    "district": "Umkhanyakude",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Jozini Local Municipality",
                    "code": "KZN272",
                    "district": "Umkhanyakude",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Mtubatuba Local Municipality",
                    "code": "KZN275",
                    "district": "Umkhanyakude",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uMhlabuyalingana Local Municipality",
                    "code": "KZN271",
                    "district": "Umkhanyakude",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Endumeni Local Municipality",
                    "code": "KZN241",
                    "district": "Umzinyathi",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Msinga Local Municipality",
                    "code": "KZN244",
                    "district": "Umzinyathi",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Nqutu Local Municipality",
                    "code": "KZN242",
                    "district": "Umzinyathi",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Umvoti Local Municipality",
                    "code": "KZN245",
                    "district": "Umzinyathi",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Alfred Duma Local Municipality",
                    "code": "KZN238",
                    "district": "Uthukela",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Inkosi Langalibalele Local Municipality",
                    "code": "KZN237",
                    "district": "Uthukela",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Okhahlamba Local Municipality",
                    "code": "KZN235",
                    "district": "Uthukela",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Abaqulusi Local Municipality",
                    "code": "KZN263",
                    "district": "Zululand",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "eDumbe Local Municipality",
                    "code": "KZN261",
                    "district": "Zululand",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Nongoma Local Municipality",
                    "code": "KZN265",
                    "district": "Zululand",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "Ulundi Local Municipality",
                    "code": "KZN266",
                    "district": "Zululand",
                },
                {
                    "province": "Kwa-Zulu Natal",
                    "municipality": "uPhongolo Local Municipality",
                    "code": "KZN262",
                    "district": "Zululand",
                },
                # Limpopo
                {
                    "province": "Limpopo",
                    "municipality": "Capricorn District Municipality",
                    "code": "DC35",
                    "district": "Capricorn",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Mopani District Municipality",
                    "code": "DC33",
                    "district": "Mopani",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Sekhukhune District Municipality",
                    "code": "DC47",
                    "district": "Sekhukhune",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Vhembe District Municipality",
                    "code": "DC34",
                    "district": "Vhembe",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Waterberg District Municipality",
                    "code": "DC36",
                    "district": "Waterberg",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Blouberg Local Municipality",
                    "code": "LIM351",
                    "district": "Capricorn",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Lepelle-Nkumpi Local Municipality",
                    "code": "LIM355",
                    "district": "Capricorn",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Molemole Local Municipality",
                    "code": "LIM353",
                    "district": "Capricorn",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Polokwane Local Municipality",
                    "code": "LIM354",
                    "district": "Capricorn",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Ba-Phalaborwa Local Municipality",
                    "code": "LIM334",
                    "district": "Mopani",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Greater Giyani Local Municipality",
                    "code": "LIM331",
                    "district": "Mopani",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Greater Letaba Local Municipality",
                    "code": "LIM332",
                    "district": "Mopani",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Greater Tzaneen Local Municipality",
                    "code": "LIM333",
                    "district": "Mopani",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Maruleng Local Municipality",
                    "code": "LIM335",
                    "district": "Mopani",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Elias Motsoaledi Local Municipality",
                    "code": "LIM472",
                    "district": "Sekhukhune",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Ephraim Mogale Local Municipality",
                    "code": "LIM471",
                    "district": "Sekhukhune",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Fetakgomo Tubatse Local Municipality",
                    "code": "LIM476",
                    "district": "Sekhukhune",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Makhuduthamaga Local Municipality",
                    "code": "LIM473",
                    "district": "Sekhukhune",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Collins Chabane Local Municipality",
                    "code": "LIM345",
                    "district": "Vhembe",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Makhado Local Municipality",
                    "code": "LIM344",
                    "district": "Vhembe",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Musina Local Municipality",
                    "code": "LIM341",
                    "district": "Vhembe",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Thulamela Local Municipality",
                    "code": "LIM343",
                    "district": "Vhembe",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Bela-Bela Local Municipality",
                    "code": "LIM366",
                    "district": "Waterberg",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Lephalale Local Municipality",
                    "code": "LIM362",
                    "district": "Waterberg",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Modimolle–Mookgophong Local Municipality",
                    "code": "LIM368",
                    "district": "Waterberg",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Mogalakwena Local Municipality",
                    "code": "LIM367",
                    "district": "Waterberg",
                },
                {
                    "province": "Limpopo",
                    "municipality": "Thabazimbi Local Municipality",
                    "code": "LIM361",
                    "district": "Waterberg",
                },
                # Mpumalanga
                {
                    "province": "Mpumalanga",
                    "municipality": "Ehlanzeni District Municipality",
                    "code": "DC32",
                    "district": "Ehlanzeni",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Gert Sibande District Municipality",
                    "code": "DC30",
                    "district": "Gert Sibande",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Nkangala District Municipality",
                    "code": "DC31",
                    "district": "Nkangala",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Bushbuckridge Local Municipality",
                    "code": "MP325",
                    "district": "Ehlanzeni",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Mbombela Local Municipality",
                    "code": "MP326",
                    "district": "Ehlanzeni",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Nkomazi Local Municipality",
                    "code": "MP324",
                    "district": "Ehlanzeni",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Thaba Chweu Local Municipality",
                    "code": "MP321",
                    "district": "Ehlanzeni",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Albert Luthuli Local Municipality",
                    "code": "MP301",
                    "district": "Gert Sibande",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Dipaleseng Local Municipality",
                    "code": "MP306",
                    "district": "Gert Sibande",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Govan Mbeki Local Municipality",
                    "code": "MP307",
                    "district": "Gert Sibande",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Lekwa Local Municipality",
                    "code": "MP305",
                    "district": "Gert Sibande",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Mkhondo Local Municipality",
                    "code": "MP303",
                    "district": "Gert Sibande",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Msukaligwa Local Municipality",
                    "code": "MP302",
                    "district": "Gert Sibande",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Pixley ka Seme Local Municipality",
                    "code": "MP304",
                    "district": "Gert Sibande",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Dr JS Moroka Local Municipality",
                    "code": "MP316",
                    "district": "Nkangala",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Emakhazeni Local Municipality",
                    "code": "MP314",
                    "district": "Nkangala",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Emalahleni Local Municipality",
                    "code": "MP312",
                    "district": "Nkangala",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Steve Tshwete Local Municipality",
                    "code": "MP313",
                    "district": "Nkangala",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Thembisile Hani Local Municipality",
                    "code": "MP315",
                    "district": "Nkangala",
                },
                {
                    "province": "Mpumalanga",
                    "municipality": "Victor Khanye Local Municipality",
                    "code": "MP311",
                    "district": "Nkangala",
                },
                # North-West
                {
                    "province": "North-West",
                    "municipality": "Bojanala Platinum District Municipality",
                    "code": "DC37",
                    "district": "Bojanala Platinum",
                },
                {
                    "province": "North-West",
                    "municipality": "Dr Kenneth Kaunda District Municipality",
                    "code": "DC40",
                    "district": "Dr Kenneth Kaunda",
                },
                {
                    "province": "North-West",
                    "municipality": "Dr Ruth Segomotsi Mompati District Municipality",
                    "code": "DC39",
                    "district": "Dr Ruth Segomotsi Mompati",
                },
                {
                    "province": "North-West",
                    "municipality": "Ngaka Modiri Molema District Municipality",
                    "code": "DC38",
                    "district": "Ngaka Modiri Molema",
                },
                {
                    "province": "North-West",
                    "municipality": "Kgetlengrivier Local Municipality",
                    "code": "NW374",
                    "district": "Bojanala Platinum",
                },
                {
                    "province": "North-West",
                    "municipality": "Madibeng Local Municipality",
                    "code": "NW372",
                    "district": "Bojanala Platinum",
                },
                {
                    "province": "North-West",
                    "municipality": "Moretele Local Municipality",
                    "code": "NW371",
                    "district": "Bojanala Platinum",
                },
                {
                    "province": "North-West",
                    "municipality": "Moses Kotane Local Municipality",
                    "code": "NW375",
                    "district": "Bojanala Platinum",
                },
                {
                    "province": "North-West",
                    "municipality": "Rustenburg Local Municipality",
                    "code": "NW373",
                    "district": "Bojanala Platinum",
                },
                {
                    "province": "North-West",
                    "municipality": "City of Matlosana Local Municipality",
                    "code": "NW403",
                    "district": "Dr Kenneth Kaunda",
                },
                {
                    "province": "North-West",
                    "municipality": "JB Marks Local Municipality",
                    "code": "NW405",
                    "district": "Dr Kenneth Kaunda",
                },
                {
                    "province": "North-West",
                    "municipality": "Maquassi Hills Local Municipality",
                    "code": "NW404",
                    "district": "Dr Kenneth Kaunda",
                },
                {
                    "province": "North-West",
                    "municipality": "Greater Taung Local Municipality",
                    "code": "NW394",
                    "district": "Dr Ruth Segomotsi Mompati",
                },
                {
                    "province": "North-West",
                    "municipality": "Kagisano-Molopo Local Municipality",
                    "code": "NW397",
                    "district": "Dr Ruth Segomotsi Mompati",
                },
                {
                    "province": "North-West",
                    "municipality": "Lekwa-Teemane Local Municipality",
                    "code": "NW396",
                    "district": "Dr Ruth Segomotsi Mompati",
                },
                {
                    "province": "North-West",
                    "municipality": "Mamusa Local Municipality",
                    "code": "NW393",
                    "district": "Dr Ruth Segomotsi Mompati",
                },
                {
                    "province": "North-West",
                    "municipality": "Naledi Local Municipality",
                    "code": "NW392",
                    "district": "Dr Ruth Segomotsi Mompati",
                },
                {
                    "province": "North-West",
                    "municipality": "Ditsobotla Local Municipality",
                    "code": "NW384",
                    "district": "Ngaka Modiri Molema",
                },
                {
                    "province": "North-West",
                    "municipality": "Mahikeng Local Municipality",
                    "code": "NW383",
                    "district": "Ngaka Modiri Molema",
                },
                {
                    "province": "North-West",
                    "municipality": "Ramotshere Moiloa Local Municipality",
                    "code": "NW385",
                    "district": "Ngaka Modiri Molema",
                },
                {
                    "province": "North-West",
                    "municipality": "Ratlou Local Municipality",
                    "code": "NW381",
                    "district": "Ngaka Modiri Molema",
                },
                {
                    "province": "North-West",
                    "municipality": "Tswaing Local Municipality",
                    "code": "NW382",
                    "district": "Ngaka Modiri Molema",
                },
            ]
            created_count = 0
            for item in defaults:
                prov_name = item["province"]
                prov_code = province_codes.get(prov_name, "")
                prov, _ = Province.objects.get_or_create(
                    name=prov_name, defaults={"code": prov_code}
                )
                _, created = Municipality.objects.get_or_create(
                    province=prov,
                    municipality_name=item["municipality"],
                    code=item["code"],
                    defaults={"district": item["district"]},
                )
                if created:
                    created_count += 1
            messages.success(request, f"Loaded {created_count} default municipalities.")
            return redirect(reverse("estimator:sys_municipalities"))

        if _handle_clear_action(
            request, Municipality.objects.all(), label="municipalities"
        ):
            return redirect(reverse("estimator:sys_municipalities"))
        if _handle_bulk_action(request, Municipality.objects.all()):
            return redirect(reverse("estimator:sys_municipalities"))
        form = SystemMunicipalityForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Municipality added successfully.")
            return redirect(reverse("estimator:sys_municipalities"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


class SystemMunicipalityUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Municipalities"
        context["upload_description"] = (
            "Upload South African provinces and municipalities from an Excel template."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = "estimator:sys_download_municipality_template"
        return context

    def form_valid(self, form):
        from .importers import MunicipalityImporter

        return _handle_upload(
            self.request,
            MunicipalityImporter,
            "estimator:sys_municipalities",
            "Municipalities",
        )


class DownloadSystemMunicipalityTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            ["Province", "Municipality Name", "Code", "District"],
            "system_municipalities_template.xlsx",
        )


# ── System Provinces ──────────────────────────────────────────────────


class SystemProvinceListView(SystemLibraryMixin, ListView):
    model = Province
    template_name = "estimator/system/province_list.html"
    context_object_name = "provinces"
    paginate_by = 50

    def get_queryset(self):
        qs = Province.objects.all()
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(models.Q(name__icontains=q) | models.Q(code__icontains=q))
        return qs.order_by("name")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemProvinceForm())
        context["f_q"] = self.request.GET.get("q", "")
        context["query_params"] = _pagination_query_params(self.request)
        return context

    def post(self, request, *args, **kwargs):
        if "load_defaults" in request.POST:
            defaults = [
                {"name": "Western Cape", "code": "WC"},
                {"name": "Northern Cape", "code": "NC"},
                {"name": "Eastern Cape", "code": "EC"},
                {"name": "Free State", "code": "FS"},
                {"name": "Gauteng", "code": "GP"},
                {"name": "Kwa-Zulu Natal", "code": "KZN"},
                {"name": "Limpopo", "code": "LP"},
                {"name": "Mpumalanga", "code": "MP"},
                {"name": "North-West", "code": "NW"},
            ]
            created_count = 0
            for item in defaults:
                _, created = Province.objects.get_or_create(
                    name=item["name"], defaults={"code": item["code"]}
                )
                if created:
                    created_count += 1
            messages.success(request, f"Loaded {created_count} default provinces.")
            return redirect(reverse("estimator:sys_provinces"))

        if _handle_clear_action(request, Province.objects.all(), label="provinces"):
            return redirect(reverse("estimator:sys_provinces"))
        if _handle_bulk_action(request, Province.objects.all()):
            return redirect(reverse("estimator:sys_provinces"))
        form = SystemProvinceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Province added successfully.")
            return redirect(reverse("estimator:sys_provinces"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


class SystemProvinceUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Provinces"
        context["upload_description"] = (
            "Upload South African provinces from an Excel template."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = "estimator:sys_download_province_template"
        return context

    def form_valid(self, form):
        from .importers import ProvinceImporter

        return _handle_upload(
            self.request, ProvinceImporter, "estimator:sys_provinces", "Provinces"
        )


class DownloadSystemProvinceTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            ["Province Name", "Code"],
            "system_provinces_template.xlsx",
        )


# ── System Materials ──────────────────────────────────────────────────


class SystemMaterialListView(SystemLibraryMixin, ListView):
    model = SystemMaterial
    template_name = "estimator/system/material_list.html"
    context_object_name = "materials"

    def get_queryset(self):
        qs = SystemMaterial.objects.all()
        unit = self.request.GET.get("unit", "").strip()
        if unit:
            qs = qs.filter(unit=unit)
        trade_name = self.request.GET.get("trade_name", "").strip()
        if trade_name:
            qs = qs.filter(trade_name=trade_name)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemMaterialForm())
        all_materials = SystemMaterial.objects.all()
        context["units"] = (
            all_materials.exclude(unit="")
            .values_list("unit", flat=True)
            .distinct()
            .order_by("unit")
        )
        context["trade_names"] = (
            all_materials.exclude(trade_name="")
            .values_list("trade_name", flat=True)
            .distinct()
            .order_by("trade_name")
        )
        context["f_q"] = self.request.GET.get("q", "")
        context["f_unit"] = self.request.GET.get("unit", "")
        context["f_trade_name"] = self.request.GET.get("trade_name", "")
        return context

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request, SystemMaterial.objects.all(), label="materials"
        ):
            return redirect(reverse("estimator:sys_materials"))
        if _handle_bulk_action(request, SystemMaterial.objects.all()):
            return redirect(reverse("estimator:sys_materials"))
        form = SystemMaterialForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(reverse("estimator:sys_materials"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemMaterialView(View):
    ALLOWED_FIELDS = {
        "pack_qty",
        "pack_cost",
        "trade_name",
        "material_code",
        "unit",
        "material_variety",
        "market_spec",
    }

    def post(self, request, pk):
        if not request.user.is_staff:
            return JsonResponse({"error": "Forbidden"}, status=403)
        mat = get_object_or_404(SystemMaterial, pk=pk)
        data = json.loads(request.body)
        field, value = data.get("field"), data.get("value")
        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": "Invalid field"}, status=400)
        try:
            if field in ("pack_qty", "pack_cost"):
                setattr(mat, field, Decimal(str(value or 0)))
            else:
                setattr(mat, field, str(value or ""))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)
        mat.save()
        return JsonResponse(
            {
                "ok": True,
                "pack_qty": format_num(mat.pack_qty),
                "pack_cost": format_num(mat.pack_cost),
                "market_rate": format_num(mat.market_rate),
            }
        )


class SystemMaterialUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Materials"
        context["upload_description"] = (
            "Upload material pricing data from an Excel template."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = "estimator:sys_download_material_template"
        return context

    def form_valid(self, form):
        from .importers import MaterialCostImporter

        return _handle_upload(
            self.request, MaterialCostImporter, "estimator:sys_materials", "Materials"
        )


class DownloadSystemMaterialTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Trade Name",
                "Material Code",
                "Unit",
                "Pack Qty",
                "Pack Cost",
                "Material Variety",
                "Market Spec",
            ],
            "system_materials_template.xlsx",
        )


# ── System Material Specs ─────────────────────────────────────────────


class SysMaterialSpecListView(SystemLibraryMixin, ListView):
    model = SystemSpecification
    template_name = "estimator/system/material_spec_list.html"
    context_object_name = "specs"
    paginate_by = 50

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request, SystemSpecification.objects.all(), label="material specs"
        ):
            return redirect(reverse("estimator:sys_material_specs"))
        if _handle_bulk_action(
            request, SystemSpecification.objects.all(), allow_toggle_active=True
        ):
            return redirect(reverse("estimator:sys_material_specs"))
        return self._handle_create(request, *args, **kwargs)

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .select_related("trade_code")
            .prefetch_related("spec_components__material")
        )
        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)
        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)
        name = self.request.GET.get("name")
        if name:
            qs = qs.filter(name=name)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .forms import SystemSpecificationComponentFormSet, SystemSpecificationForm

        context["spec_form"] = context.get("spec_form", SystemSpecificationForm())
        context["component_formset"] = context.get(
            "component_formset", SystemSpecificationComponentFormSet()
        )

        all_specs = SystemSpecification.objects.all()
        context["sections"] = (
            all_specs.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_codes"] = SystemTradeCode.objects.all()
        context["names"] = (
            all_specs.values_list("name", flat=True).distinct().order_by("name")
        )
        context["materials"] = SystemMaterial.objects.all().order_by("material_code")
        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_code"] = self.request.GET.get("trade_code", "")
        context["f_name"] = self.request.GET.get("name", "")
        context["f_q"] = self.request.GET.get("q", "")
        context["query_params"] = _pagination_query_params(self.request)
        context.update(
            _spec_datalist_context(
                SystemSpecification.objects.all(), unit_field="unit_label"
            )
        )
        return context

    def _handle_create(self, request, *args, **kwargs):
        from .forms import SystemSpecificationComponentFormSet, SystemSpecificationForm

        spec_form = SystemSpecificationForm(request.POST)
        component_formset = SystemSpecificationComponentFormSet(request.POST)
        if spec_form.is_valid():
            spec = spec_form.save()
            component_formset = SystemSpecificationComponentFormSet(
                request.POST, instance=spec
            )
            if component_formset.is_valid():
                component_formset.save()
                return redirect(reverse("estimator:sys_material_specs"))
            else:
                spec.delete()
        self.object_list = self.get_queryset()
        return self.render_to_response(
            self.get_context_data(
                spec_form=spec_form, component_formset=component_formset
            )
        )


class SystemMaterialSpecUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Material Specs"
        context["upload_description"] = (
            "Upload material specification definitions with component breakdowns."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = "estimator:sys_download_material_spec_template"
        context["columns"] = [
            (
                "Spec Name",
                "Name of the material specification (rows with same name are grouped)",
            ),
            ("Section", "Section grouping"),
            ("Trade Code Prefix", "Trade code prefix for lookup"),
            ("Unit", "Unit of measure (e.g. m3, m2)"),
            ("Material Code", "Material code to link component"),
            ("Label", "Display label for component"),
            ("Qty per Unit", "Quantity of material per spec unit"),
        ]
        return context

    def form_valid(self, form):
        from .importers import MaterialSpecImporter

        return _handle_upload(
            self.request,
            MaterialSpecImporter,
            "estimator:sys_material_specs",
            "Material Specs",
        )


class DownloadSystemMaterialSpecTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Spec Name",
                "Section",
                "Trade Code Prefix",
                "Unit",
                "Material Code",
                "Label",
                "Qty per Unit",
            ],
            "system_material_specs_template.xlsx",
        )


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSysSpecComponentView(View):
    ALLOWED_FIELDS = {"qty_per_unit", "material", "label"}

    def post(self, request, pk):
        if not request.user.is_staff:
            return JsonResponse({"error": "Forbidden"}, status=403)
        comp = get_object_or_404(SystemSpecificationComponent, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")
        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        try:
            if field == "qty_per_unit":
                comp.qty_per_unit = Decimal(str(value or 0))
            elif field == "label":
                comp.label = (value or "").strip()
            else:  # material
                if value in (None, "", 0, "0"):
                    comp.material = None
                else:
                    material = SystemMaterial.objects.filter(pk=int(value)).first()
                    if material is None:
                        return JsonResponse({"error": "Material not found"}, status=404)
                    comp.material = material
                    comp.label = material.material_code
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        comp.save()
        spec = comp.specification
        return JsonResponse(
            {
                "ok": True,
                "qty_per_unit": format_num(comp.qty_per_unit),
                "label": comp.label,
                "material_id": comp.material_id,
                "spec_id": spec.id,
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class AddSysSpecComponentView(View):
    """AJAX endpoint to add a new component to a SystemSpecification."""

    def post(self, request, pk):
        if not request.user.is_staff:
            return JsonResponse({"error": "Forbidden"}, status=403)
        spec = get_object_or_404(SystemSpecification, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}

        material_id = data.get("material")
        label = (data.get("label") or "").strip()
        qty_raw = data.get("qty_per_unit", 0)

        material = None
        if material_id not in (None, "", 0, "0"):
            try:
                material = SystemMaterial.objects.filter(pk=int(material_id)).first()
            except (TypeError, ValueError):
                material = None

        if not label and material is not None:
            label = material.material_code

        try:
            qty = Decimal(str(qty_raw or 0))
        except Exception:
            qty = Decimal("0")

        next_order = (
            spec.spec_components.aggregate(models.Max("sort_order"))["sort_order__max"]
            or 0
        ) + 1
        comp = SystemSpecificationComponent.objects.create(
            specification=spec,
            material=material,
            label=label,
            qty_per_unit=qty,
            sort_order=next_order,
        )
        return JsonResponse(
            {
                "ok": True,
                "component": {
                    "id": comp.id,
                    "material_id": comp.material_id,
                    "label": comp.label,
                    "qty_per_unit": format_num(comp.qty_per_unit),
                },
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteSysSpecComponentView(View):
    """AJAX endpoint to delete a SystemSpecificationComponent."""

    def post(self, request, pk):
        if not request.user.is_staff:
            return JsonResponse({"error": "Forbidden"}, status=403)
        comp = get_object_or_404(SystemSpecificationComponent, pk=pk)
        spec = comp.specification
        comp.delete()
        return JsonResponse(
            {
                "ok": True,
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteSystemSpecificationView(View):
    """AJAX endpoint to delete a SystemSpecification (material spec)."""

    def post(self, request, pk):
        if not request.user.is_staff:
            return JsonResponse({"error": "Forbidden"}, status=403)
        spec = get_object_or_404(SystemSpecification, pk=pk)
        spec.delete()
        return JsonResponse({"ok": True})


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemSpecificationView(View):
    """AJAX endpoint to update fields on a SystemSpecification (material spec)."""

    ALLOWED_FIELDS = {
        "section": "str",
        "name": "str",
        "unit_label": "str",
        "is_active": "bool",
    }

    def post(self, request, pk):
        if not request.user.is_staff:
            return JsonResponse({"error": "Forbidden"}, status=403)
        item = get_object_or_404(SystemSpecification, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True})


# ── System Labour Crews ───────────────────────────────────────────────


class SystemLabourCrewListView(SystemLibraryMixin, ListView):
    model = SystemLabourCrew
    template_name = "estimator/system/labour_crew_list.html"
    context_object_name = "crews"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemLabourCrewForm())
        context["f_q"] = self.request.GET.get("q", "")
        return context

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request, SystemLabourCrew.objects.all(), label="labour crews"
        ):
            return redirect(reverse("estimator:sys_labour_crews"))
        if _handle_bulk_action(request, SystemLabourCrew.objects.all()):
            return redirect(reverse("estimator:sys_labour_crews"))
        form = SystemLabourCrewForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(reverse("estimator:sys_labour_crews"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemLabourCrewView(View):
    ALLOWED_FIELDS = {
        "crew_type",
        "skilled",
        "semi_skilled",
        "general",
        "daily_production",
        "skilled_rate",
        "semi_skilled_rate",
        "general_rate",
    }

    def post(self, request, pk):
        if not request.user.is_staff:
            return JsonResponse({"error": "Forbidden"}, status=403)
        crew = get_object_or_404(SystemLabourCrew, pk=pk)
        data = json.loads(request.body)
        field, value = data.get("field"), data.get("value")
        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": "Invalid field"}, status=400)
        setattr(crew, field, value)
        crew.save()
        return JsonResponse(
            {
                "ok": True,
                "crew_size": crew.crew_size,
                "crew_daily_cost": str(crew.crew_daily_cost),
            }
        )


class SystemLabourCrewUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Labour Crews"
        context["upload_description"] = (
            "Upload labour crew compositions and daily rates."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = "estimator:sys_download_labour_crew_template"
        return context

    def form_valid(self, form):
        from .importers import LabourCostImporter

        return _handle_upload(
            self.request,
            LabourCostImporter,
            "estimator:sys_labour_crews",
            "Labour Crews",
        )


class DownloadSystemLabourCrewTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Crew Type",
                "Crew Size",
                "Skilled",
                "Semi Skilled",
                "General",
                "Daily Production",
                "Skilled Rate",
                "Semi Skilled Rate",
                "General Rate",
            ],
            "system_labour_crews_template.xlsx",
        )


# ── System Labour Specs ───────────────────────────────────────────────


class SystemLabourSpecListView(SystemLibraryMixin, ListView):
    model = SystemLabourSpecification
    template_name = "estimator/system/labour_spec_list.html"
    context_object_name = "labour_specs"
    paginate_by = 50

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request,
            SystemLabourSpecification.objects.all(),
            label="labour specs",
        ):
            return redirect(reverse("estimator:sys_labour_specs"))
        if _handle_bulk_action(
            request,
            SystemLabourSpecification.objects.all(),
            allow_toggle_active=True,
        ):
            return redirect(reverse("estimator:sys_labour_specs"))
        return self._handle_create(request, *args, **kwargs)

    def get_queryset(self):
        qs = super().get_queryset().select_related("crew")
        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)
        trade_name = self.request.GET.get("trade_name")
        if trade_name:
            qs = qs.filter(trade_name=trade_name)
        name = self.request.GET.get("name")
        if name:
            qs = qs.filter(name=name)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemLabourSpecificationForm())
        all_specs = SystemLabourSpecification.objects.all()
        context["sections"] = (
            all_specs.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_names"] = (
            all_specs.exclude(trade_name="")
            .values_list("trade_name", flat=True)
            .distinct()
            .order_by("trade_name")
        )
        context["names"] = (
            all_specs.values_list("name", flat=True).distinct().order_by("name")
        )
        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_name"] = self.request.GET.get("trade_name", "")
        context["f_name"] = self.request.GET.get("name", "")
        context["f_q"] = self.request.GET.get("q", "")
        context["crews"] = SystemLabourCrew.objects.all()
        context["query_params"] = _pagination_query_params(self.request)
        context.update(_spec_datalist_context(all_specs))
        return context

    def _handle_create(self, request, *args, **kwargs):
        form = SystemLabourSpecificationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(reverse("estimator:sys_labour_specs"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemLabourSpecView(View):
    ALLOWED_FIELDS = {
        "section": "str",
        "trade_name": "str",
        "name": "str",
        "unit": "str",
        "crew": "fk",
        "daily_production": "decimal",
        "team_mix": "decimal",
        "site_factor": "decimal",
        "tools_factor": "decimal",
        "leadership_factor": "decimal",
        "is_active": "bool",
    }

    def post(self, request, pk):
        if not request.user.is_staff:
            return JsonResponse({"error": "Forbidden"}, status=403)
        ls = get_object_or_404(SystemLabourSpecification, pk=pk)
        data = json.loads(request.body)
        field, value = data.get("field"), data.get("value")
        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": "Invalid field"}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(ls, field, Decimal(str(value)))
            elif field_type == "fk":
                if value == "" or value is None:
                    ls.crew = None
                else:
                    ls.crew = get_object_or_404(SystemLabourCrew, pk=int(value))
            elif field_type == "bool":
                setattr(ls, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(ls, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        ls.save()
        return JsonResponse(
            {
                "ok": True,
                "daily_output": str(ls.daily_output),
                "daily_cost": str(ls.daily_cost),
                "rate_per_unit": str(ls.rate_per_unit),
                "crew_type": ls.crew.crew_type if ls.crew else None,
            }
        )


class SystemLabourSpecUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Labour Specifications"
        context["upload_description"] = (
            "Upload labour specification definitions with crew assignments and factors."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = "estimator:sys_download_labour_spec_template"
        return context

    def form_valid(self, form):
        from .importers import LabourSpecImporter

        return _handle_upload(
            self.request,
            LabourSpecImporter,
            "estimator:sys_labour_specs",
            "Labour Specifications",
        )


class DownloadSystemLabourSpecTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Section",
                "Trade Name",
                "Name",
                "Unit",
                "Crew Type",
                "Daily Production",
                "Team Mix",
                "Site Factor",
                "Tools Factor",
                "Leadership Factor",
            ],
            "system_labour_specs_template.xlsx",
        )


# ── System Plant Costs ─────────────────────────────────────────────


class SystemPlantCostListView(SystemLibraryMixin, ListView):
    model = SystemPlantCost
    template_name = "estimator/system/plant_cost_list.html"
    context_object_name = "plants"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemPlantCostForm())
        context["f_q"] = self.request.GET.get("q", "")
        return context

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request, SystemPlantCost.objects.all(), label="plant costs"
        ):
            return redirect("estimator:sys_plant_costs")
        if _handle_bulk_action(request, SystemPlantCost.objects.all()):
            return redirect("estimator:sys_plant_costs")
        form = SystemPlantCostForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("estimator:sys_plant_costs")
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemPlantCostView(SystemLibraryMixin, View):
    ALLOWED_FIELDS = {
        "name": "str",
        "hourly_production": "decimal",
        "hourly_rate": "decimal",
    }

    def post(self, request, pk):
        item = get_object_or_404(SystemPlantCost, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True})


class SystemPlantCostUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Plant Costs"
        context["upload_description"] = (
            "Upload plant & equipment costs with hourly production and rates."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = "estimator:sys_download_plant_cost_template"
        return context

    def form_valid(self, form):
        from .importers import PlantCostImporter

        return _handle_upload(
            self.request,
            PlantCostImporter,
            "estimator:sys_plant_costs",
            "Plant Costs",
        )


class DownloadSystemPlantCostTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            ["Plant & Equipment", "Hourly Production", "Hourly Rate"],
            "system_plant_costs_template.xlsx",
        )


# ── System Plant Specs ─────────────────────────────────────────────


class SystemPlantSpecListView(SystemLibraryMixin, ListView):
    model = SystemPlantSpecification
    template_name = "estimator/system/plant_spec_list.html"
    context_object_name = "plant_specs"
    paginate_by = 50

    def get_queryset(self):
        return super().get_queryset().prefetch_related("components__plant_type")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemPlantSpecificationForm())
        context["plants"] = SystemPlantCost.objects.all()
        context["f_q"] = self.request.GET.get("q", "")
        context["query_params"] = _pagination_query_params(self.request)
        context.update(_spec_datalist_context(SystemPlantSpecification.objects.all()))
        return context

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request,
            SystemPlantSpecification.objects.all(),
            label="plant specs",
        ):
            return redirect("estimator:sys_plant_specs")
        if _handle_bulk_action(
            request,
            SystemPlantSpecification.objects.all(),
            allow_toggle_active=True,
        ):
            return redirect("estimator:sys_plant_specs")
        form = SystemPlantSpecificationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("estimator:sys_plant_specs")
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemPlantSpecView(SystemLibraryMixin, View):
    ALLOWED_FIELDS = {
        "section": "str",
        "trade_name": "str",
        "name": "str",
        "unit": "str",
        "daily_production": "decimal",
        "operator_factor": "decimal",
        "site_factor": "decimal",
        "is_active": "bool",
    }

    def post(self, request, pk):
        item = get_object_or_404(SystemPlantSpecification, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            elif field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse(
            {
                "ok": True,
                "daily_output": str(item.daily_output),
                "daily_cost": str(item.daily_cost),
                "rate_per_unit": str(item.rate_per_unit),
            }
        )


class SystemPlantSpecUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Plant Specifications"
        context["upload_description"] = (
            "Upload plant specification definitions with production factors."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = "estimator:sys_download_plant_spec_template"
        return context

    def form_valid(self, form):
        from .importers import PlantSpecImporter

        return _handle_upload(
            self.request,
            PlantSpecImporter,
            "estimator:sys_plant_specs",
            "Plant Specifications",
        )


class DownloadSystemPlantSpecTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Section",
                "Trade Name",
                "Plant Specification",
                "Unit",
                "Plant Type",
                "Daily Production",
                "Operator",
                "Site",
            ],
            "system_plant_specs_template.xlsx",
        )


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemPlantSpecComponentView(SystemLibraryMixin, View):
    """AJAX endpoint to update plant_type or hours on a SystemPlantSpecificationComponent."""

    ALLOWED_FIELDS = {"plant_type", "hours"}

    def post(self, request, pk):
        comp = get_object_or_404(SystemPlantSpecificationComponent, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        try:
            if field == "hours":
                comp.hours = Decimal(str(value or 0))
            else:
                if value in (None, "", 0, "0"):
                    comp.plant_type = None
                else:
                    comp.plant_type = SystemPlantCost.objects.filter(
                        pk=int(value)
                    ).first()
                    if comp.plant_type is None:
                        return JsonResponse(
                            {"error": "Plant type not found"}, status=404
                        )
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        comp.save()
        spec = comp.specification
        return JsonResponse(
            {
                "ok": True,
                "daily_cost": str(spec.daily_cost),
                "rate_per_unit": str(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class AddSystemPlantSpecComponentView(SystemLibraryMixin, View):
    """AJAX endpoint to add a new component to a SystemPlantSpecification."""

    def post(self, request, pk):
        spec = get_object_or_404(SystemPlantSpecification, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}

        plant_type_id = data.get("plant_type")
        hours_raw = data.get("hours", 0)

        plant_type = None
        if plant_type_id not in (None, "", 0, "0"):
            try:
                plant_type = SystemPlantCost.objects.filter(
                    pk=int(plant_type_id)
                ).first()
            except (TypeError, ValueError):
                plant_type = None

        try:
            hours = Decimal(str(hours_raw or 0))
        except Exception:
            hours = Decimal("0")

        next_order = (
            spec.components.aggregate(models.Max("sort_order"))["sort_order__max"] or 0
        ) + 1
        comp = SystemPlantSpecificationComponent.objects.create(
            specification=spec,
            plant_type=plant_type,
            hours=hours,
            sort_order=next_order,
        )
        return JsonResponse(
            {
                "ok": True,
                "component": {
                    "id": comp.id,
                    "plant_type_id": comp.plant_type_id,
                    "plant_type_name": comp.plant_type.name if comp.plant_type else "",
                    "hours": str(comp.hours),
                },
                "daily_cost": str(spec.daily_cost),
                "rate_per_unit": str(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteSystemPlantSpecComponentView(SystemLibraryMixin, View):
    """AJAX endpoint to delete a SystemPlantSpecificationComponent."""

    def post(self, request, pk):
        comp = get_object_or_404(SystemPlantSpecificationComponent, pk=pk)
        spec = comp.specification
        comp.delete()
        return JsonResponse(
            {
                "ok": True,
                "daily_cost": str(spec.daily_cost),
                "rate_per_unit": str(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteSystemPlantSpecificationView(SystemLibraryMixin, View):
    """AJAX endpoint to delete a SystemPlantSpecification."""

    def post(self, request, pk):
        spec = get_object_or_404(SystemPlantSpecification, pk=pk)
        spec.delete()
        return JsonResponse({"ok": True})


# ── System Preliminary Costs ──────────────────────────────────────


class SystemPreliminaryCostListView(SystemLibraryMixin, ListView):
    model = SystemPreliminaryCost
    template_name = "estimator/system/preliminary_cost_list.html"
    context_object_name = "preliminaries"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemPreliminaryCostForm())
        context["f_q"] = self.request.GET.get("q", "")
        return context

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request,
            SystemPreliminaryCost.objects.all(),
            label="preliminary costs",
        ):
            return redirect("estimator:sys_preliminary_costs")
        if _handle_bulk_action(request, SystemPreliminaryCost.objects.all()):
            return redirect("estimator:sys_preliminary_costs")
        form = SystemPreliminaryCostForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("estimator:sys_preliminary_costs")
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemPreliminaryCostView(SystemLibraryMixin, View):
    ALLOWED_FIELDS = {
        "name": "str",
        "preliminary_type": "str",
        "sum_value": "decimal",
        "amount": "decimal",
        "number_per_month": "decimal",
        "monthly_rate": "decimal",
        "months": "decimal",
    }

    def post(self, request, pk):
        item = get_object_or_404(SystemPreliminaryCost, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True, "computed_amount": str(item.computed_amount)})


class SystemPreliminaryCostUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Preliminary Costs"
        context["upload_description"] = (
            "Upload preliminary costs with type, amounts and time-based rates."
        )
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = (
            "estimator:sys_download_preliminary_cost_template"
        )
        return context

    def form_valid(self, form):
        from .importers import PreliminaryCostImporter

        return _handle_upload(
            self.request,
            PreliminaryCostImporter,
            "estimator:sys_preliminary_costs",
            "Preliminary Costs",
        )


class DownloadSystemPreliminaryCostTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Preliminary Type",
                "Name",
                "Sum",
                "Amount",
                "Number/Month",
                "Monthly Rate",
                "Months",
            ],
            "system_preliminary_costs_template.xlsx",
        )


# ── System Preliminary Specs ──────────────────────────────────────


class SystemPreliminarySpecListView(SystemLibraryMixin, ListView):
    model = SystemPreliminarySpecification
    template_name = "estimator/system/preliminary_spec_list.html"
    context_object_name = "preliminary_specs"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", SystemPreliminarySpecificationForm())
        context["preliminary_type_choices"] = (
            SystemPreliminaryCost.PRELIMINARY_TYPE_CHOICES
        )
        context["f_q"] = self.request.GET.get("q", "")
        context.update(
            _spec_datalist_context(SystemPreliminarySpecification.objects.all())
        )
        return context

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request,
            SystemPreliminarySpecification.objects.all(),
            label="preliminary specs",
        ):
            return redirect("estimator:sys_preliminary_specs")
        if _handle_bulk_action(
            request,
            SystemPreliminarySpecification.objects.all(),
            allow_toggle_active=True,
        ):
            return redirect("estimator:sys_preliminary_specs")
        form = SystemPreliminarySpecificationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("estimator:sys_preliminary_specs")
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemPreliminarySpecView(SystemLibraryMixin, View):
    ALLOWED_FIELDS = {
        "section": "str",
        "trade_name": "str",
        "name": "str",
        "unit": "str",
        "preliminary_type": "str",
        "is_active": "bool",
    }

    def post(self, request, pk):
        item = get_object_or_404(SystemPreliminarySpecification, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            elif field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True, "amount": str(item.amount)})


class SystemPreliminarySpecUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Preliminary Specifications"
        context["upload_description"] = "Upload preliminary specification definitions."
        context["parent_template"] = "estimator/system/base_system.html"
        context["download_url_name"] = (
            "estimator:sys_download_preliminary_spec_template"
        )
        return context

    def form_valid(self, form):
        from .importers import PreliminarySpecImporter

        return _handle_upload(
            self.request,
            PreliminarySpecImporter,
            "estimator:sys_preliminary_specs",
            "Preliminary Specifications",
        )


class DownloadSystemPreliminarySpecTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            ["Section", "Trade Name", "Name", "Unit", "Preliminary Type"],
            "system_preliminary_specs_template.xlsx",
        )


# ═══════════════════════════════════════════════════════════════════
# Contractor Library Views (per-Company; mirror of System Library)
# ═══════════════════════════════════════════════════════════════════


class ContractorLibraryMixin(LoginRequiredMixin, ContextMixin):
    """Resolve the user's contractor Company and scope all queries to it.

    A user can belong to multiple companies; we pick the first CONTRACTOR-type
    company. Superusers fall back to the first available CONTRACTOR company so
    the library is browsable in admin contexts.
    """

    def get_company(self):
        if not hasattr(self, "_company"):
            from app.Project.models import Company

            request = getattr(self, "request", None)
            user = getattr(request, "user", None)
            company = None
            if user is not None and user.is_authenticated:
                company = user.companies.filter(type=Company.Type.CONTRACTOR).first()
                if company is None and user.is_superuser:
                    company = Company.objects.filter(
                        type=Company.Type.CONTRACTOR
                    ).first()
            self._company = company
        return self._company

    def dispatch(self, request, *args, **kwargs):
        from django.http import HttpResponseForbidden

        if not request.user.is_authenticated:
            return super().dispatch(request, *args, **kwargs)
        if self.get_company() is None:
            from app.Account.subscription_config import Subscription

            if request.user.subscription in [
                Subscription.DEMO_TIER,
                Subscription.FREE_TIER,
            ]:
                return super().dispatch(request, *args, **kwargs)

            return HttpResponseForbidden(
                "You must be linked to a Contractor company to use the Contractor "
                "Library."
            )
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        qs = super().get_queryset()  # ty:ignore[unresolved-attribute]
        company = self.get_company()
        if company is None:
            return qs.none()
        return qs.filter(company=company)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_contractor_view"] = True
        context["contractor_company"] = self.get_company()
        return context


def _contractor_company_for(request):
    """Resolve the contractor Company for AJAX requests, or None."""
    from app.Project.models import Company

    user = request.user
    if not user.is_authenticated:
        return None
    company = user.companies.filter(type=Company.Type.CONTRACTOR).first()
    if company is None and user.is_superuser:
        company = Company.objects.filter(type=Company.Type.CONTRACTOR).first()
    return company


# ── Contractor Trade Codes ────────────────────────────────────────────


class ContractorTradeCodeListView(ContractorLibraryMixin, ListView):
    model = ContractorTradeCode
    template_name = "estimator/contractor/trade_code_list.html"
    context_object_name = "trade_codes"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", ContractorTradeCodeForm())
        context["f_q"] = self.request.GET.get("q", "")
        return context

    def post(self, request, *args, **kwargs):
        qs = ContractorTradeCode.objects.filter(company=self.get_company())
        if _handle_clear_action(request, qs, label="trade codes"):
            return redirect(reverse("estimator:ctr_trade_codes"))
        if _handle_bulk_action(request, qs):
            return redirect(reverse("estimator:ctr_trade_codes"))
        if request.POST.get("action") == "sync_system":
            from .services import sync_trade_codes_to_contractor

            result = sync_trade_codes_to_contractor(self.get_company())
            messages.success(
                request,
                f"Trade codes synced with system library — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect(reverse("estimator:ctr_trade_codes"))

        form = ContractorTradeCodeForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = self.get_company()
            obj.save()
            return redirect(reverse("estimator:ctr_trade_codes"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


class ContractorTradeCodeUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Trade Codes"
        context["upload_description"] = (
            "Upload trade code prefixes and names from an Excel template."
        )
        context["parent_template"] = "estimator/contractor/base_contractor.html"
        context["download_url_name"] = "estimator:ctr_download_trade_code_template"
        return context

    def form_valid(self, form):
        from .importers import TradeCodeImporter

        return _handle_upload(
            self.request,
            TradeCodeImporter,
            "estimator:ctr_trade_codes",
            "Trade Codes",
            company=self.get_company(),
        )


class DownloadContractorTradeCodeTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            ["Prefix", "Trade Name"],
            "contractor_trade_codes_template.xlsx",
        )


# ── Contractor Materials ──────────────────────────────────────────────


class ContractorMaterialListView(ContractorLibraryMixin, ListView):
    model = ContractorMaterial
    template_name = "estimator/contractor/material_list.html"
    context_object_name = "materials"

    def get_queryset(self):
        qs = super().get_queryset()
        unit = self.request.GET.get("unit", "").strip()
        if unit:
            qs = qs.filter(unit=unit)
        trade_name = self.request.GET.get("trade_name", "").strip()
        if trade_name:
            qs = qs.filter(trade_name=trade_name)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.get_company()
        context["form"] = context.get("form", ContractorMaterialForm(company=company))
        all_materials = ContractorMaterial.objects.filter(company=company)
        context["units"] = (
            all_materials.exclude(unit="")
            .values_list("unit", flat=True)
            .distinct()
            .order_by("unit")
        )
        context["trade_names"] = (
            all_materials.exclude(trade_name="")
            .values_list("trade_name", flat=True)
            .distinct()
            .order_by("trade_name")
        )
        context["f_q"] = self.request.GET.get("q", "")
        context["f_unit"] = self.request.GET.get("unit", "")
        context["f_trade_name"] = self.request.GET.get("trade_name", "")
        return context

    def post(self, request, *args, **kwargs):
        qs = ContractorMaterial.objects.filter(company=self.get_company())
        if _handle_clear_action(request, qs, label="materials"):
            return redirect(reverse("estimator:ctr_materials"))
        if _handle_bulk_action(request, qs):
            return redirect(reverse("estimator:ctr_materials"))
        if request.POST.get("action") == "sync_system":
            from .services import sync_materials_to_contractor

            result = sync_materials_to_contractor(self.get_company())
            messages.success(
                request,
                f"Material costs synced with system library — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect(reverse("estimator:ctr_materials"))

        form = ContractorMaterialForm(request.POST, company=self.get_company())
        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = self.get_company()
            obj.save()
            return redirect(reverse("estimator:ctr_materials"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorMaterialView(View):
    ALLOWED_FIELDS = {
        "pack_qty",
        "pack_cost",
        "trade_name",
        "material_code",
        "unit",
        "material_variety",
        "market_spec",
    }

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        mat = get_object_or_404(ContractorMaterial, pk=pk, company=company)
        data = json.loads(request.body)
        field, value = data.get("field"), data.get("value")
        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": "Invalid field"}, status=400)
        try:
            if field in ("pack_qty", "pack_cost"):
                setattr(mat, field, Decimal(str(value or 0)))
            else:
                setattr(mat, field, str(value or ""))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)
        mat.save()
        return JsonResponse(
            {
                "ok": True,
                "pack_qty": format_num(mat.pack_qty),
                "pack_cost": format_num(mat.pack_cost),
                "market_rate": format_num(mat.market_rate),
            }
        )


class ContractorMaterialUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Materials"
        context["upload_description"] = (
            "Upload material pricing data from an Excel template."
        )
        context["parent_template"] = "estimator/contractor/base_contractor.html"
        context["download_url_name"] = "estimator:ctr_download_material_template"
        return context

    def form_valid(self, form):
        from .importers import MaterialCostImporter

        return _handle_upload(
            self.request,
            MaterialCostImporter,
            "estimator:ctr_materials",
            "Materials",
            company=self.get_company(),
        )


class DownloadContractorMaterialTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Trade Name",
                "Material Code",
                "Unit",
                "Pack Qty",
                "Pack Cost",
                "Material Variety",
                "Market Spec",
            ],
            "contractor_materials_template.xlsx",
        )


# ── Contractor Material Specs ─────────────────────────────────────────


class ContractorMaterialSpecListView(ContractorLibraryMixin, ListView):
    model = ContractorSpecification
    template_name = "estimator/contractor/material_spec_list.html"
    context_object_name = "specs"
    paginate_by = 50

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .select_related("trade_code")
            .prefetch_related("spec_components__material")
        )
        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)
        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)
        name = self.request.GET.get("name")
        if name:
            qs = qs.filter(name=name)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.get_company()
        context["spec_form"] = context.get(
            "spec_form", ContractorSpecificationForm(company=company)
        )
        context["component_formset"] = context.get(
            "component_formset", ContractorSpecificationComponentFormSet()
        )

        all_specs = ContractorSpecification.objects.filter(company=company)
        context["sections"] = (
            all_specs.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_codes"] = ContractorTradeCode.objects.filter(company=company)
        context["names"] = (
            all_specs.values_list("name", flat=True).distinct().order_by("name")
        )
        context["materials"] = ContractorMaterial.objects.filter(
            company=company
        ).order_by("material_code")
        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_code"] = self.request.GET.get("trade_code", "")
        context["f_name"] = self.request.GET.get("name", "")
        context["f_q"] = self.request.GET.get("q", "")
        context["query_params"] = _pagination_query_params(self.request)
        context.update(_spec_datalist_context(all_specs, unit_field="unit_label"))
        return context

    def post(self, request, *args, **kwargs):
        qs = ContractorSpecification.objects.filter(company=self.get_company())
        if _handle_clear_action(request, qs, label="material specs"):
            return redirect(reverse("estimator:ctr_material_specs"))
        if _handle_bulk_action(request, qs, allow_toggle_active=True):
            return redirect(reverse("estimator:ctr_material_specs"))
        if request.POST.get("action") == "sync_system":
            from .services import sync_material_specs_to_contractor

            result = sync_material_specs_to_contractor(self.get_company())
            messages.success(
                request,
                f"Material specs synced with system library — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect(reverse("estimator:ctr_material_specs"))

        company = self.get_company()
        spec_form = ContractorSpecificationForm(request.POST, company=company)
        component_formset = ContractorSpecificationComponentFormSet(request.POST)
        if spec_form.is_valid():
            spec = spec_form.save(commit=False)
            spec.company = company
            spec.save()
            component_formset = ContractorSpecificationComponentFormSet(
                request.POST, instance=spec
            )
            if component_formset.is_valid():
                component_formset.save()
                return redirect(reverse("estimator:ctr_material_specs"))
            else:
                spec.delete()
        self.object_list = self.get_queryset()
        return self.render_to_response(
            self.get_context_data(
                spec_form=spec_form, component_formset=component_formset
            )
        )


class ContractorMaterialSpecUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Material Specs"
        context["upload_description"] = (
            "Upload material specification definitions with component breakdowns."
        )
        context["parent_template"] = "estimator/contractor/base_contractor.html"
        context["download_url_name"] = "estimator:ctr_download_material_spec_template"
        context["columns"] = [
            (
                "Spec Name",
                "Name of the material specification (rows with same name are grouped)",
            ),
            ("Section", "Section grouping"),
            ("Trade Code Prefix", "Trade code prefix for lookup"),
            ("Unit", "Unit of measure (e.g. m3, m2)"),
            ("Material Code", "Material code to link component"),
            ("Label", "Display label for component"),
            ("Qty per Unit", "Quantity of material per spec unit"),
        ]
        return context

    def form_valid(self, form):
        from .importers import MaterialSpecImporter

        return _handle_upload(
            self.request,
            MaterialSpecImporter,
            "estimator:ctr_material_specs",
            "Material Specs",
            company=self.get_company(),
        )


class DownloadContractorMaterialSpecTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Spec Name",
                "Section",
                "Trade Code Prefix",
                "Unit",
                "Material Code",
                "Label",
                "Qty per Unit",
            ],
            "contractor_material_specs_template.xlsx",
        )


@method_decorator(csrf_exempt, name="dispatch")
class UpdateCtrSpecComponentView(View):
    ALLOWED_FIELDS = {"qty_per_unit", "material", "label"}

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        comp = get_object_or_404(
            ContractorSpecificationComponent, pk=pk, specification__company=company
        )
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")
        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        try:
            if field == "qty_per_unit":
                comp.qty_per_unit = Decimal(str(value or 0))
            elif field == "label":
                comp.label = (value or "").strip()
            else:  # material
                if value in (None, "", 0, "0"):
                    comp.material = None
                else:
                    material = ContractorMaterial.objects.filter(
                        pk=int(value), company=company
                    ).first()
                    if material is None:
                        return JsonResponse({"error": "Material not found"}, status=404)
                    comp.material = material
                    comp.label = material.material_code
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        comp.save()
        spec = comp.specification
        return JsonResponse(
            {
                "ok": True,
                "qty_per_unit": format_num(comp.qty_per_unit),
                "label": comp.label,
                "material_id": comp.material_id,
                "spec_id": spec.id,
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class AddCtrSpecComponentView(View):
    """AJAX endpoint to add a new component to a ContractorSpecification."""

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        spec = get_object_or_404(ContractorSpecification, pk=pk, company=company)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}

        material_id = data.get("material")
        label = (data.get("label") or "").strip()
        qty_raw = data.get("qty_per_unit", 0)

        material = None
        if material_id not in (None, "", 0, "0"):
            try:
                material = ContractorMaterial.objects.filter(
                    pk=int(material_id), company=company
                ).first()
            except (TypeError, ValueError):
                material = None

        if not label and material is not None:
            label = material.material_code

        try:
            qty = Decimal(str(qty_raw or 0))
        except Exception:
            qty = Decimal("0")

        next_order = (
            spec.spec_components.aggregate(models.Max("sort_order"))["sort_order__max"]
            or 0
        ) + 1
        comp = ContractorSpecificationComponent.objects.create(
            specification=spec,
            material=material,
            label=label,
            qty_per_unit=qty,
            sort_order=next_order,
        )
        return JsonResponse(
            {
                "ok": True,
                "component": {
                    "id": comp.id,
                    "material_id": comp.material_id,
                    "label": comp.label,
                    "qty_per_unit": format_num(comp.qty_per_unit),
                },
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteCtrSpecComponentView(View):
    """AJAX endpoint to delete a ContractorSpecificationComponent."""

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        comp = get_object_or_404(
            ContractorSpecificationComponent, pk=pk, specification__company=company
        )
        spec = comp.specification
        comp.delete()
        return JsonResponse(
            {
                "ok": True,
                "spec_rate_per_unit": format_num(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteContractorSpecificationView(View):
    """AJAX endpoint to delete a ContractorSpecification."""

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        spec = get_object_or_404(ContractorSpecification, pk=pk, company=company)
        spec.delete()
        return JsonResponse({"ok": True})


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorSpecificationView(View):
    """AJAX endpoint to update fields on a ContractorSpecification (material spec)."""

    ALLOWED_FIELDS = {
        "section": "str",
        "name": "str",
        "unit_label": "str",
        "is_active": "bool",
    }

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        item = get_object_or_404(ContractorSpecification, pk=pk, company=company)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True})


# ── Contractor Labour Crews ───────────────────────────────────────────


class ContractorLabourCrewListView(ContractorLibraryMixin, ListView):
    model = ContractorLabourCrew
    template_name = "estimator/contractor/labour_crew_list.html"
    context_object_name = "crews"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", ContractorLabourCrewForm())
        context["f_q"] = self.request.GET.get("q", "")
        return context

    def post(self, request, *args, **kwargs):
        qs = ContractorLabourCrew.objects.filter(company=self.get_company())
        if _handle_clear_action(request, qs, label="labour crews"):
            return redirect(reverse("estimator:ctr_labour_crews"))
        if _handle_bulk_action(request, qs):
            return redirect(reverse("estimator:ctr_labour_crews"))
        if request.POST.get("action") == "sync_system":
            from .services import sync_labour_costs_to_contractor

            result = sync_labour_costs_to_contractor(self.get_company())
            messages.success(
                request,
                f"Labour crews synced with system library — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect(reverse("estimator:ctr_labour_crews"))

        form = ContractorLabourCrewForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = self.get_company()
            obj.save()
            return redirect(reverse("estimator:ctr_labour_crews"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorLabourCrewView(View):
    ALLOWED_FIELDS = {
        "crew_type",
        "skilled",
        "semi_skilled",
        "general",
        "daily_production",
        "skilled_rate",
        "semi_skilled_rate",
        "general_rate",
    }

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        crew = get_object_or_404(ContractorLabourCrew, pk=pk, company=company)
        data = json.loads(request.body)
        field, value = data.get("field"), data.get("value")
        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": "Invalid field"}, status=400)
        setattr(crew, field, value)
        crew.save()
        return JsonResponse(
            {
                "ok": True,
                "crew_size": crew.crew_size,
                "crew_daily_cost": str(crew.crew_daily_cost),
            }
        )


class ContractorLabourCrewUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Labour Crews"
        context["upload_description"] = (
            "Upload labour crew compositions and daily rates."
        )
        context["parent_template"] = "estimator/contractor/base_contractor.html"
        context["download_url_name"] = "estimator:ctr_download_labour_crew_template"
        return context

    def form_valid(self, form):
        from .importers import LabourCostImporter

        return _handle_upload(
            self.request,
            LabourCostImporter,
            "estimator:ctr_labour_crews",
            "Labour Crews",
            company=self.get_company(),
        )


class DownloadContractorLabourCrewTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Crew Type",
                "Crew Size",
                "Skilled",
                "Semi Skilled",
                "General",
                "Daily Production",
                "Skilled Rate",
                "Semi Skilled Rate",
                "General Rate",
            ],
            "contractor_labour_crews_template.xlsx",
        )


# ── Contractor Labour Specs ───────────────────────────────────────────


class ContractorLabourSpecListView(ContractorLibraryMixin, ListView):
    model = ContractorLabourSpecification
    template_name = "estimator/contractor/labour_spec_list.html"
    context_object_name = "labour_specs"
    paginate_by = 50

    def get_queryset(self):
        qs = super().get_queryset().select_related("crew")
        section = self.request.GET.get("section")
        if section:
            qs = qs.filter(section=section)
        trade_name = self.request.GET.get("trade_name")
        if trade_name:
            qs = qs.filter(trade_name=trade_name)
        name = self.request.GET.get("name")
        if name:
            qs = qs.filter(name=name)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.get_company()
        context["form"] = context.get(
            "form", ContractorLabourSpecificationForm(company=company)
        )
        all_specs = ContractorLabourSpecification.objects.filter(company=company)
        context["sections"] = (
            all_specs.exclude(section="")
            .values_list("section", flat=True)
            .distinct()
            .order_by("section")
        )
        context["trade_names"] = (
            all_specs.exclude(trade_name="")
            .values_list("trade_name", flat=True)
            .distinct()
            .order_by("trade_name")
        )
        context["names"] = (
            all_specs.values_list("name", flat=True).distinct().order_by("name")
        )
        context["f_section"] = self.request.GET.get("section", "")
        context["f_trade_name"] = self.request.GET.get("trade_name", "")
        context["f_name"] = self.request.GET.get("name", "")
        context["f_q"] = self.request.GET.get("q", "")
        context["crews"] = ContractorLabourCrew.objects.filter(company=company)
        context["query_params"] = _pagination_query_params(self.request)
        context.update(_spec_datalist_context(all_specs))
        return context

    def post(self, request, *args, **kwargs):
        qs = ContractorLabourSpecification.objects.filter(company=self.get_company())
        if _handle_clear_action(request, qs, label="labour specs"):
            return redirect(reverse("estimator:ctr_labour_specs"))
        if _handle_bulk_action(request, qs, allow_toggle_active=True):
            return redirect(reverse("estimator:ctr_labour_specs"))
        if request.POST.get("action") == "sync_system":
            from .services import sync_labour_specs_to_contractor

            result = sync_labour_specs_to_contractor(self.get_company())
            messages.success(
                request,
                f"Labour specs synced with system library — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect(reverse("estimator:ctr_labour_specs"))

        company = self.get_company()
        form = ContractorLabourSpecificationForm(request.POST, company=company)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = company
            obj.save()
            return redirect(reverse("estimator:ctr_labour_specs"))
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorLabourSpecView(View):
    ALLOWED_FIELDS = {
        "section": "str",
        "trade_name": "str",
        "name": "str",
        "unit": "str",
        "crew": "fk",
        "daily_production": "decimal",
        "team_mix": "decimal",
        "site_factor": "decimal",
        "tools_factor": "decimal",
        "leadership_factor": "decimal",
        "is_active": "bool",
    }

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        ls = get_object_or_404(ContractorLabourSpecification, pk=pk, company=company)
        data = json.loads(request.body)
        field, value = data.get("field"), data.get("value")
        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": "Invalid field"}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(ls, field, Decimal(str(value)))
            elif field_type == "fk":
                if value == "" or value is None:
                    ls.crew = None
                else:
                    ls.crew = get_object_or_404(
                        ContractorLabourCrew, pk=int(value), company=company
                    )
            elif field_type == "bool":
                setattr(ls, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(ls, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        ls.save()
        return JsonResponse(
            {
                "ok": True,
                "daily_output": str(ls.daily_output),
                "daily_cost": str(ls.daily_cost),
                "rate_per_unit": str(ls.rate_per_unit),
                "crew_type": ls.crew.crew_type if ls.crew else None,
            }
        )


class ContractorLabourSpecUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Labour Specifications"
        context["upload_description"] = (
            "Upload labour specification definitions with crew assignments and factors."
        )
        context["parent_template"] = "estimator/contractor/base_contractor.html"
        context["download_url_name"] = "estimator:ctr_download_labour_spec_template"
        return context

    def form_valid(self, form):
        from .importers import LabourSpecImporter

        return _handle_upload(
            self.request,
            LabourSpecImporter,
            "estimator:ctr_labour_specs",
            "Labour Specifications",
            company=self.get_company(),
        )


class DownloadContractorLabourSpecTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Section",
                "Trade Name",
                "Name",
                "Unit",
                "Crew Type",
                "Daily Production",
                "Team Mix",
                "Site Factor",
                "Tools Factor",
                "Leadership Factor",
            ],
            "contractor_labour_specs_template.xlsx",
        )


# ── Contractor Plant Costs ─────────────────────────────────────────


class ContractorPlantCostListView(ContractorLibraryMixin, ListView):
    model = ContractorPlantCost
    template_name = "estimator/contractor/plant_cost_list.html"
    context_object_name = "plants"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", ContractorPlantCostForm())
        context["f_q"] = self.request.GET.get("q", "")
        return context

    def post(self, request, *args, **kwargs):
        qs = ContractorPlantCost.objects.filter(company=self.get_company())
        if _handle_clear_action(request, qs, label="plant costs"):
            return redirect("estimator:ctr_plant_costs")
        if _handle_bulk_action(request, qs):
            return redirect("estimator:ctr_plant_costs")
        if request.POST.get("action") == "sync_system":
            from .services import sync_plant_costs_to_contractor

            result = sync_plant_costs_to_contractor(self.get_company())
            messages.success(
                request,
                f"Plant costs synced with system library — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect("estimator:ctr_plant_costs")

        form = ContractorPlantCostForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = self.get_company()
            obj.save()
            return redirect("estimator:ctr_plant_costs")
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorPlantCostView(View):
    ALLOWED_FIELDS = {
        "name": "str",
        "hourly_production": "decimal",
        "hourly_rate": "decimal",
    }

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        item = get_object_or_404(ContractorPlantCost, pk=pk, company=company)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True})


class ContractorPlantCostUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Plant Costs"
        context["upload_description"] = (
            "Upload plant & equipment costs with hourly production and rates."
        )
        context["parent_template"] = "estimator/contractor/base_contractor.html"
        context["download_url_name"] = "estimator:ctr_download_plant_cost_template"
        return context

    def form_valid(self, form):
        from .importers import PlantCostImporter

        return _handle_upload(
            self.request,
            PlantCostImporter,
            "estimator:ctr_plant_costs",
            "Plant Costs",
            company=self.get_company(),
        )


class DownloadContractorPlantCostTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            ["Plant & Equipment", "Hourly Production", "Hourly Rate"],
            "contractor_plant_costs_template.xlsx",
        )


# ── Contractor Plant Specs ─────────────────────────────────────────


class ContractorPlantSpecListView(ContractorLibraryMixin, ListView):
    model = ContractorPlantSpecification
    template_name = "estimator/contractor/plant_spec_list.html"
    context_object_name = "plant_specs"
    paginate_by = 50

    def get_queryset(self):
        return super().get_queryset().prefetch_related("components__plant_type")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.get_company()
        context["form"] = context.get(
            "form", ContractorPlantSpecificationForm(company=company)
        )
        context["plants"] = ContractorPlantCost.objects.filter(company=company)
        context["f_q"] = self.request.GET.get("q", "")
        context["query_params"] = _pagination_query_params(self.request)
        context.update(
            _spec_datalist_context(
                ContractorPlantSpecification.objects.filter(company=company)
            )
        )
        return context

    def post(self, request, *args, **kwargs):
        qs = ContractorPlantSpecification.objects.filter(company=self.get_company())
        if _handle_clear_action(request, qs, label="plant specs"):
            return redirect("estimator:ctr_plant_specs")
        if _handle_bulk_action(request, qs, allow_toggle_active=True):
            return redirect("estimator:ctr_plant_specs")
        if request.POST.get("action") == "sync_system":
            from .services import sync_plant_specs_to_contractor

            result = sync_plant_specs_to_contractor(self.get_company())
            messages.success(
                request,
                f"Plant specs synced with system library — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect("estimator:ctr_plant_specs")

        form = ContractorPlantSpecificationForm(
            request.POST, company=self.get_company()
        )
        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = self.get_company()
            obj.save()
            return redirect("estimator:ctr_plant_specs")
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorPlantSpecView(View):
    ALLOWED_FIELDS = {
        "section": "str",
        "trade_name": "str",
        "name": "str",
        "unit": "str",
        "daily_production": "decimal",
        "operator_factor": "decimal",
        "site_factor": "decimal",
        "is_active": "bool",
    }

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        item = get_object_or_404(ContractorPlantSpecification, pk=pk, company=company)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            elif field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse(
            {
                "ok": True,
                "daily_output": str(item.daily_output),
                "daily_cost": str(item.daily_cost),
                "rate_per_unit": str(item.rate_per_unit),
            }
        )


class ContractorPlantSpecUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Plant Specifications"
        context["upload_description"] = (
            "Upload plant specification definitions with production factors."
        )
        context["parent_template"] = "estimator/contractor/base_contractor.html"
        context["download_url_name"] = "estimator:ctr_download_plant_spec_template"
        return context

    def form_valid(self, form):
        from .importers import PlantSpecImporter

        return _handle_upload(
            self.request,
            PlantSpecImporter,
            "estimator:ctr_plant_specs",
            "Plant Specifications",
            company=self.get_company(),
        )


class DownloadContractorPlantSpecTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Section",
                "Trade Name",
                "Plant Specification",
                "Unit",
                "Plant Type",
                "Daily Production",
                "Operator",
                "Site",
            ],
            "contractor_plant_specs_template.xlsx",
        )


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorPlantSpecComponentView(View):
    """AJAX endpoint to update plant_type or hours on a contractor plant spec component."""

    ALLOWED_FIELDS = {"plant_type", "hours"}

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        comp = get_object_or_404(
            ContractorPlantSpecificationComponent,
            pk=pk,
            specification__company=company,
        )
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        try:
            if field == "hours":
                comp.hours = Decimal(str(value or 0))
            else:
                if value in (None, "", 0, "0"):
                    comp.plant_type = None
                else:
                    comp.plant_type = ContractorPlantCost.objects.filter(
                        pk=int(value), company=company
                    ).first()
                    if comp.plant_type is None:
                        return JsonResponse(
                            {"error": "Plant type not found"}, status=404
                        )
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        comp.save()
        spec = comp.specification
        return JsonResponse(
            {
                "ok": True,
                "daily_cost": str(spec.daily_cost),
                "rate_per_unit": str(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class AddContractorPlantSpecComponentView(View):
    """AJAX endpoint to add a new component to a ContractorPlantSpecification."""

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        spec = get_object_or_404(ContractorPlantSpecification, pk=pk, company=company)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}

        plant_type_id = data.get("plant_type")
        hours_raw = data.get("hours", 0)

        plant_type = None
        if plant_type_id not in (None, "", 0, "0"):
            try:
                plant_type = ContractorPlantCost.objects.filter(
                    pk=int(plant_type_id), company=company
                ).first()
            except (TypeError, ValueError):
                plant_type = None

        try:
            hours = Decimal(str(hours_raw or 0))
        except Exception:
            hours = Decimal("0")

        next_order = (
            spec.components.aggregate(models.Max("sort_order"))["sort_order__max"] or 0
        ) + 1
        comp = ContractorPlantSpecificationComponent.objects.create(
            specification=spec,
            plant_type=plant_type,
            hours=hours,
            sort_order=next_order,
        )
        return JsonResponse(
            {
                "ok": True,
                "component": {
                    "id": comp.id,
                    "plant_type_id": comp.plant_type_id,
                    "plant_type_name": comp.plant_type.name if comp.plant_type else "",
                    "hours": str(comp.hours),
                },
                "daily_cost": str(spec.daily_cost),
                "rate_per_unit": str(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteContractorPlantSpecComponentView(View):
    """AJAX endpoint to delete a ContractorPlantSpecificationComponent."""

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        comp = get_object_or_404(
            ContractorPlantSpecificationComponent,
            pk=pk,
            specification__company=company,
        )
        spec = comp.specification
        comp.delete()
        return JsonResponse(
            {
                "ok": True,
                "daily_cost": str(spec.daily_cost),
                "rate_per_unit": str(spec.rate_per_unit),
            }
        )


@method_decorator(csrf_exempt, name="dispatch")
class DeleteContractorPlantSpecificationView(View):
    """AJAX endpoint to delete a ContractorPlantSpecification."""

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        spec = get_object_or_404(ContractorPlantSpecification, pk=pk, company=company)
        spec.delete()
        return JsonResponse({"ok": True})


# ── Contractor Preliminary Costs ──────────────────────────────────


class ContractorPreliminaryCostListView(ContractorLibraryMixin, ListView):
    model = ContractorPreliminaryCost
    template_name = "estimator/contractor/preliminary_cost_list.html"
    context_object_name = "preliminaries"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get("form", ContractorPreliminaryCostForm())
        context["f_q"] = self.request.GET.get("q", "")
        return context

    def post(self, request, *args, **kwargs):
        qs = ContractorPreliminaryCost.objects.filter(company=self.get_company())
        if _handle_clear_action(request, qs, label="preliminary costs"):
            return redirect("estimator:ctr_preliminary_costs")
        if _handle_bulk_action(request, qs):
            return redirect("estimator:ctr_preliminary_costs")
        if request.POST.get("action") == "sync_system":
            from .services import sync_preliminary_costs_to_contractor

            result = sync_preliminary_costs_to_contractor(self.get_company())
            messages.success(
                request,
                f"Preliminary costs synced with system library — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect("estimator:ctr_preliminary_costs")

        form = ContractorPreliminaryCostForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = self.get_company()
            obj.save()
            return redirect("estimator:ctr_preliminary_costs")
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorPreliminaryCostView(View):
    ALLOWED_FIELDS = {
        "name": "str",
        "preliminary_type": "str",
        "sum_value": "decimal",
        "amount": "decimal",
        "number_per_month": "decimal",
        "monthly_rate": "decimal",
        "months": "decimal",
    }

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        item = get_object_or_404(ContractorPreliminaryCost, pk=pk, company=company)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True, "computed_amount": str(item.computed_amount)})


class ContractorPreliminaryCostUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Preliminary Costs"
        context["upload_description"] = (
            "Upload preliminary costs with type, amounts and time-based rates."
        )
        context["parent_template"] = "estimator/contractor/base_contractor.html"
        context["download_url_name"] = (
            "estimator:ctr_download_preliminary_cost_template"
        )
        return context

    def form_valid(self, form):
        from .importers import PreliminaryCostImporter

        return _handle_upload(
            self.request,
            PreliminaryCostImporter,
            "estimator:ctr_preliminary_costs",
            "Preliminary Costs",
            company=self.get_company(),
        )


class DownloadContractorPreliminaryCostTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            [
                "Preliminary Type",
                "Name",
                "Sum",
                "Amount",
                "Number/Month",
                "Monthly Rate",
                "Months",
            ],
            "contractor_preliminary_costs_template.xlsx",
        )


# ── Contractor Preliminary Specs ──────────────────────────────────


class ContractorPreliminarySpecListView(ContractorLibraryMixin, ListView):
    model = ContractorPreliminarySpecification
    template_name = "estimator/contractor/preliminary_spec_list.html"
    context_object_name = "preliminary_specs"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = context.get(
            "form",
            ContractorPreliminarySpecificationForm(company=self.get_company()),
        )
        context["preliminary_type_choices"] = (
            SystemPreliminaryCost.PRELIMINARY_TYPE_CHOICES
        )
        company = self.get_company()
        context["f_q"] = self.request.GET.get("q", "")
        context.update(
            _spec_datalist_context(
                ContractorPreliminarySpecification.objects.filter(company=company)
            )
        )
        return context

    def post(self, request, *args, **kwargs):
        qs = ContractorPreliminarySpecification.objects.filter(
            company=self.get_company()
        )
        if _handle_clear_action(request, qs, label="preliminary specs"):
            return redirect("estimator:ctr_preliminary_specs")
        if _handle_bulk_action(request, qs, allow_toggle_active=True):
            return redirect("estimator:ctr_preliminary_specs")
        if request.POST.get("action") == "sync_system":
            from .services import sync_preliminary_specs_to_contractor

            result = sync_preliminary_specs_to_contractor(self.get_company())
            messages.success(
                request,
                f"Preliminary specs synced with system library — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect("estimator:ctr_preliminary_specs")

        form = ContractorPreliminarySpecificationForm(
            request.POST, company=self.get_company()
        )
        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = self.get_company()
            obj.save()
            return redirect("estimator:ctr_preliminary_specs")
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data(form=form))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorPreliminarySpecView(View):
    ALLOWED_FIELDS = {
        "section": "str",
        "trade_name": "str",
        "name": "str",
        "unit": "str",
        "preliminary_type": "str",
        "is_active": "bool",
    }

    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        item = get_object_or_404(
            ContractorPreliminarySpecification, pk=pk, company=company
        )
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        field = data.get("field")
        value = data.get("value")

        if field not in self.ALLOWED_FIELDS:
            return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)

        field_type = self.ALLOWED_FIELDS[field]
        try:
            if field_type == "decimal":
                setattr(item, field, Decimal(str(value)))
            elif field_type == "bool":
                setattr(item, field, bool(value) and value not in ("false", "0", 0))
            else:
                setattr(item, field, str(value))
        except Exception:
            return JsonResponse({"error": "Invalid value"}, status=400)

        item.save()
        return JsonResponse({"ok": True, "amount": str(item.amount)})


class ContractorPreliminarySpecUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["upload_title"] = "Upload Preliminary Specifications"
        context["upload_description"] = "Upload preliminary specification definitions."
        context["parent_template"] = "estimator/contractor/base_contractor.html"
        context["download_url_name"] = (
            "estimator:ctr_download_preliminary_spec_template"
        )
        return context

    def form_valid(self, form):
        from .importers import PreliminarySpecImporter

        return _handle_upload(
            self.request,
            PreliminarySpecImporter,
            "estimator:ctr_preliminary_specs",
            "Preliminary Specifications",
            company=self.get_company(),
        )


class DownloadContractorPreliminarySpecTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            ["Section", "Trade Name", "Name", "Unit", "Preliminary Type"],
            "contractor_preliminary_specs_template.xlsx",
        )


# ── Item Library (Project / Contractor / System) ──────────────────


_ITEM_LIBRARY_COLUMNS = [
    (
        "Trade Code",
        "Full trade code label (e.g. 'PRE-Preliminaries' or just the prefix)",
    ),
    ("Accounts Code", "Optional finance/accounts mapping code"),
    ("Component", "Component / category prefix (e.g. 'Concrete 25MPa/19mm - ')"),
    ("Material Specification", "Name of an existing material spec; left blank if none"),
    (
        "Labour & Plant Specification",
        "Name of a labour and/or plant spec; resolved against both tables",
    ),
    ("Preliminaries", "Name of an existing preliminary spec; left blank if none"),
    ("Item Description", "Required — the BoQ row description"),
    ("Unit", "Unit of measurement (e.g. m3, m2, t, SUM)"),
    ("Item Code", "Optional — short identifier used to pick this entry on a BoQ row"),
]
_ITEM_LIBRARY_HEADERS = [c[0] for c in _ITEM_LIBRARY_COLUMNS]


class ItemLibraryListView(ProjectEstimatorMixin, ListView):
    model = ProjectItemLibraryEntry
    template_name = "estimator/item_library_list.html"
    context_object_name = "entries"
    paginate_by = 100

    def get_queryset(self):
        project = self.get_project()
        qs = (
            ProjectItemLibraryEntry.objects.filter(project=project)
            .select_related(
                "trade_code",
                "material_spec",
                "labour_spec",
                "plant_spec",
                "preliminary_spec",
            )
            .order_by("item_code", "display_order", "id")
        )
        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(description__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        project = self.get_project()
        ctx["trade_codes"] = ProjectTradeCode.objects.filter(project=project)
        ctx["material_specs"] = ProjectSpecification.objects.filter(
            project=project
        ).order_by("name")
        ctx["labour_specs"] = ProjectLabourSpecification.objects.filter(
            project=project
        ).order_by("name")
        ctx["plant_specs"] = ProjectPlantSpecification.objects.filter(
            project=project
        ).order_by("name")
        ctx["labour_plant_spec_names"] = sorted(
            set(ctx["labour_specs"].values_list("name", flat=True))
            | set(ctx["plant_specs"].values_list("name", flat=True))
        )
        ctx["preliminary_specs"] = ProjectPreliminarySpecification.objects.filter(
            project=project
        ).order_by("name")
        ctx["f_trade_code"] = self.request.GET.get("trade_code", "")
        ctx["f_q"] = self.request.GET.get("q", "")
        ctx["query_params"] = _pagination_query_params(self.request)
        return ctx

    def post(self, request, *args, **kwargs):
        project = self.get_project()
        action = request.POST.get("action")
        if _handle_clear_action(
            request,
            ProjectItemLibraryEntry.objects.filter(project=project),
            label="item library entries",
        ):
            return redirect(
                reverse(
                    "estimator:item_library",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if action == "sync_system":
            from .services import sync_item_library_from_system

            result = sync_item_library_from_system(project)
            messages.success(
                request,
                f"Item Library synced from system — "
                f"{result['updated']} updated, {result['created']} new.",
            )
            return redirect(
                reverse(
                    "estimator:item_library",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        if action == "sync_contractor":
            from .services import sync_item_library_from_contractor

            result = sync_item_library_from_contractor(project)
            _flash_sync_result(request, result, "Item Library")
            return redirect(
                reverse(
                    "estimator:item_library",
                    kwargs={"project_pk": self.kwargs["project_pk"]},
                )
            )
        return redirect(
            reverse(
                "estimator:item_library",
                kwargs={"project_pk": self.kwargs["project_pk"]},
            )
        )


class ItemLibraryUploadView(CloneFromProjectMixin, ProjectEstimatorMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_success_url(self):
        return reverse(
            "estimator:item_library", kwargs={"project_pk": self.kwargs["project_pk"]}
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/base.html"
        ctx["upload_title"] = "Upload Item Library"
        ctx["upload_description"] = (
            "Upload pre-configured BoQ template rows from an Excel sheet."
        )
        ctx["download_url_name"] = "estimator:download_item_library_template"
        ctx["columns"] = _ITEM_LIBRARY_COLUMNS
        ctx["notes"] = [
            "Spec columns are matched against existing project specs by name. "
            "Unmatched names are stored blank and reported as warnings.",
            "Existing entries with the same component + description are updated, not duplicated.",
        ]
        return ctx

    def form_valid(self, form):
        from .importers import ItemLibraryImporter

        return _handle_upload(
            self.request,
            ItemLibraryImporter,
            self.get_success_url(),
            "Item Library",
            project=self.get_project(),
        )


class DownloadItemLibraryTemplateView(View):
    def get(self, request, project_pk):
        return _generate_template(_ITEM_LIBRARY_HEADERS, "item_library_template.xlsx")


# Contractor scope


class ContractorItemLibraryListView(ContractorLibraryMixin, ListView):
    model = ContractorItemLibraryEntry
    template_name = "estimator/contractor/item_library_list.html"
    context_object_name = "entries"
    paginate_by = 100

    def get_queryset(self):
        company = self.get_company()
        if company is None:
            return ContractorItemLibraryEntry.objects.none()
        qs = (
            ContractorItemLibraryEntry.objects.filter(company=company)
            .select_related(
                "trade_code",
                "material_spec",
                "labour_spec",
                "plant_spec",
                "preliminary_spec",
            )
            .order_by("item_code", "display_order", "id")
        )
        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(description__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        company = self.get_company()
        if company is None:
            ctx["trade_codes"] = ContractorTradeCode.objects.none()
            ctx["material_specs"] = ContractorSpecification.objects.none()
            ctx["labour_specs"] = ContractorLabourSpecification.objects.none()
            ctx["plant_specs"] = ContractorPlantSpecification.objects.none()
            ctx["preliminary_specs"] = ContractorPreliminarySpecification.objects.none()
        else:
            ctx["trade_codes"] = ContractorTradeCode.objects.filter(company=company)
            ctx["material_specs"] = ContractorSpecification.objects.filter(
                company=company
            ).order_by("name")
            ctx["labour_specs"] = ContractorLabourSpecification.objects.filter(
                company=company
            ).order_by("name")
            ctx["plant_specs"] = ContractorPlantSpecification.objects.filter(
                company=company
            ).order_by("name")
            ctx["preliminary_specs"] = (
                ContractorPreliminarySpecification.objects.filter(
                    company=company
                ).order_by("name")
            )
        ctx["labour_plant_spec_names"] = sorted(
            set(ctx["labour_specs"].values_list("name", flat=True))
            | set(ctx["plant_specs"].values_list("name", flat=True))
        )
        ctx["f_trade_code"] = self.request.GET.get("trade_code", "")
        ctx["f_q"] = self.request.GET.get("q", "")
        ctx["query_params"] = _pagination_query_params(self.request)
        return ctx

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request,
            ContractorItemLibraryEntry.objects.filter(company=self.get_company()),
            label="item library entries",
        ):
            return redirect(reverse("estimator:ctr_item_library"))
        if request.POST.get("action") == "sync_system":
            from .services import sync_item_library_to_contractor

            result = sync_item_library_to_contractor(self.get_company())
            messages.success(
                request,
                f"Item Library synced with system — "
                f"{result['updated']} updated, {result['created']} new.",
            )
        return redirect(reverse("estimator:ctr_item_library"))


class ContractorItemLibraryUploadView(ContractorLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/contractor/base_contractor.html"
        ctx["upload_title"] = "Upload Item Library"
        ctx["upload_description"] = (
            "Upload pre-configured BoQ template rows from an Excel sheet."
        )
        ctx["download_url_name"] = "estimator:ctr_download_item_library_template"
        ctx["columns"] = _ITEM_LIBRARY_COLUMNS
        return ctx

    def form_valid(self, form):
        from .importers import ItemLibraryImporter

        return _handle_upload(
            self.request,
            ItemLibraryImporter,
            "estimator:ctr_item_library",
            "Item Library",
            company=self.get_company(),
        )


class DownloadContractorItemLibraryTemplateView(ContractorLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            _ITEM_LIBRARY_HEADERS, "contractor_item_library_template.xlsx"
        )


# System scope


class SystemItemLibraryListView(SystemLibraryMixin, ListView):
    model = SystemItemLibraryEntry
    template_name = "estimator/system/item_library_list.html"
    context_object_name = "entries"
    paginate_by = 100

    def get_queryset(self):
        qs = (
            SystemItemLibraryEntry.objects.all()
            .select_related(
                "trade_code",
                "material_spec",
                "labour_spec",
                "plant_spec",
                "preliminary_spec",
            )
            .order_by("item_code", "display_order", "id")
        )
        trade_code = self.request.GET.get("trade_code")
        if trade_code:
            qs = qs.filter(trade_code__id=trade_code)
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(description__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["trade_codes"] = SystemTradeCode.objects.all()
        ctx["material_specs"] = SystemSpecification.objects.all().order_by("name")
        ctx["labour_specs"] = SystemLabourSpecification.objects.all().order_by("name")
        ctx["plant_specs"] = SystemPlantSpecification.objects.all().order_by("name")
        ctx["labour_plant_spec_names"] = sorted(
            set(ctx["labour_specs"].values_list("name", flat=True))
            | set(ctx["plant_specs"].values_list("name", flat=True))
        )
        ctx["preliminary_specs"] = (
            SystemPreliminarySpecification.objects.all().order_by("name")
        )
        ctx["f_trade_code"] = self.request.GET.get("trade_code", "")
        ctx["f_q"] = self.request.GET.get("q", "")
        ctx["query_params"] = _pagination_query_params(self.request)
        return ctx

    def post(self, request, *args, **kwargs):
        if _handle_clear_action(
            request,
            SystemItemLibraryEntry.objects.all(),
            label="item library entries",
        ):
            return redirect(reverse("estimator:sys_item_library"))
        return redirect(reverse("estimator:sys_item_library"))


class SystemItemLibraryUploadView(SystemLibraryMixin, FormView):
    template_name = "estimator/upload_generic.html"
    form_class = ExcelImportForm

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["parent_template"] = "estimator/system/base_system.html"
        ctx["upload_title"] = "Upload Item Library"
        ctx["upload_description"] = (
            "Upload pre-configured BoQ template rows from an Excel sheet."
        )
        ctx["download_url_name"] = "estimator:sys_download_item_library_template"
        ctx["columns"] = _ITEM_LIBRARY_COLUMNS
        return ctx

    def form_valid(self, form):
        from .importers import ItemLibraryImporter

        return _handle_upload(
            self.request,
            ItemLibraryImporter,
            "estimator:sys_item_library",
            "Item Library",
        )


class DownloadSystemItemLibraryTemplateView(SystemLibraryMixin, View):
    def get(self, request):
        return _generate_template(
            _ITEM_LIBRARY_HEADERS, "system_item_library_template.xlsx"
        )


# ─────────────────────────────────────────────────────────────────────
# Item Library — inline edit / create / delete (project, contractor, system)
# ─────────────────────────────────────────────────────────────────────


_ITEM_LIBRARY_TEXT_FIELDS = {
    "item_code",
    "accounts_code",
    "component",
    "description",
    "unit",
}
_ITEM_LIBRARY_INT_FIELDS = {"display_order"}
# FK field → raw-name mirror field, kept in sync so the UI can flag unlinked
# spec values in red while preserving the originally uploaded name.
_ITEM_LIBRARY_SPEC_NAME_ATTRS = {
    "material_spec": "material_spec_name",
    "preliminary_spec": "preliminary_spec_name",
}


def _apply_item_library_field(entry, field, value, fk_models):
    """Mutate `entry` based on `{field, value}`. Returns None on success or an
    error JsonResponse. `fk_models` maps fk field name → related model class."""
    if field in _ITEM_LIBRARY_TEXT_FIELDS:
        setattr(entry, field, (value or "").strip() if value is not None else "")
        return None
    if field in _ITEM_LIBRARY_INT_FIELDS:
        try:
            setattr(entry, field, int(value or 0))
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid value"}, status=400)
        return None
    if field == "labour_plant_spec":
        # Combined column: `value` is a spec name resolved against both
        # the labour and plant tables. Whichever side matches wins; the
        # other is cleared so we don't leave a stale pairing.
        labour_model, labour_scope = fk_models.get("labour_spec", (None, {}))
        plant_model, plant_scope = fk_models.get("plant_spec", (None, {}))
        if value in (None, "", 0, "0"):
            entry.labour_spec = None
            entry.plant_spec = None
            entry.labour_plant_spec_name = ""
            return None
        name = str(value).strip()
        entry.labour_spec = (
            labour_model.objects.filter(name=name, **labour_scope).first()
            if labour_model
            else None
        )
        entry.plant_spec = (
            plant_model.objects.filter(name=name, **plant_scope).first()
            if plant_model
            else None
        )
        entry.labour_plant_spec_name = name
        return None
    if field in fk_models:
        model_cls, scope_filter = fk_models[field]
        # Keep the raw-name mirror in sync so the column reads as linked
        # (no red flag) once a spec is chosen, and reverts to blank when cleared.
        name_attr = _ITEM_LIBRARY_SPEC_NAME_ATTRS.get(field)
        if value in (None, "", 0, "0"):
            setattr(entry, field, None)
            if name_attr:
                setattr(entry, name_attr, "")
            return None
        try:
            obj = model_cls.objects.filter(pk=int(value), **scope_filter).first()
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid id"}, status=400)
        if obj is None:
            return JsonResponse({"error": f"{field} not found"}, status=404)
        setattr(entry, field, obj)
        if name_attr:
            setattr(entry, name_attr, obj.name)
        return None
    return JsonResponse({"error": f'Field "{field}" not allowed'}, status=400)


@method_decorator(csrf_exempt, name="dispatch")
class UpdateItemLibraryEntryView(View):
    """AJAX: update a single field on a ProjectItemLibraryEntry."""

    def post(self, request, project_pk, pk):
        entry = get_object_or_404(ProjectItemLibraryEntry, pk=pk, project_id=project_pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        scope = {"project_id": project_pk}
        fk_models = {
            "trade_code": (ProjectTradeCode, scope),
            "material_spec": (ProjectSpecification, scope),
            "labour_spec": (ProjectLabourSpecification, scope),
            "plant_spec": (ProjectPlantSpecification, scope),
            "preliminary_spec": (ProjectPreliminarySpecification, scope),
        }
        err = _apply_item_library_field(
            entry, data.get("field"), data.get("value"), fk_models
        )
        if err is not None:
            return err
        entry.save()
        return JsonResponse({"ok": True})


class CreateItemLibraryEntryView(ProjectEstimatorMixin, View):
    """Create a ProjectItemLibraryEntry from the Add Item form (all fields
    entered up front), appended at the end of its item-code group."""

    FORM_FIELDS = (
        "item_code",
        "trade_code",
        "component",
        "description",
        "unit",
        "material_spec",
        "labour_plant_spec",
        "preliminary_spec",
    )

    def post(self, request, project_pk):
        project = self.get_project()
        list_url = reverse("estimator:item_library", kwargs={"project_pk": project_pk})
        entry = ProjectItemLibraryEntry(project=project)
        scope = {"project_id": project_pk}
        fk_models = {
            "trade_code": (ProjectTradeCode, scope),
            "material_spec": (ProjectSpecification, scope),
            "labour_spec": (ProjectLabourSpecification, scope),
            "plant_spec": (ProjectPlantSpecification, scope),
            "preliminary_spec": (ProjectPreliminarySpecification, scope),
        }
        for field in self.FORM_FIELDS:
            if field in request.POST:
                err = _apply_item_library_field(
                    entry, field, request.POST.get(field), fk_models
                )
                if err is not None:
                    messages.error(request, "Could not add item — invalid input.")
                    return redirect(list_url)

        if not (entry.item_code or entry.component or entry.description):
            messages.error(
                request,
                "Enter at least an item code, component or description.",
            )
            return redirect(list_url)

        last = ProjectItemLibraryEntry.objects.filter(
            project=project, item_code=entry.item_code
        ).aggregate(m=models.Max("display_order"))["m"]
        entry.display_order = (last or 0) + 1
        entry.save()
        messages.success(request, "Library item added.")
        return redirect(list_url)


@method_decorator(csrf_exempt, name="dispatch")
class ReorderItemLibraryEntriesView(ProjectEstimatorMixin, View):
    """AJAX: persist a drag-and-drop ordering. Body: {"order": [id, ...]}.

    Assigns display_order by the received position so the manual sequence
    sticks (within each item-code group, which remains the primary sort)."""

    def post(self, request, project_pk):
        project = self.get_project()
        try:
            order = json.loads(request.body).get("order", [])
        except (json.JSONDecodeError, AttributeError):
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        entries = {
            e.pk: e for e in ProjectItemLibraryEntry.objects.filter(project=project)
        }
        to_update = []
        for pos, raw_id in enumerate(order):
            try:
                e = entries.get(int(raw_id))
            except (TypeError, ValueError):
                continue
            if e is not None and e.display_order != pos:
                e.display_order = pos
                to_update.append(e)
        if to_update:
            ProjectItemLibraryEntry.objects.bulk_update(to_update, ["display_order"])
        return JsonResponse({"ok": True, "updated": len(to_update)})


class DeleteItemLibraryEntryView(ProjectEstimatorMixin, View):
    def post(self, request, project_pk, pk):
        entry = get_object_or_404(ProjectItemLibraryEntry, pk=pk, project_id=project_pk)
        entry.delete()
        messages.success(request, "Library entry deleted.")
        return redirect(
            reverse("estimator:item_library", kwargs={"project_pk": project_pk})
        )


class BulkDeleteItemLibraryEntriesView(ProjectEstimatorMixin, View):
    def post(self, request, project_pk):
        ids = request.POST.getlist("entry_ids")
        count, _ = ProjectItemLibraryEntry.objects.filter(
            pk__in=ids, project_id=project_pk
        ).delete()
        if count:
            messages.success(
                request, f"Deleted {count} library entr{'y' if count == 1 else 'ies'}."
            )
        else:
            messages.info(request, "No entries selected.")
        return redirect(
            reverse("estimator:item_library", kwargs={"project_pk": project_pk})
        )


@method_decorator(csrf_exempt, name="dispatch")
class UpdateContractorItemLibraryEntryView(View):
    def post(self, request, pk):
        company = _contractor_company_for(request)
        if company is None:
            return JsonResponse({"error": "Forbidden"}, status=403)
        entry = get_object_or_404(ContractorItemLibraryEntry, pk=pk, company=company)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        scope = {"company": company}
        fk_models = {
            "trade_code": (ContractorTradeCode, scope),
            "material_spec": (ContractorSpecification, scope),
            "labour_spec": (ContractorLabourSpecification, scope),
            "plant_spec": (ContractorPlantSpecification, scope),
            "preliminary_spec": (ContractorPreliminarySpecification, scope),
        }
        err = _apply_item_library_field(
            entry, data.get("field"), data.get("value"), fk_models
        )
        if err is not None:
            return err
        entry.save()
        return JsonResponse({"ok": True})


class CreateContractorItemLibraryEntryView(ContractorLibraryMixin, View):
    def post(self, request):
        company = self.get_company()
        ContractorItemLibraryEntry.objects.create(company=company, description="")
        messages.success(
            request, "New library entry added — fill in the fields inline."
        )
        return redirect(reverse("estimator:ctr_item_library"))


class DeleteContractorItemLibraryEntryView(ContractorLibraryMixin, View):
    def post(self, request, pk):
        company = self.get_company()
        entry = get_object_or_404(ContractorItemLibraryEntry, pk=pk, company=company)
        entry.delete()
        messages.success(request, "Library entry deleted.")
        return redirect(reverse("estimator:ctr_item_library"))


class BulkDeleteContractorItemLibraryEntriesView(ContractorLibraryMixin, View):
    def post(self, request):
        company = self.get_company()
        ids = request.POST.getlist("entry_ids")
        count, _ = ContractorItemLibraryEntry.objects.filter(
            pk__in=ids, company=company
        ).delete()
        if count:
            messages.success(
                request, f"Deleted {count} library entr{'y' if count == 1 else 'ies'}."
            )
        else:
            messages.info(request, "No entries selected.")
        return redirect(reverse("estimator:ctr_item_library"))


@method_decorator(csrf_exempt, name="dispatch")
class UpdateSystemItemLibraryEntryView(View):
    def post(self, request, pk):
        if not request.user.is_staff:
            return JsonResponse({"error": "Forbidden"}, status=403)
        entry = get_object_or_404(SystemItemLibraryEntry, pk=pk)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        fk_models = {
            "trade_code": (SystemTradeCode, {}),
            "material_spec": (SystemSpecification, {}),
            "labour_spec": (SystemLabourSpecification, {}),
            "plant_spec": (SystemPlantSpecification, {}),
            "preliminary_spec": (SystemPreliminarySpecification, {}),
        }
        err = _apply_item_library_field(
            entry, data.get("field"), data.get("value"), fk_models
        )
        if err is not None:
            return err
        entry.save()
        return JsonResponse({"ok": True})


class CreateSystemItemLibraryEntryView(SystemLibraryMixin, View):
    def post(self, request):
        SystemItemLibraryEntry.objects.create(description="")
        messages.success(
            request, "New library entry added — fill in the fields inline."
        )
        return redirect(reverse("estimator:sys_item_library"))


class DeleteSystemItemLibraryEntryView(SystemLibraryMixin, View):
    def post(self, request, pk):
        entry = get_object_or_404(SystemItemLibraryEntry, pk=pk)
        entry.delete()
        messages.success(request, "Library entry deleted.")
        return redirect(reverse("estimator:sys_item_library"))


class BulkDeleteSystemItemLibraryEntriesView(SystemLibraryMixin, View):
    def post(self, request):
        ids = request.POST.getlist("entry_ids")
        count, _ = SystemItemLibraryEntry.objects.filter(pk__in=ids).delete()
        if count:
            messages.success(
                request, f"Deleted {count} library entr{'y' if count == 1 else 'ies'}."
            )
        else:
            messages.info(request, "No entries selected.")
        return redirect(reverse("estimator:sys_item_library"))
