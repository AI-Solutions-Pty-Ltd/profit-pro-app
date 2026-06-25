"""
Individual sheet importers for each data tab.

Each importer handles a single Excel file upload using update_or_create
(safe for re-uploads) rather than delete-all-then-create.

Sheet matching is fuzzy: case-insensitive substring search across multiple
keyword variants (plural/singular, abbreviations, US/UK spelling).
Falls back to the active sheet when no keyword matches — so single-sheet
uploads just work.

Each importer accepts optional `project` and `company` parameters:
- When `project` is set: writes to Project* models scoped to that project.
- When `company` is set: writes to Contractor* models scoped to that company.
- When both are None: writes to System* models (library/admin mode).
"""

from decimal import Decimal, InvalidOperation

import openpyxl

from app.Account.models import Municipality

from .models import (
    ContractorItemLibraryEntry,
    ContractorLabourCrew,
    ContractorLabourSpecification,
    ContractorMaterial,
    ContractorPlantCost,
    ContractorPlantSpecification,
    ContractorPlantSpecificationComponent,
    ContractorPreliminaryCost,
    ContractorPreliminarySpecification,
    ContractorSpecification,
    ContractorSpecificationComponent,
    ContractorTradeCode,
    ProjectItemLibraryEntry,
    ProjectLabourCrew,
    ProjectLabourSpecification,
    ProjectMaterial,
    ProjectPlantCost,
    ProjectPlantSpecification,
    ProjectPlantSpecificationComponent,
    ProjectPreliminaryCost,
    ProjectPreliminarySpecification,
    ProjectSpecification,
    ProjectSpecificationComponent,
    ProjectTradeCode,
    SystemItemLibraryEntry,
    SystemLabourCrew,
    SystemLabourSpecification,
    SystemMaterial,
    SystemPlantCost,
    SystemPlantSpecification,
    SystemPlantSpecificationComponent,
    SystemPreliminaryCost,
    SystemPreliminarySpecification,
    SystemSpecification,
    SystemSpecificationComponent,
    SystemTradeCode,
)


def _find_sheet(wb, keywords):
    """Find a worksheet by fuzzy name matching."""
    for name in wb.sheetnames:
        normalised = name.lower().strip()
        for kw in keywords:
            if kw in normalised:
                return wb[name]
    return wb.active


def _find_sheet_with_name(wb, keywords):
    """Like _find_sheet but returns (worksheet, matched_sheet_name, fell_back)."""
    for name in wb.sheetnames:
        normalised = name.lower().strip()
        for kw in keywords:
            if kw in normalised:
                return wb[name], name, False
    return wb.active, wb.active.title, True


def _col_offset(ws):
    """Return 0-indexed offset for the first non-empty column (handles sheets
    where column A is blank, e.g. the master-file layout)."""
    mc = ws.min_column or 1
    return max(0, mc - 1)


def _find_header_row(ws, required_keywords, max_scan=10):
    """Return the 1-indexed row that contains every required keyword
    (case-insensitive substring). Falls back to row 1 when no match."""
    required = [kw.lower() for kw in required_keywords]
    last = min(ws.max_row or 1, max_scan)
    for r in range(1, last + 1):
        parts = []
        for c in range(1, (ws.max_column or 1) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                parts.append(str(v).lower())
        joined = " | ".join(parts)
        if all(kw in joined for kw in required):
            return r
    return 1


def _safe_decimal(value):
    if value is None:
        return None
    if isinstance(value, str) and (value.startswith("#") or not value.strip()):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _safe_str(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("#"):
        return ""
    return s


def _norm_trade(s):
    """Normalise a trade token for matching: alphanumerics only, upper-case.

    So "CFR", "cfr-", "C.F.R." all collapse to "CFR" and match an existing
    trade code with prefix "CFR-".
    """
    return "".join(ch for ch in (s or "") if ch.isalnum()).upper()


def _unique_trade_prefix(model, scope_kwargs, name):
    """Pick a prefix unique within the scope (<=20 chars)."""
    existing = set(
        model.objects.filter(**scope_kwargs).values_list("prefix", flat=True)
    )
    base = "".join(ch for ch in name if ch.isalnum())[:18].upper() or "TRADE"
    candidate, n = base, 1
    while candidate in existing:
        suffix = str(n)
        candidate = base[: 20 - len(suffix)] + suffix
        n += 1
    return candidate


def resolve_trade_code(raw, project=None, company=None):
    """Resolve (find-or-create) a TradeCode for a free-text trade value,
    scoped to project / company / system. Returns None if ``raw`` is blank.

    Matches on trade_name (case-insensitive) or the full ``prefix+trade_name``
    string; creates a new code (so uploads never silently drop a trade)
    when nothing matches.
    """
    raw = (raw or "").strip()
    if not raw:
        return None
    if project is not None:
        model, scope = ProjectTradeCode, {"project": project}
    elif company is not None:
        model, scope = ContractorTradeCode, {"company": company}
    else:
        model, scope = SystemTradeCode, {}

    raw_norm = _norm_trade(raw)
    for tc in model.objects.filter(**scope):
        if raw_norm and raw_norm in {
            _norm_trade(tc.trade_name),
            _norm_trade(tc.prefix),
            _norm_trade(f"{tc.prefix}{tc.trade_name}"),
        }:
            return tc
    return model.objects.create(
        prefix=_unique_trade_prefix(model, scope, raw),
        trade_name=raw[:100],
        **scope,
    )


# ── Trade Codes ──────────────────────────────────────────────────


class TradeCodeImporter:
    """Import Trade Codes from Excel.

    Expected columns: Prefix | Trade Name
    """

    SHEET_KEYWORDS = [
        "trade code",
        "trade codes",
        "tradecode",
        "tradecodes",
    ]

    def __init__(self, path, project=None, company=None):
        self.path = path
        self.project = project
        self.company = company

    def run(self):
        wb = openpyxl.load_workbook(self.path, data_only=True)
        ws = _find_sheet(wb, self.SHEET_KEYWORDS)
        created = updated = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            prefix = _safe_str(row[0]) if len(row) > 0 else ""
            trade_name = _safe_str(row[1]) if len(row) > 1 else ""
            if not prefix:
                skipped += 1
                continue

            if self.project:
                _, is_new = ProjectTradeCode.objects.update_or_create(
                    project=self.project,
                    prefix=prefix,
                    defaults={"trade_name": trade_name},
                )
            elif self.company:
                _, is_new = ContractorTradeCode.objects.update_or_create(
                    company=self.company,
                    prefix=prefix,
                    defaults={"trade_name": trade_name},
                )
            else:
                _, is_new = SystemTradeCode.objects.update_or_create(
                    prefix=prefix,
                    defaults={"trade_name": trade_name},
                )
            if is_new:
                created += 1
            else:
                updated += 1
        return {"created": created, "updated": updated, "skipped": skipped}


# ── Municipalities ────────────────────────────────────────────────


class MunicipalityImporter:
    """Import Municipalities from Excel.

    Expected columns: Province | Municipality Name | Code | District
    """

    SHEET_KEYWORDS = [
        "municipality",
        "municipalities",
        "province",
        "provinces",
    ]

    def __init__(self, path, project=None, company=None):
        self.path = path
        self.project = project
        self.company = company

    def run(self):
        wb = openpyxl.load_workbook(self.path, data_only=True)
        ws = _find_sheet(wb, self.SHEET_KEYWORDS)
        created = updated = skipped = 0

        # Find header row
        header_row_idx = _find_header_row(ws, ["province", "municipality", "code"])

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or not any(row):
                continue
            province = _safe_str(row[0]) if len(row) > 0 else ""
            municipality_name = _safe_str(row[1]) if len(row) > 1 else ""
            code = _safe_str(row[2]) if len(row) > 2 else ""
            district = _safe_str(row[3]) if len(row) > 3 else ""

            if not province or not municipality_name or not code:
                skipped += 1
                continue

            obj, is_new = Municipality.objects.update_or_create(
                province=province,
                municipality_name=municipality_name,
                code=code,
                defaults={"district": district},
            )
            if is_new:
                created += 1
            else:
                updated += 1
        return {"created": created, "updated": updated, "skipped": skipped}


# ── Material Costs ───────────────────────────────────────────────


class MaterialCostImporter:
    """Import Material costs from Excel.

    New layout (preferred):
        Trade Name | Material Code | Unit | Pack Qty | Pack Cost | Variety | Spec

    Legacy layout (still accepted — pack_qty defaults to 1):
        Trade Name | Material Code | Unit | Market Rate | Variety | Spec

    Column positions are resolved by header row 1 (case-insensitive). If the
    header doesn't include "Pack Qty"/"Pack Cost" the importer falls back to
    "Market Rate" with pack_qty=1.
    """

    SHEET_KEYWORDS = [
        "material cost",
        "material costs",
        "materialcost",
        "materialcosts",
        "mat cost",
        "mat costs",
        "materials code",
        "material code",
    ]

    HEADER_ALIASES = {
        "trade_name": ["trade name", "trade code", "trade"],
        "material_code": ["material code", "code", "material"],
        "unit": ["unit"],
        "pack_qty": ["pack qty", "pack qty.", "qty", "quantity", "pack quantity"],
        "pack_cost": ["pack cost", "cost", "pack price", "price"],
        "market_rate": ["market rate", "rate", "unit rate"],
        "material_variety": ["variety", "material variety"],
        "market_spec": [
            "spec",
            "market spec",
            "market spec / strength",
            "specification",
        ],
    }

    def __init__(self, file_path, project=None, company=None):
        self.file_path = file_path
        self.project = project
        self.company = company

    def _resolve_columns(self, header_row):
        """Return {field_name: column_index} from a header row."""
        normalized = [(_safe_str(c) or "").strip().lower() for c in header_row]
        col_map = {}
        for field, aliases in self.HEADER_ALIASES.items():
            for alias in aliases:
                if alias in normalized:
                    col_map[field] = normalized.index(alias)
                    break
        return col_map

    @staticmethod
    def _cell(row, idx):
        if idx is None:
            return None
        return row[idx] if idx < len(row) else None

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = _find_sheet(wb, self.SHEET_KEYWORDS)
        created = updated = 0

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return {"created": 0, "updated": 0}

        col_map = self._resolve_columns(rows[0])
        # Fall back to fixed positions for older templates that don't have
        # recognisable headers (Trade Name | Material Code | Unit | Market Rate
        # | Variety | Spec).
        if not col_map.get("material_code"):
            col_map = {
                "trade_name": 0,
                "material_code": 1,
                "unit": 2,
                "market_rate": 3,
                "material_variety": 4,
                "market_spec": 5,
            }

        for row in rows[1:]:
            if not row:
                continue
            mat_code = _safe_str(self._cell(row, col_map.get("material_code")))
            if not mat_code:
                continue

            trade_name = _safe_str(self._cell(row, col_map.get("trade_name")))
            unit = _safe_str(self._cell(row, col_map.get("unit")))
            variety = _safe_str(self._cell(row, col_map.get("material_variety")))
            spec = _safe_str(self._cell(row, col_map.get("market_spec")))

            pack_qty_raw = self._cell(row, col_map.get("pack_qty"))
            pack_cost_raw = self._cell(row, col_map.get("pack_cost"))
            market_rate_raw = self._cell(row, col_map.get("market_rate"))

            if pack_qty_raw is not None or pack_cost_raw is not None:
                pack_qty = _safe_decimal(pack_qty_raw) or Decimal("1")
                pack_cost = _safe_decimal(pack_cost_raw) or Decimal("0")
            else:
                # Legacy template: a single Market Rate column.
                pack_qty = Decimal("1")
                pack_cost = _safe_decimal(market_rate_raw) or Decimal("0")

            if pack_qty <= 0:
                pack_qty = Decimal("1")

            defaults = {
                "trade_name": trade_name,
                "unit": unit,
                "pack_qty": pack_qty,
                "pack_cost": pack_cost,
                "material_variety": variety,
                "market_spec": spec,
            }

            if self.project:
                _, was_created = ProjectMaterial.objects.update_or_create(
                    project=self.project,
                    material_code=mat_code,
                    defaults=defaults,
                )
            elif self.company:
                _, was_created = ContractorMaterial.objects.update_or_create(
                    company=self.company,
                    material_code=mat_code,
                    defaults=defaults,
                )
            else:
                _, was_created = SystemMaterial.objects.update_or_create(
                    material_code=mat_code,
                    defaults=defaults,
                )
            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        return {"created": created, "updated": updated}


# ── Labour Costs ─────────────────────────────────────────────────


class LabourCostImporter:
    """Import Labour Crew costs from Excel.

    Supports TWO layouts:
    - Offset layout (from Complete_Upload): data starts row 3, cols offset by 1
    - Simple layout (from template): data starts row 2, no offset

    Auto-detects by checking if row 2 col B contains "Crew Type".
    """

    SHEET_KEYWORDS = [
        "labour cost",
        "labour costs",
        "labourcost",
        "labourcosts",
        "labor cost",
        "labor costs",
        "crew cost",
        "crew costs",
        "crew",
    ]

    def __init__(self, file_path, project=None, company=None):
        self.file_path = file_path
        self.project = project
        self.company = company

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = _find_sheet(wb, self.SHEET_KEYWORDS)
        created = updated = 0

        row2_b = _safe_str(ws.cell(row=2, column=2).value).lower()
        is_offset = "crew type" in row2_b

        data_start = 3 if is_offset else 2
        col_offset = 1 if is_offset else 0

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            ncols = len(row) if row else 0
            c = col_offset
            crew_type = _safe_str(row[0 + c]) if ncols > c else ""
            if not crew_type:
                continue

            rate_base = 8 if is_offset else 5
            defaults = {
                "crew_size": int(row[1 + c] or 0) if ncols > 1 + c else 0,
                "skilled": int(row[2 + c] or 0) if ncols > 2 + c else 0,
                "semi_skilled": int(row[3 + c] or 0) if ncols > 3 + c else 0,
                "general": int(row[4 + c] or 0) if ncols > 4 + c else 0,
                "skilled_rate": _safe_decimal(row[rate_base]) or Decimal("0")
                if ncols > rate_base
                else Decimal("0"),
                "semi_skilled_rate": _safe_decimal(row[rate_base + 1]) or Decimal("0")
                if ncols > rate_base + 1
                else Decimal("0"),
                "general_rate": _safe_decimal(row[rate_base + 2]) or Decimal("0")
                if ncols > rate_base + 2
                else Decimal("0"),
            }

            if self.project:
                _, was_created = ProjectLabourCrew.objects.update_or_create(
                    project=self.project,
                    crew_type=crew_type,
                    defaults=defaults,
                )
            elif self.company:
                _, was_created = ContractorLabourCrew.objects.update_or_create(
                    company=self.company,
                    crew_type=crew_type,
                    defaults=defaults,
                )
            else:
                _, was_created = SystemLabourCrew.objects.update_or_create(
                    crew_type=crew_type,
                    defaults=defaults,
                )
            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        return {"created": created, "updated": updated}


# ── Material Specifications ──────────────────────────────────────


class MaterialSpecImporter:
    """Import Material Specifications from Excel.

    Supports wide format and multi-row format. Auto-detects.
    """

    SHEET_KEYWORDS = [
        "materials specification",
        "materials specifications",
        "materials spec",
        "materials specs",
        "material specification",
        "material specifications",
        "material spec",
        "material specs",
        "materialspec",
        "materialspecs",
        "mat spec",
        "mat specs",
        "specification code",
        "spec code",
    ]

    def __init__(self, file_path, project=None, company=None):
        self.file_path = file_path
        self.project = project
        self.company = company

    def _get_trade_code(self, trade_str):
        """Resolve (find-or-create) a trade code from a free-text value.

        The master sheet stores the full ``PREFIX-Trade Name`` string in the
        Trade Code column, so use the normalised resolver (matches on prefix,
        name, or prefix+name) rather than an exact prefix lookup.
        """
        return resolve_trade_code(trade_str, project=self.project, company=self.company)

    def _get_material(self, mat_code):
        """Resolve material by code, using project / contractor / system models."""
        if not mat_code:
            return None
        if self.project:
            return ProjectMaterial.objects.filter(
                project=self.project, material_code=mat_code
            ).first()
        if self.company:
            return ContractorMaterial.objects.filter(
                company=self.company, material_code=mat_code
            ).first()
        return SystemMaterial.objects.filter(material_code=mat_code).first()

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws, sheet_name, fell_back = _find_sheet_with_name(wb, self.SHEET_KEYWORDS)
        if fell_back:
            # Never silently import an arbitrary sheet — that produced the
            # "all over the place" results when the Materials Specification
            # sheet name didn't match.
            raise ValueError(
                "Could not find a 'Materials Specification' sheet in the "
                f"workbook (sheets: {', '.join(wb.sheetnames)}). Rename the "
                "sheet to 'Materials Specification' and re-upload."
            )

        row3_vals = [c.value for c in ws[3]] if ws.max_row >= 3 else []
        is_wide = any("specification code" in _safe_str(v).lower() for v in row3_vals)

        result = self._import_wide(ws) if is_wide else self._import_multirow(ws)
        result["sheet_used"] = sheet_name
        return result

    def _import_wide(self, ws):
        created = updated = 0

        for row in ws.iter_rows(min_row=4, values_only=True):
            ncols = len(row) if row else 0
            if ncols < 4:
                continue

            section = _safe_str(row[0])
            trade_code_str = _safe_str(row[1])
            unit = _safe_str(row[2]) or "m3"
            spec_name = _safe_str(row[3])

            if not spec_name:
                continue

            trade_code = self._get_trade_code(trade_code_str)

            if self.project:
                spec, was_created = ProjectSpecification.objects.update_or_create(
                    project=self.project,
                    name=spec_name,
                    defaults={
                        "section": section,
                        "trade_code": trade_code,
                        "unit_label": unit,
                    },
                )
                if not was_created:
                    spec.spec_components.all().delete()
            elif self.company:
                spec, was_created = ContractorSpecification.objects.update_or_create(
                    company=self.company,
                    name=spec_name,
                    defaults={
                        "section": section,
                        "trade_code": trade_code,
                        "unit_label": unit,
                    },
                )
                if not was_created:
                    spec.spec_components.all().delete()
            else:
                spec, was_created = SystemSpecification.objects.update_or_create(
                    name=spec_name,
                    defaults={
                        "section": section,
                        "trade_code": trade_code,
                        "unit_label": unit,
                    },
                )
                if not was_created:
                    spec.spec_components.all().delete()

            if was_created:
                created += 1
            else:
                updated += 1

            if self.project:
                component_model = ProjectSpecificationComponent
            elif self.company:
                component_model = ContractorSpecificationComponent
            else:
                component_model = SystemSpecificationComponent
            for i in range(4):
                mat_col = 4 + i
                qty_col = 8 + i
                mat_code = _safe_str(row[mat_col]) if ncols > mat_col else ""
                qty = _safe_decimal(row[qty_col]) if ncols > qty_col else None

                if not mat_code:
                    continue

                mat = self._get_material(mat_code)
                component_model.objects.create(
                    specification=spec,
                    material=mat,
                    label=mat_code,
                    qty_per_unit=qty or Decimal("0"),
                    sort_order=i,
                )

        return {"created": created, "updated": updated}

    def _import_multirow(self, ws):
        specs_data: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ncols = len(row) if row else 0
            spec_name = _safe_str(row[0]) if ncols > 0 else ""
            if not spec_name:
                continue

            if spec_name not in specs_data:
                specs_data[spec_name] = {
                    "section": _safe_str(row[1]) if ncols > 1 else "",
                    "trade_code_prefix": _safe_str(row[2]) if ncols > 2 else "",
                    "unit": _safe_str(row[3]) if ncols > 3 else "m3",
                    "components": [],
                }

            mat_code_str = _safe_str(row[4]) if ncols > 4 else ""
            if mat_code_str:
                specs_data[spec_name]["components"].append(
                    {
                        "material_code": mat_code_str,
                        "label": (_safe_str(row[5]) if ncols > 5 else "")
                        or mat_code_str,
                        "qty_per_unit": (_safe_decimal(row[6]) if ncols > 6 else None)
                        or Decimal("0"),
                    }
                )

        created = updated = 0

        for name, data in specs_data.items():
            trade_code = self._get_trade_code(data["trade_code_prefix"])

            if self.project:
                spec, was_created = ProjectSpecification.objects.update_or_create(
                    project=self.project,
                    name=name,
                    defaults={
                        "section": data["section"],
                        "trade_code": trade_code,
                        "unit_label": data["unit"],
                    },
                )
                if not was_created:
                    spec.spec_components.all().delete()
            elif self.company:
                spec, was_created = ContractorSpecification.objects.update_or_create(
                    company=self.company,
                    name=name,
                    defaults={
                        "section": data["section"],
                        "trade_code": trade_code,
                        "unit_label": data["unit"],
                    },
                )
                if not was_created:
                    spec.spec_components.all().delete()
            else:
                spec, was_created = SystemSpecification.objects.update_or_create(
                    name=name,
                    defaults={
                        "section": data["section"],
                        "trade_code": trade_code,
                        "unit_label": data["unit"],
                    },
                )
                if not was_created:
                    spec.spec_components.all().delete()

            if was_created:
                created += 1
            else:
                updated += 1

            if self.project:
                component_model = ProjectSpecificationComponent
            elif self.company:
                component_model = ContractorSpecificationComponent
            else:
                component_model = SystemSpecificationComponent
            for i, comp in enumerate(data["components"]):
                mat = self._get_material(comp["material_code"])
                component_model.objects.create(
                    specification=spec,
                    material=mat,
                    label=comp["label"],
                    qty_per_unit=comp["qty_per_unit"],
                    sort_order=i,
                )

        return {"created": created, "updated": updated}


# ── Labour Specifications ────────────────────────────────────────


class LabourSpecImporter:
    """Import Labour Specifications from Excel.

    Supports two-header and simple layouts. Auto-detects.
    """

    SHEET_KEYWORDS = [
        "labour spec",
        "labour specs",
        "labour specification",
        "labour specifications",
        "labourspec",
        "labourspecs",
        "labor spec",
        "labor specs",
        "labor specification",
        "labor specifications",
    ]

    def __init__(self, file_path, project=None, company=None):
        self.file_path = file_path
        self.project = project
        self.company = company

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = _find_sheet(wb, self.SHEET_KEYWORDS)
        created = updated = 0

        row1_vals = (
            [_safe_str(c.value).lower() for c in ws[1]] if ws.max_row >= 1 else []
        )
        has_group_header = any("productivity" in v for v in row1_vals)
        data_start = 3 if has_group_header else 2

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            ncols = len(row) if row else 0
            name = _safe_str(row[2]) if ncols > 2 else ""
            if not name:
                continue

            trade_code = resolve_trade_code(
                _safe_str(row[1]) if ncols > 1 else "",
                project=self.project,
                company=self.company,
            )
            if trade_code is None:
                # Trade is compulsory — skip rows without one.
                continue

            crew_type_str = _safe_str(row[4]) if ncols > 4 else ""
            if self.project:
                crew = (
                    ProjectLabourCrew.objects.filter(
                        project=self.project, crew_type=crew_type_str
                    ).first()
                    if crew_type_str
                    else None
                )
            elif self.company:
                crew = (
                    ContractorLabourCrew.objects.filter(
                        company=self.company, crew_type=crew_type_str
                    ).first()
                    if crew_type_str
                    else None
                )
            else:
                crew = (
                    SystemLabourCrew.objects.filter(crew_type=crew_type_str).first()
                    if crew_type_str
                    else None
                )

            defaults = {
                "section": _safe_str(row[0]) if ncols > 0 else "",
                "trade_code": trade_code,
                "unit": _safe_str(row[3]) if ncols > 3 else "",
                "crew": crew,
                "daily_production": (_safe_decimal(row[5]) if ncols > 5 else None)
                or Decimal("0"),
                "team_mix": (_safe_decimal(row[6]) if ncols > 6 else None)
                or Decimal("1"),
                "site_factor": (_safe_decimal(row[7]) if ncols > 7 else None)
                or Decimal("1"),
                "tools_factor": (_safe_decimal(row[8]) if ncols > 8 else None)
                or Decimal("1"),
                "leadership_factor": (_safe_decimal(row[9]) if ncols > 9 else None)
                or Decimal("1"),
            }

            if self.project:
                _, was_created = ProjectLabourSpecification.objects.update_or_create(
                    project=self.project,
                    name=name,
                    defaults=defaults,
                )
            elif self.company:
                _, was_created = ContractorLabourSpecification.objects.update_or_create(
                    company=self.company,
                    name=name,
                    defaults=defaults,
                )
            else:
                _, was_created = SystemLabourSpecification.objects.update_or_create(
                    name=name,
                    defaults=defaults,
                )
            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        return {"created": created, "updated": updated}


# ── Plant Costs ─────────────────────────────────────────────────


class PlantCostImporter:
    """Import Plant & Equipment costs from Excel.

    Expected columns: Plant & Equipment | Hourly Production | Hourly Rate
    """

    SHEET_KEYWORDS = [
        "plant cost",
        "plant costs",
        "plantcost",
        "plantcosts",
        "plant and equipment",
        "plant & equipment",
    ]

    def __init__(self, file_path, project=None, company=None):
        self.file_path = file_path
        self.project = project
        self.company = company

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws, sheet_name, fell_back = _find_sheet_with_name(wb, self.SHEET_KEYWORDS)
        all_sheets = list(wb.sheetnames)
        created = updated = 0

        co = _col_offset(ws)
        header_row = _find_header_row(ws, ["plant", "hourly"])
        data_start = header_row + 1

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            ncols = len(row) if row else 0
            name = _safe_str(row[co]) if ncols > co else ""
            if not name or name.lower() == "total":
                continue

            defaults = {
                "hourly_production": (
                    _safe_decimal(row[co + 1]) if ncols > co + 1 else None
                )
                or Decimal("0"),
                "hourly_rate": (_safe_decimal(row[co + 2]) if ncols > co + 2 else None)
                or Decimal("0"),
            }

            if self.project:
                _, was_created = ProjectPlantCost.objects.update_or_create(
                    project=self.project,
                    name=name,
                    defaults=defaults,
                )
            elif self.company:
                _, was_created = ContractorPlantCost.objects.update_or_create(
                    company=self.company,
                    name=name,
                    defaults=defaults,
                )
            else:
                _, was_created = SystemPlantCost.objects.update_or_create(
                    name=name,
                    defaults=defaults,
                )
            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        return {
            "created": created,
            "updated": updated,
            "sheet_used": sheet_name,
            "fell_back": fell_back,
            "all_sheets": all_sheets,
        }


# ── Plant Specifications ────────────────────────────────────────


class PlantSpecImporter:
    """Import Plant Specifications from Excel.

    Master layout (one row = one spec with up to 4 plant components):
        Section | Trade Name | Specification | Unit
        | Plant Type 1 | Hours 1 | Plant Type 2 | Hours 2
        | Plant 3 | Hours 3 | Plant 4 | Hours 4
        | Daily Production | Operator | Site | (Daily Output/Daily Cost/Rate/BoQ/Total — ignored)

    Template layout (one row = one spec with a single plant component):
        Section | Trade Name | Plant Specification | Unit | Plant Type
        | Daily Production | Operator | Site

    The two layouts are distinguished by checking whether the header row
    contains more than one "plant" column.
    """

    SHEET_KEYWORDS = [
        "plant spec",
        "plant specs",
        "plant specification",
        "plant specifications",
        "plantspec",
        "plantspecs",
    ]

    def __init__(self, file_path, project=None, company=None):
        self.file_path = file_path
        self.project = project
        self.company = company

    def _lookup_plant_cost(self, name):
        if not name:
            return None
        if self.project:
            return ProjectPlantCost.objects.filter(
                project=self.project, name=name
            ).first()
        if self.company:
            return ContractorPlantCost.objects.filter(
                company=self.company, name=name
            ).first()
        return SystemPlantCost.objects.filter(name=name).first()

    def _header_is_master(self, ws, header_row, co):
        """Return True if the header row has multiple plant columns (master layout)."""
        plant_header_count = 0
        for c in range(co, (ws.max_column or 1)):
            v = ws.cell(row=header_row, column=c + 1).value
            if v and "plant" in str(v).lower():
                plant_header_count += 1
        return plant_header_count > 1

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws, sheet_name, fell_back = _find_sheet_with_name(wb, self.SHEET_KEYWORDS)
        all_sheets = list(wb.sheetnames)
        created = updated = 0

        co = _col_offset(ws)
        header_row = _find_header_row(ws, ["section", "plant"])
        data_start = header_row + 1
        is_master = self._header_is_master(ws, header_row, co)

        if self.project:
            spec_model = ProjectPlantSpecification
            comp_model = ProjectPlantSpecificationComponent
        elif self.company:
            spec_model = ContractorPlantSpecification
            comp_model = ContractorPlantSpecificationComponent
        else:
            spec_model = SystemPlantSpecification
            comp_model = SystemPlantSpecificationComponent

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            ncols = len(row) if row else 0
            name = _safe_str(row[co + 2]) if ncols > co + 2 else ""
            if not name:
                continue

            trade_code = resolve_trade_code(
                _safe_str(row[co + 1]) if ncols > co + 1 else "",
                project=self.project,
                company=self.company,
            )
            if trade_code is None:
                continue

            if is_master:
                # Master layout: 4 plant-type/hours pairs at cols 4..11,
                # productivity at cols 12..14.
                component_specs = []
                for i in range(4):
                    pt_col = co + 4 + (i * 2)
                    hr_col = pt_col + 1
                    pt_name = _safe_str(row[pt_col]) if ncols > pt_col else ""
                    if not pt_name:
                        continue
                    hours = _safe_decimal(row[hr_col]) if ncols > hr_col else None
                    component_specs.append((pt_name, hours or Decimal("0")))
                daily_prod = _safe_decimal(row[co + 12]) if ncols > co + 12 else None
                operator = _safe_decimal(row[co + 13]) if ncols > co + 13 else None
                site = _safe_decimal(row[co + 14]) if ncols > co + 14 else None
            else:
                # Template layout: single plant type at col 4, productivity at 5..7.
                pt_name = _safe_str(row[co + 4]) if ncols > co + 4 else ""
                component_specs = [(pt_name, Decimal("1"))] if pt_name else []
                daily_prod = _safe_decimal(row[co + 5]) if ncols > co + 5 else None
                operator = _safe_decimal(row[co + 6]) if ncols > co + 6 else None
                site = _safe_decimal(row[co + 7]) if ncols > co + 7 else None

            defaults = {
                "section": _safe_str(row[co + 0]) if ncols > co + 0 else "",
                "trade_code": trade_code,
                "unit": _safe_str(row[co + 3]) if ncols > co + 3 else "",
                "daily_production": daily_prod or Decimal("0"),
                "operator_factor": operator or Decimal("1"),
                "site_factor": site or Decimal("1"),
            }

            if self.project:
                spec, was_created = spec_model.objects.update_or_create(
                    project=self.project,
                    name=name,
                    defaults=defaults,
                )
            elif self.company:
                spec, was_created = spec_model.objects.update_or_create(
                    company=self.company,
                    name=name,
                    defaults=defaults,
                )
            else:
                spec, was_created = spec_model.objects.update_or_create(
                    name=name,
                    defaults=defaults,
                )

            # Rebuild components from scratch — avoids stale entries on re-import.
            spec.components.all().delete()
            for i, (pt_name, hours) in enumerate(component_specs):
                comp_model.objects.create(
                    specification=spec,
                    plant_type=self._lookup_plant_cost(pt_name),
                    hours=hours,
                    sort_order=i,
                )

            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        return {
            "created": created,
            "updated": updated,
            "sheet_used": sheet_name,
            "fell_back": fell_back,
            "all_sheets": all_sheets,
        }


# ── Preliminary Costs ───────────────────────────────────────────


class PreliminaryCostImporter:
    """Import Preliminary Costs from Excel.

    Supports two layouts:

    1. Flat template: one row per item with a `Preliminary Type` column
       (as produced by the downloaded template).
    2. Hierarchical master layout: section-header rows (e.g. "Fixed
       Contractual Requirements", "Time - Site Personnel") define the
       preliminary_type for all data rows beneath them, until the next
       header.
    """

    SHEET_KEYWORDS = [
        "preliminary cost",
        "preliminary costs",
        "preliminarycost",
        "preliminarycosts",
        "preliminaries cost",
        "preliminaries costs",
        "prelim cost",
        "prelim costs",
    ]

    # Maps a matched header phrase (lowercased, normalised) to the type code.
    # Keys are checked as substrings against the first cell of each row.
    TYPE_PHRASE_MAP = [
        ("fixed contractual", "fixed_contractual"),
        ("fixed - contractual", "fixed_contractual"),
        ("fixed-contractual", "fixed_contractual"),
        ("fixed facilities", "fixed_facilities"),
        ("fixed - facilities", "fixed_facilities"),
        ("fixed-facilities", "fixed_facilities"),
        ("time contractual", "time_contractual"),
        ("time - contractual", "time_contractual"),
        ("time-contractual", "time_contractual"),
        ("time small tool", "time_small_tools"),
        ("time - small tool", "time_small_tools"),
        ("time-small tool", "time_small_tools"),
        ("time plant", "time_plant_equipment"),
        ("time - plant", "time_plant_equipment"),
        ("time-plant", "time_plant_equipment"),
        ("time company", "time_company_overheads"),
        ("time - company", "time_company_overheads"),
        ("time-company", "time_company_overheads"),
        ("time site personnel", "time_site_personnel"),
        ("time - site personnel", "time_site_personnel"),
        ("time-site personnel", "time_site_personnel"),
        # keep these after the more-specific time- entries
        ("time facilities", "time_facilities"),
        ("time - facilities", "time_facilities"),
        ("time-facilities", "time_facilities"),
    ]

    VALID_TYPE_CODES = {
        code for code, _ in SystemPreliminaryCost.PRELIMINARY_TYPE_CHOICES
    }

    SKIP_PHRASES = {
        "fixed preliminaries",
        "time related preliminaries",
        "total",
    }

    # Header labels recognised in the master-workbook layout. Each maps a
    # lowercased cell value to the model field it supplies.
    COLUMN_LABELS = {
        "items": "name_col",
        "name": "name_col",
        "sum": "sum_value",
        "number/month": "number_per_month",
        "number / month": "number_per_month",
        "monthly rate": "monthly_rate",
        "months": "months",
        "amount": "amount",
    }

    # Fallback name column when a section header doesn't label it (the time
    # sections omit a "Name" header — col 0 is trade code, col 1 is blank or
    # the type phrase, col 2 is the item name).
    DEFAULT_NAME_COL = 2

    def __init__(self, file_path, project=None, company=None):
        self.file_path = file_path
        self.project = project
        self.company = company

    def _match_type(self, phrase):
        p = phrase.lower().strip()
        for needle, code in self.TYPE_PHRASE_MAP:
            if needle in p:
                return code
        return None

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws, sheet_name, fell_back = _find_sheet_with_name(wb, self.SHEET_KEYWORDS)
        all_sheets = list(wb.sheetnames)

        co = _col_offset(ws)

        # Detect layout: flat template has a header row mentioning both
        # "preliminary type" (or just "type") and "name".
        flat_header_row = None
        for r in range(1, min((ws.max_row or 1) + 1, 8)):
            parts = [
                _safe_str(ws.cell(row=r, column=c).value).lower()
                for c in range(1, (ws.max_column or 1) + 1)
            ]
            joined = " | ".join(parts)
            if ("preliminary type" in joined or "type" in joined) and "name" in joined:
                flat_header_row = r
                break

        if flat_header_row is not None:
            return self._import_flat(
                ws, flat_header_row, co, sheet_name, fell_back, all_sheets
            )
        return self._import_by_labels(ws, co, sheet_name, fell_back, all_sheets)

    def _upsert(self, name, ptype, defaults):
        if self.project:
            _, was_created = ProjectPreliminaryCost.objects.update_or_create(
                project=self.project,
                name=name,
                preliminary_type=ptype,
                defaults=defaults,
            )
        elif self.company:
            _, was_created = ContractorPreliminaryCost.objects.update_or_create(
                company=self.company,
                name=name,
                preliminary_type=ptype,
                defaults=defaults,
            )
        else:
            _, was_created = SystemPreliminaryCost.objects.update_or_create(
                name=name,
                preliminary_type=ptype,
                defaults=defaults,
            )
        return was_created

    def _import_flat(self, ws, header_row, co, sheet_name, fell_back, all_sheets):
        created = updated = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            ncols = len(row) if row else 0
            raw_type = _safe_str(row[co + 0]) if ncols > co + 0 else ""
            name = _safe_str(row[co + 1]) if ncols > co + 1 else ""
            if not name:
                continue

            # Accept either a code (fixed_contractual) or a display phrase.
            if raw_type in self.VALID_TYPE_CODES:
                ptype = raw_type
            else:
                ptype = self._match_type(raw_type) or ""
            if not ptype:
                continue

            defaults = {
                "sum_value": (_safe_decimal(row[co + 2]) if ncols > co + 2 else None)
                or Decimal("0"),
                "amount": (_safe_decimal(row[co + 3]) if ncols > co + 3 else None)
                or Decimal("0"),
                "number_per_month": (
                    _safe_decimal(row[co + 4]) if ncols > co + 4 else None
                )
                or Decimal("0"),
                "monthly_rate": (_safe_decimal(row[co + 5]) if ncols > co + 5 else None)
                or Decimal("0"),
                "months": (_safe_decimal(row[co + 6]) if ncols > co + 6 else None)
                or Decimal("0"),
            }
            if self._upsert(name, ptype, defaults):
                created += 1
            else:
                updated += 1

        return {
            "created": created,
            "updated": updated,
            "sheet_used": sheet_name,
            "fell_back": fell_back,
            "all_sheets": all_sheets,
        }

    def _detect_column_map(self, row):
        """Scan a row for known header labels and return {field: col_index}.

        Matching is by exact (case-insensitive, whitespace-stripped) cell value
        so unrelated text like "Trade Name" doesn't claim the "Name" slot.
        When a label appears more than once (the fixed section has two
        "Amount" columns) the left-most occurrence wins.
        """
        col_map = {}
        for ci, cell in enumerate(row or ()):
            if cell is None:
                continue
            s = str(cell).strip().lower()
            if not s:
                continue
            field = self.COLUMN_LABELS.get(s)
            if field is not None:
                col_map.setdefault(field, ci)
        return col_map

    def _scan_type_match(self, row):
        """Return the first type-phrase match found anywhere in the row."""
        for cell in row or ():
            if cell is None:
                continue
            m = self._match_type(str(cell))
            if m:
                return m
        return None

    def _import_by_labels(self, ws, co, sheet_name, fell_back, all_sheets):
        """Import the master-workbook layout.

        The trade-code column is ignored. The type phrase may appear in any
        column (repeated per row in the fixed section, and once as a header in
        each time section). Numeric columns are resolved by header label
        ('Sum', 'Amount', 'Number/Month', 'Monthly Rate', 'Months'), so the
        parser is tolerant of shifted columns and duplicate 'Amount' cells.
        """
        created = updated = 0
        current_type = None
        col_map = {}

        for row in ws.iter_rows(min_row=1, values_only=True):
            type_match = self._scan_type_match(row)
            new_col_map = self._detect_column_map(row)

            if new_col_map:
                col_map = new_col_map
                if type_match:
                    current_type = type_match
                continue  # header row — never a data row

            if type_match:
                current_type = type_match

            if not current_type or not col_map:
                continue

            name_col = col_map.get("name_col", self.DEFAULT_NAME_COL)
            name = _safe_str(row[name_col]) if len(row) > name_col else ""
            if not name:
                continue

            name_lower = name.lower().strip()
            # Skip rows where the "name" cell is actually a section header
            # (e.g. the time sections list the type phrase at the name column).
            if self._match_type(name):
                continue
            if any(phrase in name_lower for phrase in self.SKIP_PHRASES):
                continue

            is_time = current_type.startswith("time_")
            defaults = {
                "sum_value": Decimal("0"),
                "amount": Decimal("0"),
                "number_per_month": Decimal("0"),
                "monthly_rate": Decimal("0"),
                "months": Decimal("0"),
            }
            for field in defaults:
                idx = col_map.get(field)
                if idx is None or idx >= len(row):
                    continue
                val = _safe_decimal(row[idx])
                if val is not None:
                    defaults[field] = val

            # Drop rows with no numeric payload — prevents the label/subtotal
            # rows that slip through from creating zero-value records.
            if is_time:
                if (
                    defaults["number_per_month"] == 0
                    and defaults["monthly_rate"] == 0
                    and defaults["months"] == 0
                    and defaults["amount"] == 0
                ):
                    continue
            else:
                if defaults["sum_value"] == 0 and defaults["amount"] == 0:
                    continue

            if self._upsert(name, current_type, defaults):
                created += 1
            else:
                updated += 1

        return {
            "created": created,
            "updated": updated,
            "sheet_used": sheet_name,
            "fell_back": fell_back,
            "all_sheets": all_sheets,
        }


# ── Preliminary Specifications ──────────────────────────────────


class PreliminarySpecImporter:
    """Import Preliminary Specifications from Excel.

    Expected columns: Section | Trade Name | Name | Unit | Preliminary Type

    The Preliminary Type column is matched against the choices on
    ``SystemPreliminaryCost.PRELIMINARY_TYPE_CHOICES`` (either the value or the
    display label, case-insensitive). Spec amount is now derived from
    preliminary costs of the matching type, so it is no longer imported.
    """

    _NAME_TO_TYPE = {
        "fixed-contractual requirements": "fixed_contractual",
        "fixed contractual requirements": "fixed_contractual",
        "fixed-facilities required": "fixed_facilities",
        "fixed facilities": "fixed_facilities",
        "time-contractual requirements": "time_contractual",
        "time contractual requirements": "time_contractual",
        "time-facilities for contractor": "time_facilities",
        "time-facilities": "time_facilities",
        "time-small tools & accessories": "time_small_tools",
        "time-small tool allowances": "time_small_tools",
        "time-plant and equipment": "time_plant_equipment",
        "time-company and head office overheads": "time_company_overheads",
        "time-company & head office overheads": "time_company_overheads",
        "time-site personnel": "time_site_personnel",
    }

    @classmethod
    def _resolve_type(cls, raw, name):
        """Resolve a preliminary_type from the cell value, falling back to name."""
        from .models import SystemPreliminaryCost

        valid_values = {v for v, _ in SystemPreliminaryCost.PRELIMINARY_TYPE_CHOICES}
        valid_labels = {
            label.strip().lower(): v
            for v, label in SystemPreliminaryCost.PRELIMINARY_TYPE_CHOICES
        }
        candidate = (_safe_str(raw) or "").strip()
        if candidate:
            if candidate in valid_values:
                return candidate
            mapped = valid_labels.get(candidate.lower())
            if mapped:
                return mapped
        return cls._NAME_TO_TYPE.get((name or "").strip().lower(), "")

    SHEET_KEYWORDS = [
        "preliminary spec",
        "preliminary specs",
        "preliminary specification",
        "preliminary specifications",
        "preliminariesspec",
        "preliminaries spec",
        "prelim spec",
        "prelim specs",
    ]

    def __init__(self, file_path, project=None, company=None):
        self.file_path = file_path
        self.project = project
        self.company = company

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws, sheet_name, fell_back = _find_sheet_with_name(wb, self.SHEET_KEYWORDS)
        all_sheets = list(wb.sheetnames)
        created = updated = 0

        co = _col_offset(ws)
        # "Unit" is present in both the downloaded template ("Section | Trade
        # Name | Name | Unit | Amount") and the master workbook ("Element |
        # Trade Name | Preliminaries | Unit | Amount"), and never appears as
        # a cell value in data rows — so it uniquely anchors the header row.
        header_row = _find_header_row(ws, ["unit"])
        data_start = header_row + 1

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            ncols = len(row) if row else 0
            name = _safe_str(row[co + 2]) if ncols > co + 2 else ""
            if not name:
                continue

            trade_code = resolve_trade_code(
                _safe_str(row[co + 1]) if ncols > co + 1 else "",
                project=self.project,
                company=self.company,
            )
            if trade_code is None:
                continue

            raw_type = row[co + 4] if ncols > co + 4 else None
            defaults = {
                "section": _safe_str(row[co + 0]) if ncols > co + 0 else "",
                "trade_code": trade_code,
                "unit": _safe_str(row[co + 3]) if ncols > co + 3 else "",
                "preliminary_type": self._resolve_type(raw_type, name),
            }

            if self.project:
                _, was_created = (
                    ProjectPreliminarySpecification.objects.update_or_create(
                        project=self.project,
                        name=name,
                        defaults=defaults,
                    )
                )
            elif self.company:
                _, was_created = (
                    ContractorPreliminarySpecification.objects.update_or_create(
                        company=self.company,
                        name=name,
                        defaults=defaults,
                    )
                )
            else:
                _, was_created = (
                    SystemPreliminarySpecification.objects.update_or_create(
                        name=name,
                        defaults=defaults,
                    )
                )
            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        return {
            "created": created,
            "updated": updated,
            "sheet_used": sheet_name,
            "fell_back": fell_back,
            "all_sheets": all_sheets,
        }


# ── Item Library ─────────────────────────────────────────────────


class ItemLibraryImporter:
    """Import Item Library entries from Excel.

    Expected columns:
        Trade Code | Accounts Code | Component | Material Spec |
        Labour & Plant Spec | Preliminaries | Item Description | Unit |
        Item Code

    The "Labour & Plant Spec" column resolves against both labour and
    plant specification tables — whichever names match are linked.
    Item Code is optional, free-text.
    """

    SHEET_KEYWORDS = [
        "item library",
        "items library",
        "item lib",
        "itemlibrary",
        "library",
    ]

    def __init__(self, file_path, project=None, company=None):
        self.file_path = file_path
        self.project = project
        self.company = company

    def _get_trade_code(self, value):
        if not value:
            return None
        if self.project:
            qs = ProjectTradeCode.objects.filter(project=self.project)
        elif self.company:
            qs = ContractorTradeCode.objects.filter(company=self.company)
        else:
            qs = SystemTradeCode.objects.all()
        # Try exact prefix match first, then full "prefix+trade_name" match
        tc = qs.filter(prefix=value).first()
        if tc:
            return tc
        for candidate in qs:
            if candidate.trade_code == value:
                return candidate
        return None

    def _get_material_spec(self, name):
        if not name:
            return None
        if self.project:
            return ProjectSpecification.objects.filter(
                project=self.project, name=name
            ).first()
        if self.company:
            return ContractorSpecification.objects.filter(
                company=self.company, name=name
            ).first()
        return SystemSpecification.objects.filter(name=name).first()

    def _get_labour_spec(self, name):
        if not name:
            return None
        if self.project:
            return ProjectLabourSpecification.objects.filter(
                project=self.project, name=name
            ).first()
        if self.company:
            return ContractorLabourSpecification.objects.filter(
                company=self.company, name=name
            ).first()
        return SystemLabourSpecification.objects.filter(name=name).first()

    def _get_plant_spec(self, name):
        if not name:
            return None
        if self.project:
            return ProjectPlantSpecification.objects.filter(
                project=self.project, name=name
            ).first()
        if self.company:
            return ContractorPlantSpecification.objects.filter(
                company=self.company, name=name
            ).first()
        return SystemPlantSpecification.objects.filter(name=name).first()

    def _get_prelim_spec(self, name):
        if not name:
            return None
        if self.project:
            return ProjectPreliminarySpecification.objects.filter(
                project=self.project, name=name
            ).first()
        if self.company:
            return ContractorPreliminarySpecification.objects.filter(
                company=self.company, name=name
            ).first()
        return SystemPreliminarySpecification.objects.filter(name=name).first()

    def run(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws, sheet_name, fell_back = _find_sheet_with_name(wb, self.SHEET_KEYWORDS)
        all_sheets = list(wb.sheetnames)

        created = updated = skipped = 0
        warnings = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            ncols = len(row) if row else 0
            trade_code_str = _safe_str(row[0]) if ncols > 0 else ""
            accounts_code = _safe_str(row[1]) if ncols > 1 else ""
            component = _safe_str(row[2]) if ncols > 2 else ""
            material_spec_name = _safe_str(row[3]) if ncols > 3 else ""
            labour_plant_name = _safe_str(row[4]) if ncols > 4 else ""
            prelim_spec_name = _safe_str(row[5]) if ncols > 5 else ""
            description = _safe_str(row[6]) if ncols > 6 else ""
            unit = _safe_str(row[7]) if ncols > 7 else ""
            item_code = _safe_str(row[8]) if ncols > 8 else ""

            if not description:
                skipped += 1
                continue

            # Drop placeholder text from the template sheet ("Dropdown from
            # specification") — those are header hints, not real names.
            if material_spec_name.lower().startswith("dropdown"):
                material_spec_name = ""
            if labour_plant_name.lower().startswith("dropdown"):
                labour_plant_name = ""
            if prelim_spec_name.lower().startswith("dropdown"):
                prelim_spec_name = ""

            trade_code = self._get_trade_code(trade_code_str)
            material_spec = self._get_material_spec(material_spec_name)
            labour_spec = self._get_labour_spec(labour_plant_name)
            plant_spec = self._get_plant_spec(labour_plant_name)
            prelim_spec = self._get_prelim_spec(prelim_spec_name)

            if material_spec_name and not material_spec:
                warnings.append(
                    f'Row {idx}: material spec "{material_spec_name}" not found'
                )
            if labour_plant_name and not (labour_spec or plant_spec):
                warnings.append(
                    f'Row {idx}: labour/plant spec "{labour_plant_name}" not found'
                )
            if prelim_spec_name and not prelim_spec:
                warnings.append(
                    f'Row {idx}: preliminary spec "{prelim_spec_name}" not found'
                )

            defaults = {
                "trade_code": trade_code,
                "item_code": item_code,
                "accounts_code": accounts_code,
                "component": component,
                "unit": unit,
                "material_spec": material_spec,
                "labour_spec": labour_spec,
                "plant_spec": plant_spec,
                "preliminary_spec": prelim_spec,
                # Preserve the raw names so unmatched values stay visible and
                # can be flagged as unlinked in the UI. When a spec matched,
                # store its canonical name so the value reads as linked.
                "material_spec_name": (
                    material_spec.name if material_spec else material_spec_name
                ),
                "labour_plant_spec_name": labour_plant_name,
                "preliminary_spec_name": (
                    prelim_spec.name if prelim_spec else prelim_spec_name
                ),
                "display_order": idx,
            }

            if self.project:
                _, was_created = ProjectItemLibraryEntry.objects.update_or_create(
                    project=self.project,
                    description=description,
                    component=component,
                    defaults=defaults,
                )
            elif self.company:
                _, was_created = ContractorItemLibraryEntry.objects.update_or_create(
                    company=self.company,
                    description=description,
                    component=component,
                    defaults=defaults,
                )
            else:
                _, was_created = SystemItemLibraryEntry.objects.update_or_create(
                    description=description,
                    component=component,
                    defaults=defaults,
                )
            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        return {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "warnings": warnings,
            "sheet_used": sheet_name,
            "fell_back": fell_back,
            "all_sheets": all_sheets,
        }
