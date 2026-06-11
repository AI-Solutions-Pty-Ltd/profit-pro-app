import openpyxl

wb = openpyxl.load_workbook("test_export_2.xlsx")
print("Sheetnames:", wb.sheetnames)
ws = wb.active
bill_cell = None
for row in ws.iter_rows(min_row=1, max_row=20):
    if (
        row[0].value
        and isinstance(row[0].value, str)
        and row[0].value.startswith("BILL NO.")
    ):
        bill_cell = row[0]
        break

if bill_cell:
    print("Bill Header Cell:", bill_cell.value)
    print("Bold:", bill_cell.font.bold)
    print(
        "Fill Color:", bill_cell.fill.start_color.rgb if bill_cell.fill else "No Fill"
    )
else:
    print("Could not find bill cell")
